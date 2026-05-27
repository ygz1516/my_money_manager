# -*- coding: utf-8 -*-
import sys
import os
import time
import json
from collections import defaultdict      # ← 新加这一行
from werkzeug.utils import secure_filename
import pandas as pd
from flask import send_from_directory, session
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash, abort
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from extensions import db, login_manager
from flask_migrate import Migrate
import re
from simpleeval import simple_eval
from flask import current_app
from utils import calculate_salary as utils_calculate_salary
from utils import send_email_with_attachment, export_salary_table
import logging
from sqlalchemy import text, inspect
from dateutil.relativedelta import relativedelta
from models import Unit, Employee, SalaryItem, SalaryRecord, User, SpecialSalaryItem, AssessmentOption, EmployeeSpecialItem, EmployeeAssessment, UserTask, SpecialItemTemplate, EmployeeSpecialGrant, AuxiliaryForm, BatchCalcTemplate, BatchCalcSnapshot
current_dir = os.path.dirname(os.path.abspath(__file__))
from sqlalchemy.orm.attributes import flag_modified
import shutil
import subprocess
import tempfile
from sqlalchemy.orm.attributes import flag_modified
import zipfile
import io
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
from sqlalchemy import func
from models import AuxiliaryForm, BatchCalcTemplate
from models import SpecialItemTemplate, EmployeeSpecialGrant, SalaryRecord, SalaryItem, Employee, Unit
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return decorated_function
import re
from simpleeval import simple_eval


def evaluate_formula(formula, variables):
    if not formula:
        return 0.0
    try:
        # 统一引号
        formula = formula.replace('“', '"').replace('”', '"').replace('‘', "'").replace('’', "'")
        formula = formula.replace('（', '(').replace('）', ')')
        formula = formula.lstrip()
        if formula.startswith('='):
            formula = formula[1:]
        formula = formula.replace('＝', '=')

        # 清除数字外面多余的括号
        import re
        formula = re.sub(r'\((\d+(?:\.\d+)?)\)+', r'\1', formula)

        # 变量替换
        for var in sorted(variables.keys(), key=len, reverse=True):
            val = variables[var]
            escaped_var = re.escape(var)
            if isinstance(val, str):
                cleaned = val.strip().strip('()').replace(',', '').replace('，', '').strip()
                if var == '身份证号':  # 身份证号保持字符串
                    replacement = f'"{cleaned}"'
                else:
                    try:
                        num = float(cleaned)
                        replacement = str(num)
                    except ValueError:
                        replacement = f'"{cleaned}"'
            else:
                replacement = str(val)
            pattern = r'(?<![a-zA-Z0-9\u4e00-\u9fa5%])' + escaped_var + r'(?![a-zA-Z0-9\u4e00-\u9fa5%])'
            formula = re.sub(pattern, replacement, formula)

        # 定义 IF 函数
        def if_func(condition, true_val, false_val):
            return true_val if condition else false_val

        # 注册常用数学函数
        result = simple_eval(formula, functions={
            'IF': if_func,
            'int': int,
            'float': float,
            'round': round,
            'max': max,
            'min': min,
            'abs': abs,
        })
        return round(float(result), 2)
    except Exception as e:
        current_app.logger.error(f"公式计算失败: {formula}, 错误: {e}")
        return 0.0
def topological_sort_calc_fields(calc_fields, known_vars):
    """
    对计算字段进行拓扑排序，返回计算顺序。
    若存在循环依赖，抛出 ValueError。
    """
    graph = {f['name']: set() for f in calc_fields}
    name_to_field = {f['name']: f for f in calc_fields}

    # 构建依赖图：解析每个计算公式中的变量名
    import re
    for f in calc_fields:
        formula = f.get('formula', '')
        # 简单提取所有标识符（字母、数字、下划线、中文）
        tokens = re.findall(r'[a-zA-Z_][a-zA-Z0-9_]*|[\u4e00-\u9fa5]+', formula)
        for token in tokens:
            if token in name_to_field and token != f['name']:
                graph[f['name']].add(token)

    # Kahn 算法
    in_degree = {name: 0 for name in graph}
    for name, deps in graph.items():
        for dep in deps:
            if dep in in_degree:
                in_degree[name] += 1

    queue = [name for name in graph if in_degree[name] == 0]
    result = []
    while queue:
        node = queue.pop(0)
        result.append(node)
        for other, deps in graph.items():
            if node in deps:
                in_degree[other] -= 1
                if in_degree[other] == 0:
                    queue.append(other)

    if len(result) != len(calc_fields):
        raise ValueError(f"自定义字段存在循环依赖，请检查公式。未解析字段: {set(graph.keys()) - set(result)}")

    # 按结果顺序返回字段对象
    sorted_fields = [name_to_field[name] for name in result]
    return sorted_fields


def evaluate_extra_fields(template, input_values, context_vars=None):
    if not template.extra_fields:
        return input_values.copy() if input_values else {}

    try:
        extra_fields = json.loads(template.extra_fields)
    except:
        extra_fields = []

    if not extra_fields:
        return input_values.copy() if input_values else {}

    # 初始化变量池
    variables = context_vars.copy() if context_vars else {}

    # ========== ★ 新增：对 input_values 进行类型转换 ==========
    converted_input = {}
    for f in extra_fields:
        name = f.get('name')
        if name in input_values:
            val = input_values[name]
            if f.get('type') == 'number':
                try:
                    if isinstance(val, str):
                        val = val.strip().strip('()').replace(',', '').replace('，', '').strip()
                    converted_input[name] = float(val) if val != '' else 0.0
                except:
                    converted_input[name] = 0.0
            elif f.get('type') == 'calculation':
                # 计算字段不应作为输入，但为安全起见，忽略或按数字尝试转换
                try:
                    converted_input[name] = float(val) if val != '' else 0.0
                except:
                    pass  # 忽略，后面会重新计算
            else:
                converted_input[name] = str(val) if val is not None else ''
    variables.update(converted_input)
    # =====================================================

    # 分离普通字段和计算字段
    normal_fields = [f for f in extra_fields if f.get('type') != 'calculation']
    calc_fields = [f for f in extra_fields if f.get('type') == 'calculation']

    # 确保普通字段有默认值（如果用户未提供）
    for f in normal_fields:
        name = f['name']
        if name not in variables:
            default_val = f.get('default')
            if f.get('type') == 'number':
                try:
                    variables[name] = float(default_val) if default_val not in (None, '') else 0.0
                except:
                    variables[name] = 0.0
            else:
                variables[name] = default_val if default_val is not None else ''

    # 如果没有计算字段，直接返回
    if not calc_fields:
        return {f['name']: variables.get(f['name']) for f in extra_fields}

    # 拓扑排序
    try:
        sorted_calcs = topological_sort_calc_fields(calc_fields, variables.keys())
    except ValueError as e:
        current_app.logger.error(f"自定义字段依赖排序失败: {e}")
        raise

    # 按顺序计算
    for field in sorted_calcs:
        formula = field.get('formula', '')
        if not formula:
            variables[field['name']] = 0.0
            continue
        try:
            emp_id = context_vars.get('employee_id') if context_vars else None
            year = context_vars.get('year') if context_vars else None
            month = context_vars.get('month') if context_vars else None  # ← 新增：提取 month
            salary_item_names = json.loads(template.salary_items) if template.salary_items else None

            value = evaluate_formula_with_context(
                formula,
                employee_id=emp_id,
                year=year,
                extra_vars=variables,
                salary_item_names=salary_item_names,
                month=month  # ← 新增：传递 month
            )
        except Exception as e:
            current_app.logger.warning(f"计算字段 {field['name']} 公式失败: {e}")
            value = 0.0

        variables[field['name']] = value

    # 整理最终结果
    result = {}
    for f in extra_fields:
        name = f['name']
        result[name] = variables.get(name)
    return result


def evaluate_formula_with_context(formula, employee_id, year, extra_vars=None, salary_item_names=None, month=None):
    """
    增强版公式求值，支持：
    - 上年合计（上年[工资项]合计、上年工资收入总额）
    - 本年合计（本年[工资项]合计、本年收入总额）
    - 指定年份/月份的工资项合计（例如：2026年5月基本工资）
    - 单月工资项取值（例如：2026年5月绩效工资）
    - special_模板名(参数) 引用其他特殊事项金额
    - 员工属性注入（姓名、岗位级别等）
    - row_字段名 兼容语法
    - 本月/上月关键字（自动使用当前系统月份作为基准）
    - special_模板名(本月) / special_模板名(上月)
    - 参数中的点号自动转为逗号（兼容旧写法）
    """
    if not formula:
        return 0.0

    import re
    from models import SpecialItemTemplate, EmployeeSpecialGrant, SalaryRecord, SalaryItem
    from datetime import datetime

    # 初始化变量池
    variables = extra_vars.copy() if extra_vars else {}

    # ------------------- 基准月份自动确定 -------------------
    # 如果没有传入月份，则使用当前系统月份（自然月）
    if month is None:
        month = datetime.now().month
    # 确保 month 为整数且在1-12范围内
    month = int(month)
    if month < 1:
        month = 1
    if month > 12:
        month = 12

    # ------------------- 1. 注入员工基本信息 -------------------
    if employee_id:
        emp = db.session.get(Employee, employee_id)
        if emp:
            variables.update({
                '姓名': emp.name,
                '人员身份': emp.employee_identity or '',
                '岗位级别': emp.position_level or '',
                '人员类型': emp.employee_type or '',
                '性别': emp.gender or '',
                '是否退役军人': 1 if emp.is_veteran else 0,
                '身份证号': emp.id_card or '',
            })

    def replace_special(match):
        tpl_name = match.group(1)
        args = match.group(2).strip() if match.group(2) else ''
        # 兼容点号分隔的参数（例如 2026.3 -> 2026,3）
        args = args.replace('.', ',')
        tpl = SpecialItemTemplate.query.filter_by(name=tpl_name).first()
        if not tpl:
            return '0'

        # 处理 "上月" 关键字
        if args == '上月':
            cur_year = year
            cur_month = month
            if cur_month == 1:
                last_year = cur_year - 1
                last_month = 12
            else:
                last_year = cur_year
                last_month = cur_month - 1
            grant = EmployeeSpecialGrant.query.filter_by(
                employee_id=employee_id,
                template_id=tpl.id,
                year=last_year,
                month=last_month
            ).first()
            return str(grant.amount if grant else 0)

        # 处理 "本月" 关键字
        if args == '本月':
            cur_year = year
            cur_month = month
            grant = EmployeeSpecialGrant.query.filter_by(
                employee_id=employee_id,
                template_id=tpl.id,
                year=cur_year,
                month=cur_month
            ).first()
            return str(grant.amount if grant else 0)

        parts = [p.strip() for p in args.split(',') if p.strip()]
        try:
            if len(parts) == 1:
                p = parts[0]
                if p.isdigit():
                    if len(p) == 4:  # 年份
                        grant = EmployeeSpecialGrant.query.filter_by(
                            employee_id=employee_id, template_id=tpl.id, year=int(p), month=None
                        ).first()
                        return str(grant.amount if grant else 0)
                    else:  # 月份（1-12）
                        grant = EmployeeSpecialGrant.query.filter_by(
                            employee_id=employee_id, template_id=tpl.id, year=year, month=int(p)
                        ).first()
                        return str(grant.amount if grant else 0)
            elif len(parts) == 2:
                y = int(parts[0])
                m = int(parts[1])
                grant = EmployeeSpecialGrant.query.filter_by(
                    employee_id=employee_id, template_id=tpl.id, year=y, month=m
                ).first()
                return str(grant.amount if grant else 0)
            elif len(parts) == 3:
                y = int(parts[0])
                start_m = int(parts[1])
                end_m = int(parts[2])
                grants = EmployeeSpecialGrant.query.filter(
                    EmployeeSpecialGrant.employee_id == employee_id,
                    EmployeeSpecialGrant.template_id == tpl.id,
                    EmployeeSpecialGrant.year == y,
                    EmployeeSpecialGrant.month >= start_m,
                    EmployeeSpecialGrant.month <= end_m
                ).all()
                total = sum(g.amount for g in grants)
                return str(total)
        except Exception:
            pass
        return '0'

    formula = re.sub(r'special_([a-zA-Z\u4e00-\u9fa5_][a-zA-Z0-9\u4e00-\u9fa5_]*)\(([^)]*)\)', replace_special, formula)

    # ------------------- 3. 转换 row_字段名 为 字段名（批量计算兼容） -------------------
    formula = re.sub(r'\brow_([a-zA-Z\u4e00-\u9fa5_][a-zA-Z0-9\u4e00-\u9fa5_]*)\b', r'\1', formula)

    # ------------------- 4. 处理本年/上年合计及指定年月取值（原逻辑保留并增强） -------------------
    # 4.1 本年合计（本年[工资项]合计）
    if '本年' in formula and '合计' in formula:
        current_year = year
        records_current = SalaryRecord.query.filter(
            SalaryRecord.employee_id == employee_id,
            SalaryRecord.month.like(f'{current_year}-%')
        ).all()
        pattern_current = r'本年([\u4e00-\u9fa5a-zA-Z0-9%]+)合计'
        current_matches = re.findall(pattern_current, formula)
        for item_name in current_matches:
            total = sum(float(rec.details.get(item_name, 0)) for rec in records_current if rec.details)
            variables[f'本年{item_name}合计'] = total

    # 本年收入总额
    if '本年收入总额' in formula:
        if 'records_current' not in locals():
            records_current = SalaryRecord.query.filter(
                SalaryRecord.employee_id == employee_id,
                SalaryRecord.month.like(f'{year}-%')
            ).all()
        income_items = [item.name for item in SalaryItem.query.filter_by(item_type='income')]
        total_income = 0.0
        for rec in records_current:
            details = rec.details if isinstance(rec.details, dict) else json.loads(rec.details)
            for inc in income_items:
                total_income += float(details.get(inc, 0))
        variables['本年收入总额'] = total_income

    # 4.2 上年合计
    if '上年' in formula:
        prev_year = int(year) - 1
        records_prev = SalaryRecord.query.filter(
            SalaryRecord.employee_id == employee_id,
            SalaryRecord.month.like(f'{prev_year}-%')
        ).all()
        all_items = {item.name: item.item_type for item in SalaryItem.query.all()}

        # 处理上年[工资项]合计（例如 上年基本工资合计）
        pattern_prev = r'上年([\u4e00-\u9fa5a-zA-Z0-9%]+)合计'
        matches = re.findall(pattern_prev, formula)
        total_income = 0.0
        if matches:
            if salary_item_names:
                item_names = [name for name in salary_item_names if all_items.get(name) == 'income']
            else:
                item_names = list(set([m for m in matches if m != '工资收入总额']))
            for name in item_names:
                total = sum(float(rec.details.get(name, 0)) for rec in records_prev if rec.details)
                variables[f'上年{name}合计'] = total
                total_income += total

        # ★★★ 核心修复：独立处理“上年工资收入总额” ★★★
        if '上年工资收入总额' in formula:
            if '上年工资收入总额' not in variables:
                total_income = 0.0
                for rec in records_prev:
                    details = rec.details if isinstance(rec.details, dict) else json.loads(rec.details)
                    for name, val in details.items():
                        if all_items.get(name) == 'income':
                            try:
                                total_income += float(val)
                            except:
                                pass
            variables['上年工资收入总额'] = total_income

    # 4.3 指定年份的工资项合计（例如：2026年基本工资合计）
    dyn_pattern = r'(\d{4})年([\u4e00-\u9fa5a-zA-Z0-9%]+)合计'
    for (yyyy, item_name) in re.findall(dyn_pattern, formula):
        target_year = int(yyyy)
        records = SalaryRecord.query.filter(
            SalaryRecord.employee_id == employee_id,
            SalaryRecord.month.like(f'{target_year}-%')
        ).all()
        total = sum(float(rec.details.get(item_name, 0)) for rec in records if rec.details)
        variables[f'{yyyy}年{item_name}合计'] = round(total, 2)

    # 处理本月/上月工资项（基于确定的基准月份）
    current_year = year
    current_month = month

    def replace_this_month(match):
        item_name = match.group(1)
        return f"{current_year}年{current_month}月{item_name}"

    def replace_last_month(match):
        item_name = match.group(1)
        if current_month == 1:
            last_year = current_year - 1
            last_month = 12
        else:
            last_year = current_year
            last_month = current_month - 1
        return f"{last_year}年{last_month}月{item_name}"

    formula = re.sub(r'本月([\u4e00-\u9fa5a-zA-Z0-9%]+)', replace_this_month, formula)
    formula = re.sub(r'上月([\u4e00-\u9fa5a-zA-Z0-9%]+)', replace_last_month, formula)

    # 4.4 单月工资项取值（年份可选）
    monthly_item_pattern = r'(?:(\d{4})年)?(\d{1,2})月([\u4e00-\u9fa5a-zA-Z0-9%]+)'
    for (yyyy_str, mm_str, item_name) in re.findall(monthly_item_pattern, formula):
        target_year = int(yyyy_str) if yyyy_str else year
        target_month = int(mm_str)
        month_str = f"{target_year}-{str(target_month).zfill(2)}"
        record = SalaryRecord.query.filter_by(employee_id=employee_id, month=month_str).first()
        amount = 0.0
        if record and record.details:
            details = record.details if isinstance(record.details, dict) else json.loads(record.details)
            amount = details.get(item_name, 0.0)
        var_name = f"{target_year}年{mm_str}月{item_name}"
        variables[var_name] = amount

    # 5. 最终计算
    return evaluate_formula(formula, variables)
def create_app():
    app = Flask(__name__)
    app.config.from_pyfile('config.py')
    os.makedirs(app.instance_path, exist_ok=True)
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(app.instance_path, "hr_salary.db")}'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['UPLOAD_FOLDER'] = os.path.join(app.instance_path, 'uploads')
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

    db.init_app(app)
    login_manager.init_app(app)
    login_manager.login_view = 'login'
    migrate = Migrate(app, db)

    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

    with app.app_context():
        db.create_all()
        try:
            inspector = inspect(db.engine)
            if 'print_template' not in inspector.get_table_names():
                db.create_all()
                print("创建 print_template 表")
        except Exception as e:
            print(f"检查/创建 print_template 表时出错: {str(e)}")
        try:
            inspector = inspect(db.engine)
            indexes = inspector.get_indexes('employee')
            for idx in indexes:
                if idx['unique'] and 'id_card' in idx['name']:
                    db.session.execute(text(f"DROP INDEX IF EXISTS {idx['name']}"))
                    db.session.commit()
                    print(f"已删除 id_card 唯一索引: {idx['name']}")
        except Exception as e:
            print(f"处理 id_card 索引时出错: {str(e)}")
        try:
            inspector = inspect(db.engine)
            columns = [col['name'] for col in inspector.get_columns('employee')]
            if 'phone' not in columns:
                db.session.execute(text("ALTER TABLE employee ADD COLUMN phone VARCHAR(20) DEFAULT ''"))
                db.session.commit()
                print("成功添加 phone 列到 employee 表")
        except Exception as e:
            print(f"添加 phone 列时出错: {str(e)}")
            db.session.rollback()
        try:
            inspector = inspect(db.engine)
            columns = [col['name'] for col in inspector.get_columns('employee')]
            if 'extra_fields' not in columns:
                db.session.execute(text("ALTER TABLE employee ADD COLUMN extra_fields TEXT DEFAULT '{}'"))
                db.session.commit()
                print("成功添加 extra_fields 列到 employee 表")
        except Exception as e:
            print(f"添加 extra_fields 列时出错: {str(e)}")
            db.session.rollback()
        try:
            inspector = inspect(db.engine)
            columns = [col['name'] for col in inspector.get_columns('employee')]
            if 'manual_retirement_date' not in columns:
                db.session.execute(text("ALTER TABLE employee ADD COLUMN manual_retirement_date DATE"))
                db.session.commit()
                print("成功添加 manual_retirement_date 列到 employee 表")
        except Exception as e:
            print(f"添加 manual_retirement_date 列时出错: {str(e)}")
            db.session.rollback()
        try:
            inspector = inspect(db.engine)
            columns = [col['name'] for col in inspector.get_columns('employee')]
            if 'active_for_annual_special' not in columns:
                db.session.execute(
                    text("ALTER TABLE employee ADD COLUMN active_for_annual_special BOOLEAN DEFAULT 1 NOT NULL")
                )
                db.session.commit()
                print("成功添加 active_for_annual_special 列到 employee 表")
        except Exception as e:
            print(f"添加 active_for_annual_special 列时出错: {str(e)}")
            db.session.rollback()
        # 添加 remark 列到 employee_special_item
        try:
            inspector = inspect(db.engine)
            columns = [col['name'] for col in inspector.get_columns('special_item_template')]
            if 'salary_items' not in columns:
                db.session.execute(text("ALTER TABLE special_item_template ADD COLUMN salary_items TEXT"))
                db.session.commit()
                print("成功添加 salary_items 列到 special_item_template 表")
        except Exception as e:
            print(f"添加 salary_items 列时出错: {str(e)}")
            db.session.rollback()

        # 新添加的代码
        try:
            inspector = inspect(db.engine)
            columns = [col['name'] for col in inspector.get_columns('special_item_template')]
            if 'extra_fields' not in columns:
                db.session.execute(text("ALTER TABLE special_item_template ADD COLUMN extra_fields TEXT"))
                db.session.commit()
                print("成功添加 extra_fields 列到 special_item_template 表")
        except Exception as e:
            print(f"添加 extra_fields 列时出错: {str(e)}")
            db.session.rollback()
        # 添加 remark 列到 salary_record
        try:
            inspector = inspect(db.engine)
            columns = [col['name'] for col in inspector.get_columns('special_item_template')]
            if 'salary_items' not in columns:
                db.session.execute(text("ALTER TABLE special_item_template ADD COLUMN salary_items TEXT"))
                db.session.commit()
                print("成功添加 salary_items 列到 special_item_template 表")
        except Exception as e:
            print(f"添加 salary_items 列时出错: {str(e)}")
            db.session.rollback()

        try:
            inspector = inspect(db.engine)
            columns = [col['name'] for col in inspector.get_columns('salary_record')]
            if 'remark' not in columns:
                db.session.execute(text("ALTER TABLE salary_record ADD COLUMN remark VARCHAR(200)"))
                db.session.commit()
                print("成功添加 remark 列到 salary_record 表")
        except Exception as e:
            print(f"添加 remark 列到 salary_record 时出错: {str(e)}")
            db.session.rollback()
        try:
            inspector = inspect(db.engine)
            columns = [col['name'] for col in inspector.get_columns('print_template')]
            if 'order' not in columns:
                db.session.execute(text("ALTER TABLE print_template ADD COLUMN [order] INTEGER DEFAULT 0 NOT NULL"))
                db.session.commit()
                print("成功添加 order 列到 print_template 表")
        except Exception as e:
            print(f"添加 order 列时出错: {str(e)}")
            db.session.rollback()
        try:
            inspector = inspect(db.engine)
            columns = [col['name'] for col in inspector.get_columns('print_template')]
            if 'updated_at' not in columns:
                db.session.execute(text("ALTER TABLE print_template ADD COLUMN updated_at DATETIME"))
                db.session.commit()
                print("成功添加 updated_at 列到 print_template 表")
        except Exception as e:
            print(f"添加 updated_at 列时出错: {str(e)}")
            db.session.rollback()

            # 添加 updated_by 列
        try:
            inspector = inspect(db.engine)
            columns = [col['name'] for col in inspector.get_columns('print_template')]
            if 'updated_by' not in columns:
                db.session.execute(text("ALTER TABLE print_template ADD COLUMN updated_by INTEGER REFERENCES user(id)"))
                db.session.commit()
                print("成功添加 updated_by 列到 print_template 表")
        except Exception as e:
            print(f"添加 updated_by 列时出错: {str(e)}")
            db.session.rollback()
        # 创建管理员用户
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            init_password = os.getenv('ADMIN_INIT_PASSWORD', 'admin123')
            admin = User(username='admin', is_admin=True)
            admin.set_password(init_password)
            db.session.add(admin)
            db.session.commit()
            print(f"创建管理员用户: admin/{init_password}")

    @login_manager.user_loader
    def load_user(user_id):
        return db.session.get(User, int(user_id))

    @login_manager.unauthorized_handler
    def unauthorized():
        # 仅当客户端明确期望 JSON 响应（如 AJAX 请求）时才返回 JSON
        if request.is_json or request.headers.get('Accept') == 'application/json':
            return jsonify({'error': '未登录或Session已过期，请刷新页面重新登录'}), 401
        # 否则正常重定向到登录页面
        return redirect(url_for('login'))

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if request.method == 'POST':
            username = request.form['username']
            password = request.form['password']
            user = User.query.filter_by(username=username).first()
            if user and user.check_password(password):
                login_user(user)
                next_page = request.args.get('next')
                return redirect(next_page or url_for('dashboard'))
            else:
                flash("无效的用户名或密码", "danger")
        return render_template('login.html')

    @app.route('/', endpoint='dashboard')
    @login_required
    def dashboard():
        session['last_module'] = 'dashboard'
        now = datetime.now()

        # ========== 新增：查询未来一年内退休的员工 ==========
        from dateutil.relativedelta import relativedelta
        today = now.date()
        one_year_later = today + relativedelta(years=1)
        retiring_soon = []
        employees = Employee.query.filter(Employee.active_for_payroll == True, Employee.employee_type == '在职').all()
        for emp in employees:
            ret_date = emp.retirement_date
            if ret_date and today <= ret_date <= one_year_later:
                months_left = (ret_date.year - today.year) * 12 + (ret_date.month - today.month)
                retiring_soon.append({
                    'name': emp.name,
                    'unit': emp.unit.name,
                    'retirement_date': ret_date.strftime('%Y-%m-%d'),
                    'months_left': months_left
                })
        retiring_soon.sort(key=lambda x: x['months_left'])
        # =================================================

        if request.args.get('embedded') == 'true':
            return render_template('dashboard.html', now=now, retiring_soon=retiring_soon)

        style = session.get('dashboard_style', 'sidebar')
        if style == 'sidebar':
            return render_template('dashboard_iframe.html', now=now, retiring_soon=retiring_soon)
        return render_template('dashboard.html', now=now, retiring_soon=retiring_soon)

    @app.route('/logout')
    @login_required
    def logout():
        logout_user()
        return redirect(url_for('login'))

    @app.route('/switch_dashboard_style')
    @login_required
    def switch_dashboard_style():
        """切换仪表盘风格：card 或 sidebar"""
        current_style = session.get('dashboard_style', 'sidebar')
        new_style = 'sidebar' if current_style == 'card' else 'card'
        session['dashboard_style'] = new_style
        return redirect(url_for('dashboard'))
    # 单位管理
    @app.route('/units', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def manage_units():
        if request.method == 'POST':
            if 'delete_id' in request.form:
                # 获取管理员密码
                admin_password = request.form.get('admin_password', '')
                if not current_user.check_password(admin_password):
                    flash('管理员密码错误，删除失败', 'danger')
                    return redirect(url_for('manage_units'))

                unit_id = request.form['delete_id']
                unit = db.session.get(Unit, unit_id)
                if unit:
                    Employee.query.filter_by(unit_id=unit_id).delete()
                    db.session.delete(unit)
                    db.session.commit()
                    flash(f"单位 '{unit.name}' 及其员工已删除", "success")
                else:
                    flash("单位不存在", "danger")
                return redirect(url_for('manage_units'))
            name = request.form['name']
            unit = Unit(name=name)
            db.session.add(unit)
            db.session.commit()
            flash(f"单位 '{name}' 添加成功", "success")
            return redirect(url_for('manage_units'))
        units = Unit.query.all()
        return render_template('units.html', units=units)

    # 员工管理
    @app.route('/employees', methods=['GET', 'POST'])
    @login_required
    def manage_employees():
        if request.method == 'POST':
            # 权限检查
            if not current_user.is_admin:
                flash('您没有权限修改员工信息', 'danger')
                return redirect(url_for('manage_employees'))
            # 下面是原有的业务逻辑（从第二个 if 里复制过来，保持缩进一致）
            data = request.form
            employee_id = data.get('employee_id')
            required_fields = ['unit_id', 'name', 'id_card', 'gender', 'employee_type', 'employee_identity']
            for field in required_fields:
                if not data.get(field):
                    flash(f"必填字段 {field} 不能为空", "danger")
                    return redirect(url_for('manage_employees'))
            id_card = data.get('id_card', '').strip()
            existing_emp = Employee.query.filter_by(id_card=id_card).first()
            if existing_emp:
                # 若为编辑自身，允许通过
                if employee_id and str(existing_emp.id) == employee_id:
                    pass
                else:
                    # 若已存在同类型员工，拒绝
                    if existing_emp.employee_type == data.get('employee_type', '在职'):
                        flash(f"该人员类型下已存在相同身份证号的员工: {existing_emp.name}", "danger")
                        return redirect(url_for('manage_employees'))
            email = data.get('email', '')
            if email and not validate_email(email):
                flash("邮箱格式不正确", "danger")
                return redirect(url_for('manage_employees'))
            if employee_id:
                employee = db.session.get(Employee, employee_id)
                if not employee:
                    flash("员工不存在", "danger")
                    return redirect(url_for('manage_employees'))
            else:
                employee = Employee()
            try:
                employee.unit_id = data['unit_id']
                employee.name = data['name']
                employee.id_card = data['id_card']
                employee.gender = data['gender']
                employee.birth_date = datetime.strptime(data['birth_date'], '%Y-%m-%d').date() if data.get('birth_date') else None
                employee.join_date = datetime.strptime(data['join_date'], '%Y-%m-%d').date() if data.get('join_date') else None
                employee.employee_type = data['employee_type']
                employee.employee_identity = data['employee_identity']
                employee.position_level = data.get('position_level', '')
                employee.salary_level = data.get('salary_level', '')
                employee.bank_account = data.get('bank_account', '')
                employee.is_veteran = 'is_veteran' in data
                employee.active_for_payroll = 'active_for_payroll' in data
                employee.email = data.get('email', '')
                employee.phone = data.get('phone', '')  # 新增这一行
                manual_retirement = data.get('manual_retirement_date', '')
                if manual_retirement:
                    employee.manual_retirement_date = datetime.strptime(manual_retirement, '%Y-%m-%d').date()
                else:
                    employee.manual_retirement_date = None
                employee.extra_fields = data.get('extra_fields', '{}')
                if not employee_id:
                    db.session.add(employee)
                db.session.commit()
                flash(f"员工 '{data['name']}' {'更新' if employee_id else '添加'}成功", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"操作失败: {str(e)}", "danger")
            return redirect(url_for('manage_employees'))
        from collections import defaultdict  # 放在文件开头（已有，可确认）

        # 替换原来的 employees 和 units 查询部分
        units = Unit.query.all()

        # 按员工类型分组并排序
        employees = Employee.query.all()
        employees_by_type = defaultdict(list)
        for emp in employees:
            etype = emp.employee_type or '未分类'
            employees_by_type[etype].append(emp)

        type_order = ['在职', '退休', '死亡', '其他', '未分类']
        sorted_type_groups = []
        for t in type_order:
            if t in employees_by_type:
                sorted_type_groups.append((t, employees_by_type[t]))
        for t, emps in employees_by_type.items():
            if t not in type_order:
                sorted_type_groups.append((t, emps))

        employee_id = request.args.get('employee_id')
        employee = db.session.get(Employee, employee_id) if employee_id else None
        return render_template('employees.html',
                               employees_by_type=sorted_type_groups,
                               units=units,
                               employee=employee,
                               today=datetime.now().strftime('%Y-%m-%d'))
    @app.route('/employees/edit/<int:employee_id>')
    @login_required
    def edit_employee(employee_id):
        return redirect(f"{url_for('manage_employees', employee_id=employee_id)}#employee-form")

    @app.route('/employees/delete/<int:employee_id>', methods=['POST'])
    @login_required
    def delete_employee(employee_id):
        if not current_user.is_admin:
            flash('您没有权限删除员工', 'danger')
            return redirect(url_for('manage_employees'))

        employee = db.session.get(Employee, employee_id)  # ← 必须获取对象
        if employee:
            EmployeeSpecialItem.query.filter_by(employee_id=employee_id).delete()
            if 'EmployeeAssessment' in globals():
                EmployeeAssessment.query.filter_by(employee_id=employee_id).delete()
            SalaryRecord.query.filter_by(employee_id=employee_id).delete()
            db.session.delete(employee)
            db.session.commit()
            flash(f"员工 '{employee.name}' 已删除", "success")
        else:
            flash("员工不存在", "danger")
        return redirect(url_for('manage_employees'))

    def validate_email(email):
        """简单的邮箱格式验证"""
        if not email:
            return True
        import re
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email) is not None
    def validate_id_card(id_card):
        if not id_card:
            return False, "身份证号不能为空"
        id_card = str(id_card).strip().upper()
        id_card = id_card.replace(' ', '').replace('-', '')
        if len(id_card) not in (15, 18):
            return False, "身份证号长度错误(应为15或18位)"
        if len(id_card) == 18 and not id_card[:17].isdigit():
            return False, "身份证前17位必须为数字"
        if len(id_card) == 18:
            factors = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
            check_codes = ['1', '0', 'X', '9', '8', '7', '6', '5', '4', '3', '2']
            try:
                total = sum(int(id_card[i]) * factors[i] for i in range(17))
                check_code = check_codes[total % 11]
                if id_card[17] != check_code:
                    return False, "身份证校验码错误"
            except:
                return False, "身份证校验码计算错误"
        elif len(id_card) == 15:
            try:
                birth_date_str = '19' + id_card[6:12]
                datetime.strptime(birth_date_str, '%Y%m%d')
            except:
                return False, "身份证包含无效的出生日期"
        return True, ""

    # 员工导入
    @app.route('/employees/import', methods=['POST'])
    @login_required
    def import_employees():
        if not current_user.is_admin:
            flash('您没有权限导入员工数据', 'danger')
            return redirect(url_for('manage_employees'))
        if 'file' not in request.files:
            flash('没有选择文件', 'danger')
            return redirect(url_for('manage_employees'))
        file = request.files['file']
        if file.filename == '':
            flash('没有选择文件', 'danger')
            return redirect(url_for('manage_employees'))
        try:
            df = pd.read_excel(file, dtype={'身份证号': str, '银行卡号': str})
            df.columns = df.columns.str.strip()
            import numpy as np
            df = df.replace(to_replace=r'^#\w+[!/]*.*$|^#REF!$|^#N/A$|^#VALUE!$', value=np.nan, regex=True)

            def safe_str(row, key, default=''):
                val = row.get(key)
                if pd.isna(val):
                    return default
                s = str(val).strip()
                # 如果值是 Excel 错误字符串，也返回默认值
                if s.startswith('#') and any(c in s for c in ('N/A', 'REF!', 'VALUE!', 'NAME?', 'NUM!', 'NULL!')):
                    return default
                return s

            def safe_date(row, key):
                val = row.get(key)
                if pd.isna(val):
                    return None
                if isinstance(val, str) and val.strip().startswith('#'):
                    return None
                try:
                    if isinstance(val, str):
                        return datetime.strptime(val.strip(), '%Y-%m-%d').date()
                    elif isinstance(val, (datetime, pd.Timestamp)):
                        return val.date()
                except:
                    pass
                return None
            count = 0
            errors = []
            imported_ids = set()
            for index, row in df.iterrows():
                row_num = index + 2
                try:
                    if pd.isna(row.get('姓名')) or pd.isna(row.get('身份证号')):
                        errors.append(f"第{row_num}行: 缺少姓名或身份证号")
                        continue
                    id_card = str(row['身份证号']).strip().upper()
                    id_card = id_card.replace(' ', '').replace('-', '')
                    valid, msg = validate_id_card(id_card)
                    if not valid:
                        errors.append(f"第{row_num}行: 身份证号格式错误 - {msg}")
                        continue
                    if id_card in imported_ids:
                        errors.append(f"第{row_num}行: 身份证号 {id_card} 在本文件中重复")
                        continue
                    existing = Employee.query.filter_by(id_card=id_card).first()
                    if existing:
                        existing_type = existing.employee_type or ''
                        import_type = safe_str(row, '人员类型', '在职')
                        if existing_type == import_type:
                            errors.append(f"第{row_num}行: 身份证号 {id_card} 在'{import_type}'类型中已存在")
                            continue
                    gender = '男'
                    birth_date = None
                    if len(id_card) == 18:
                        gender_num = int(id_card[16])
                        gender = '男' if gender_num % 2 == 1 else '女'
                        birth_date_str = id_card[6:14]
                        try:
                            birth_date = datetime.strptime(birth_date_str, '%Y%m%d').date()
                        except:
                            errors.append(f"第{row_num}行: 无法从身份证解析出生日期")
                            continue
                    elif len(id_card) == 15:
                        gender_num = int(id_card[14])
                        gender = '男' if gender_num % 2 == 1 else '女'
                        birth_date_str = '19' + id_card[6:12]
                        try:
                            birth_date = datetime.strptime(birth_date_str, '%Y%m%d').date()
                        except:
                            errors.append(f"第{row_num}行: 无法从身份证解析出生日期")
                            continue
                    unit_name = str(row.get('单位', '重庆碚区江北蚕种场')).strip()
                    if not unit_name:
                        unit_name = '重庆碚区江北蚕种场'
                    unit = Unit.query.filter_by(name=unit_name).first()
                    if not unit:
                        unit = Unit(name=unit_name)
                        db.session.add(unit)
                        db.session.flush()

                    name = safe_str(row, '姓名')
                    employee_type = safe_str(row, '人员类型', '在职')
                    employee_identity = safe_str(row, '人员身份', '工勤')
                    position_level = safe_str(row, '岗位级别', '')
                    salary_level = safe_str(row, '薪资级别', '')
                    bank_account = safe_str(row, '银行卡号', '')
                    # 银行卡号只保留数字
                    bank_account = ''.join(filter(str.isdigit, bank_account))
                    is_veteran = safe_str(row, '是否退役军人', '否') == '是'
                    email = safe_str(row, '邮箱', '').lower()
                    if email and not validate_email(email):
                        errors.append(f"第{row_num}行: 邮箱格式不正确，已清空")
                        email = ''
                    phone = safe_str(row, '电话号码', '')
                    # 入职日期和退休时间使用 safe_date

                    join_date = safe_date(row, '入职日期')
                    manual_retirement = safe_date(row, '退休时间')

                    employee = Employee(
                        name=name,
                        id_card=id_card,
                        gender=gender,
                        birth_date=birth_date,
                        join_date=join_date,
                        employee_type=employee_type,
                        employee_identity=employee_identity,
                        position_level=position_level,
                        salary_level=salary_level,
                        bank_account=bank_account,
                        is_veteran=is_veteran,
                        email=email,
                        phone=phone,
                        manual_retirement_date=manual_retirement,
                        unit_id=unit.id,
                        active_for_payroll=True,
                        extra_fields='{}'
                    )
                    db.session.add(employee)
                    imported_ids.add(id_card)
                    count += 1
                except Exception as e:
                    errors.append(f"第{row_num}行: 处理错误 - {str(e)}")
                    continue
            db.session.commit()
            if errors:
                error_msg = f'成功导入 {count} 名员工, 但有 {len(errors)} 条错误'
                if len(errors) > 5:
                    error_msg += f"(显示前5条): {'; '.join(errors[:5])}..."
                else:
                    error_msg += f": {'; '.join(errors)}"
                flash(error_msg, 'warning')
            else:
                flash(f'成功导入 {count} 名员工', 'success')
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"导入员工失败: {str(e)}", exc_info=True)
            flash(f'导入失败: {str(e)}', 'danger')
        return redirect(url_for('manage_employees'))

    # 员工导出
    @app.route('/employees/export')
    @login_required
    def export_employees():
        employees = Employee.query.all()
        data = [{
            'ID': emp.id,
            '姓名': emp.name,
            '单位': emp.unit.name,
            '身份证号': str(emp.id_card) if emp.id_card else '',
            '性别': emp.gender,
            '出生日期': emp.birth_date.strftime('%Y-%m-%d') if emp.birth_date else '',
            '人员类型': emp.employee_type,
            '人员身份': emp.employee_identity,
            '入职日期': emp.join_date.strftime('%Y-%m-%d') if emp.join_date else '',
            '银行卡号': emp.bank_account,
            '是否退役军人': '是' if emp.is_veteran else '否',
            '邮箱': emp.email,
            '电话号码': emp.phone,
            '扩展信息': emp.extra_fields  # ← 新增这一列
        } for emp in employees]
        export_dir = os.path.join(current_dir, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = f"员工列表_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        filepath = os.path.join(export_dir, filename)
        df = pd.DataFrame(data)
        df.to_excel(filepath, index=False)
        return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # 工资项管理
    @app.route('/salary_items', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def manage_salary_items():
        if request.method == 'POST':
            # 删除工资项
            if 'delete_id' in request.form:
                item_id = request.form['delete_id']
                item = db.session.get(SalaryItem, item_id)
                if item:
                    used_count = 0
                    for record in SalaryRecord.query.all():
                        if record.details and item.name in record.details:
                            used_count += 1
                    if used_count > 0:
                        flash(f"工资项 '{item.name}' 已在 {used_count} 条工资记录中使用，无法删除", "danger")
                    else:
                        db.session.delete(item)
                        db.session.commit()
                        flash(f"工资项 '{item.name}' 已删除", "success")
                else:
                    flash("工资项不存在", "danger")
                return redirect(url_for('manage_salary_items'))
            # 添加工资项
            name = request.form.get('name')
            item_type = request.form.get('item_type')
            type_mapping = {'收入项': 'income', '扣款项': 'deduction', '计算项': 'calculation'}
            db_item_type = type_mapping.get(item_type, 'income')
            formula = request.form.get('formula', '')
            default_value = float(request.form.get('default_value', 0))
            if formula:
                try:
                    validate_formula(formula)
                except ValueError as e:
                    flash(f"公式验证失败: {str(e)}", "danger")
                    return redirect(url_for('manage_salary_items'))
            max_order = db.session.query(db.func.max(SalaryItem.order)).scalar() or 0
            new_item = SalaryItem(name=name, item_type=db_item_type, formula=formula, default_value=default_value, order=max_order + 1)
            db.session.add(new_item)
            db.session.commit()
            flash(f"工资项 '{name}' 添加成功", "success")
            return redirect(url_for('manage_salary_items'))
        items = SalaryItem.query.order_by(SalaryItem.order).all()
        income_count = SalaryItem.query.filter_by(item_type='income').count()
        deduction_count = SalaryItem.query.filter_by(item_type='deduction').count()
        calculation_count = SalaryItem.query.filter_by(item_type='calculation').count()
        return render_template('salary_items.html', items=items, income_count=income_count,
                               deduction_count=deduction_count, calculation_count=calculation_count, now=datetime.now())

    def validate_formula(formula):
        if not formula:
            return
        illegal_chars = ["￥", "$", "&", "<", ">", "'", "\""]
        for char in illegal_chars:
            if char in formula:
                raise ValueError(f"公式中包含非法字符: '{char}'")
        if not re.match(r'^[\u4e00-\u9fa5a-zA-Z0-9\+\-\*\/\(\)\s\.]+$', formula):
            raise ValueError("公式包含不支持的字符")
        if formula.count('(') != formula.count(')'):
            raise ValueError("括号不匹配")
        try:
            test_formula = re.sub(r'[\u4e00-\u9fa5]+', '1', formula)
            simple_eval(test_formula)
        except Exception as e:
            raise ValueError(f"公式语法错误: {str(e)}")

    @app.route('/salary_items/edit/<int:item_id>', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def edit_salary_item(item_id):
        item = db.session.get(SalaryItem, item_id)
        if not item:
            abort(404)

        if request.method == 'POST':
            from models import PrintTemplate
            try:
                new_name = request.form.get('name', '').strip()
                old_name = item.name
                item_type_display = request.form.get('item_type')
                type_mapping = {'收入项': 'income', '扣款项': 'deduction', '计算项': 'calculation'}
                new_item_type = type_mapping.get(item_type_display, 'income')
                formula = request.form.get('formula', '')
                default_value = float(request.form.get('default_value', 0))
                calc_direction = request.form.get('calc_direction', 'income')

                # 1. 更新当前工资项基本信息
                item.name = new_name
                item.item_type = new_item_type
                item.formula = formula
                item.default_value = default_value
                item.calc_direction = calc_direction

                # 2. 如果名称发生变化，执行数据迁移
                if new_name != old_name:
                    try:
                        # 全表扫描（数据量大时可分批，但一般人事系统规模不大）
                        records = SalaryRecord.query.all()
                        for rec in records:
                            # 处理 details 字段
                            if rec.details is None:
                                rec.details = {}
                            if isinstance(rec.details, str):
                                try:
                                    rec.details = json.loads(rec.details)
                                except:
                                    rec.details = {}
                            if old_name in rec.details:
                                rec.details[new_name] = rec.details.pop(old_name)
                                flag_modified(rec, "details")

                            # 处理 item_remarks 字段
                            if rec.item_remarks:
                                if isinstance(rec.item_remarks, str):
                                    try:
                                        remarks = json.loads(rec.item_remarks)
                                    except:
                                        remarks = {}
                                else:
                                    remarks = rec.item_remarks
                                if old_name in remarks:
                                    remarks[new_name] = remarks.pop(old_name)
                                    rec.item_remarks = json.dumps(remarks, ensure_ascii=False)
                                    flag_modified(rec, "item_remarks")

                        # 处理 EmployeeSpecialGrant.extra_data
                        grants = EmployeeSpecialGrant.query.all()
                        for grant in grants:
                            if grant.extra_data:
                                try:
                                    extra = json.loads(grant.extra_data)
                                    changed = False

                                    def recurse(d):
                                        nonlocal changed
                                        if isinstance(d, dict):
                                            new_d = {}
                                            for k, v in d.items():
                                                new_k = k.replace(old_name, new_name) if isinstance(k,
                                                                                                    str) and old_name in k else k
                                                if new_k != k:
                                                    changed = True
                                                new_d[new_k] = recurse(v)
                                            return new_d
                                        elif isinstance(d, str):
                                            new_v = d.replace(old_name, new_name)
                                            if new_v != d:
                                                changed = True
                                            return new_v
                                        else:
                                            return d

                                    new_extra = recurse(extra)
                                    if changed:
                                        grant.extra_data = json.dumps(new_extra, ensure_ascii=False)
                                        flag_modified(grant, "extra_data")
                                except:
                                    pass

                        # 更新特殊事项模板公式
                        special_templates = SpecialItemTemplate.query.order_by(SpecialItemTemplate.order).all()
                        for st in special_templates:
                            if st.formula and old_name in st.formula:
                                st.formula = st.formula.replace(old_name, new_name)

                        # 更新其他工资项公式
                        all_items = SalaryItem.query.filter(SalaryItem.id != item.id).all()
                        for si in all_items:
                            if si.formula and old_name in si.formula:
                                si.formula = si.formula.replace(old_name, new_name)

                        # 更新打印模板配置
                        templates = PrintTemplate.query.all()
                        for tpl in templates:
                            config = json.loads(tpl.config)
                            changed = False
                            if 'columns' in config:
                                for i, col in enumerate(config['columns']):
                                    if col == old_name:
                                        config['columns'][i] = new_name
                                        changed = True
                            if 'column_display' in config:
                                new_display = {}
                                for k, v in config['column_display'].items():
                                    new_k = k if k != old_name else new_name
                                    if new_k != k:
                                        changed = True
                                    new_display[new_k] = v
                                config['column_display'] = new_display
                            if changed:
                                tpl.config = json.dumps(config, ensure_ascii=False)

                        db.session.commit()
                        flash(f'工资项 "{old_name}" 已更名为 "{new_name}"，所有历史数据及关联配置已同步更新', 'success')
                    except Exception as e:
                        db.session.rollback()
                        flash(f'迁移数据失败: {str(e)}，请联系管理员', 'danger')
                        current_app.logger.error(f"工资项改名迁移失败: {e}", exc_info=True)
                        return redirect(url_for('edit_salary_item', item_id=item_id))
                else:
                    db.session.commit()
                    flash(f'工资项 "{item.name}" 更新成功', 'success')

                return redirect(url_for('manage_salary_items'))

            except Exception as e:
                db.session.rollback()
                flash(f'更新失败: {str(e)}', 'danger')
                current_app.logger.error(f"编辑工资项失败: {e}", exc_info=True)
                return redirect(url_for('edit_salary_item', item_id=item_id))

        # GET 请求
        reverse_mapping = {'income': '收入项', 'deduction': '扣款项', 'calculation': '计算项'}
        item_type_display = reverse_mapping.get(item.item_type, '收入项')
        return render_template('edit_salary_item.html', item=item, item_type_display=item_type_display)

    @app.route('/salary_items/import', methods=['POST'])
    @login_required
    @admin_required
    def import_salary_items():
        if 'file' not in request.files:
            flash('没有选择文件', 'danger')
            return redirect(url_for('manage_salary_items'))
        file = request.files['file']
        if file.filename == '':
            flash('没有选择文件', 'danger')
            return redirect(url_for('manage_salary_items'))
        try:
            df = pd.read_excel(file)
            count = 0
            for _, row in df.iterrows():
                existing = SalaryItem.query.filter_by(name=row['项目名称']).first()
                if existing:
                    continue
                new_item = SalaryItem(name=row['项目名称'], item_type=row['项目类型'], formula=row.get('计算公式', ''), default_value=row.get('默认值', 0.0))
                db.session.add(new_item)
                count += 1
            db.session.commit()
            flash(f'成功导入 {count} 个工资项', 'success')
        except Exception as e:
            flash(f'导入失败: {str(e)}', 'danger')
        return redirect(url_for('manage_salary_items'))

    @app.route('/salary_items/save_order', methods=['POST'])
    @login_required
    @admin_required
    def save_salary_items_order():
        order_data = request.get_json()
        for item in order_data:
            salary_item = db.session.get(SalaryItem, item['id'])
            if salary_item:
                salary_item.order = item['order']
        try:
            db.session.commit()
            return jsonify(success=True)
        except Exception as e:
            db.session.rollback()
            return jsonify(success=False, error=str(e)), 500

    @app.route('/special_items/advanced_import', methods=['POST'])
    @login_required
    @admin_required
    def advanced_import_special():
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '未上传文件'}), 400
        file = request.files['file']
        template_id = request.form.get('template_id')
        year = int(request.form.get('year'))
        month = request.form.get('month')
        if month and month.strip():
            month = int(month)
        else:
            month = None
        mapping = request.form.get('mapping')

        template = db.session.get(SpecialItemTemplate, template_id)
        if not template:
            return jsonify({'success': False, 'error': '模板不存在'}), 404

        try:
            df = pd.read_excel(file)
        except Exception as e:
            return jsonify({'success': False, 'error': f'读取Excel失败: {str(e)}'}), 400

        # 如果没有传递 mapping，自动根据中文列名构建映射
        if not mapping:
            col_map = {}
            for col in df.columns:
                col_stripped = col.strip()
                if '姓名' in col_stripped:
                    col_map['employee_name'] = col
                elif '金额' in col_stripped:
                    col_map['amount'] = col
                elif '备注' in col_stripped:
                    col_map['remark'] = col
                elif '工作月数' in col_stripped:
                    col_map['work_months'] = col
                elif '未休天数' in col_stripped:
                    col_map['unused_days'] = col
                elif '考核系数' in col_stripped:
                    col_map['assessment_coefficient'] = col
            mapping = json.dumps(col_map)

        if not mapping:
            return jsonify({'success': False, 'error': '缺少列映射配置'}), 400
        mapping = json.loads(mapping)

        count = 0
        errors = []
        for idx, row in df.iterrows():
            try:
                emp_name = row.get(mapping.get('employee_name')) if mapping.get('employee_name') else None
                # 仅使用姓名匹配员工，不再使用身份证
                employee = None
                if emp_name and pd.notna(emp_name):
                    employees = Employee.query.filter_by(name=str(emp_name).strip()).all()
                    if len(employees) == 1:
                        employee = employees[0]
                    elif len(employees) > 1:
                        errors.append(f"行{idx + 2}: 存在多个同名员工 '{emp_name}'，请手动处理")
                        continue
                    else:
                        errors.append(f"行{idx + 2}: 未找到员工 '{emp_name}'")
                        continue
                else:
                    errors.append(f"行{idx + 2}: 员工姓名为空")
                    continue

                amount_field = mapping.get('amount')
                if not amount_field:
                    errors.append(f"行{idx + 2}: 未指定金额列")
                    continue
                amount = float(row[amount_field]) if pd.notna(row[amount_field]) else 0.0

                remark_field = mapping.get('remark')
                remark = str(row[remark_field]) if remark_field and pd.notna(row[remark_field]) else ''

                extra = {}
                for extra_field in ['unused_days', 'assessment_coefficient', 'work_months']:
                    if mapping.get(extra_field) and mapping[extra_field] in row and pd.notna(row[mapping[extra_field]]):
                        extra[extra_field] = float(row[mapping[extra_field]])

                grant = EmployeeSpecialGrant(
                    employee_id=employee.id,
                    template_id=template.id,
                    year=year,
                    month=month,
                    amount=amount,
                    remark=remark,
                    extra_data=json.dumps(extra, ensure_ascii=False),
                    source_file=file.filename,
                    grant_date=datetime.now().date()
                )
                db.session.add(grant)
                count += 1
            except Exception as e:
                errors.append(f"行{idx + 2}: {str(e)}")
        db.session.commit()
        return jsonify({'success': True, 'count': count, 'errors': errors})

    @app.route('/api/special_templates', methods=['GET'])
    @login_required
    def api_special_templates():
        active_only = request.args.get('active_only', 'true').lower() == 'true'
        if active_only:
            templates = SpecialItemTemplate.query.filter_by(is_active=True).order_by(SpecialItemTemplate.order).all()
        else:
            templates = SpecialItemTemplate.query.order_by(SpecialItemTemplate.order).all()
        return jsonify([{
            'id': t.id, 'name': t.name, 'frequency': t.frequency,
            'calculation_type': t.calculation_type, 'formula': t.formula,
            'extra_fields': t.extra_fields, 'base_amount': t.base_amount,
            'is_active': t.is_active  # 新增这一行，前端需要
        } for t in templates])

    @app.route('/api/special_templates/<int:template_id>', methods=['DELETE'])
    @login_required
    @admin_required
    def delete_special_template(template_id):
        """删除特殊事项模板（需验证管理员密码，且检查是否存在发放记录）"""
        data = request.get_json()
        password = data.get('admin_password', '')
        if not current_user.check_password(password):
            return jsonify({'success': False, 'error': '管理员密码错误'}), 403

        template = SpecialItemTemplate.query.get_or_404(template_id)

        # 检查是否有发放记录
        grant_count = EmployeeSpecialGrant.query.filter_by(template_id=template_id).count()
        if grant_count > 0:
            return jsonify({
                'success': False,
                'error': f'该模板已有 {grant_count} 条发放记录，请先清空记录或确认删除（目前不支持带记录删除）。'
            }), 400

        # 删除关联的发放记录（理论上已被上面拦截，但保留以防未来调整）
        EmployeeSpecialGrant.query.filter_by(template_id=template_id).delete()
        db.session.delete(template)
        db.session.commit()

        return jsonify({'success': True, 'message': f'模板 “{template.name}” 已删除'})
    @app.route('/api/employee_special_amounts', methods=['GET'])
    @login_required
    def api_employee_special_amounts():
        employee_id = request.args.get('employee_id')
        year = request.args.get('year')
        month = request.args.get('month')
        if not employee_id or not year:
            return jsonify({'error': '缺少参数'}), 400
        employee = db.session.get(Employee, employee_id)
        if not employee:
            return jsonify({'error': '员工不存在'}), 404
        month_int = int(month) if month else None
        results = {}
        templates = SpecialItemTemplate.query.filter_by(is_active=True).all()
        for tpl in templates:
            if tpl.frequency == 'monthly' and month_int:
                amount = employee.get_special_amount(tpl.name, int(year), month_int)
            elif tpl.frequency in ('yearly', 'once'):
                amount = employee.get_special_amount(tpl.name, int(year))
            else:
                amount = 0.0
            if amount != 0:
                results[tpl.name] = amount
        return jsonify(results)

    @app.route('/api/employee_retirement_preview')
    @login_required
    def api_employee_retirement_preview():
        birth_date_str = request.args.get('birth_date')
        gender = request.args.get('gender')
        identity = request.args.get('identity')
        if not birth_date_str:
            return jsonify({'retirement_date': ''})
        try:
            birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
            if gender == '男':
                base_age = 60
                base_date = date(1964, 12, 1)  # ✅ 修正
            else:
                if identity == '工勤':
                    base_age = 55
                else:
                    base_age = 55  # 如需女性管理人员60岁，请修改此处
                base_date = date(1969, 12, 1)  # ✅ 修正
            months_diff = (birth_date.year - base_date.year) * 12 + (
                        birth_date.month - base_date.month) if birth_date >= base_date else 0
            delay = min(36, max(0, (months_diff + 3) // 4))
            total_months = base_age * 12 + delay
            ret_date = birth_date + relativedelta(months=total_months)
            return jsonify({'retirement_date': ret_date.strftime('%Y-%m-%d')})
        except Exception as e:
            return jsonify({'retirement_date': ''})

    @app.route('/api/employees/batch_toggle_non_regular', methods=['POST'])
    @login_required
    @admin_required
    def batch_toggle_non_regular():
        """批量切换非在职员工是否参与年度特殊事项（仅影响 special_annual.html）"""
        data = request.get_json()
        emp_ids = data.get('employee_ids', [])
        value = data.get('active', True)  # True 启用, False 停用
        if not emp_ids:
            return jsonify({'success': False, 'error': '未选择员工'}), 400

        updated = Employee.query.filter(Employee.id.in_(emp_ids)).update(
            {Employee.active_for_annual_special: bool(value)}, synchronize_session='fetch'
        )
        db.session.commit()
        return jsonify({'success': True, 'updated': updated})

    @app.route('/api/employees/batch_toggle_payroll', methods=['POST'])
    @login_required
    @admin_required
    def batch_toggle_payroll():
        """批量切换员工的工资计算状态（启用/停用）"""
        data = request.get_json()
        emp_ids = data.get('employee_ids', [])
        value = data.get('active', True)  # True=启用, False=停用
        if not emp_ids:
            return jsonify({'success': False, 'error': '未选择员工'}), 400

        updated = Employee.query.filter(Employee.id.in_(emp_ids)).update(
            {Employee.active_for_payroll: bool(value)}, synchronize_session='fetch'
        )
        db.session.commit()
        return jsonify({'success': True, 'updated': updated})
    @app.route('/special_items', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def manage_special_items():
        if request.method == 'POST':
            if 'delete_id' in request.form:
                # ✅ 新增：密码保护
                admin_pwd = request.form.get('admin_password', '')
                if not current_user.check_password(admin_pwd):
                    flash('管理员密码错误，删除失败', 'danger')
                    return redirect(url_for('manage_special_items'))

                item_id = request.form['delete_id']
                item = db.session.get(SpecialItemTemplate, item_id)
                if item:
                    EmployeeSpecialGrant.query.filter_by(template_id=item_id).delete()
                    db.session.delete(item)
                    db.session.commit()
                    flash(f"特殊事项模板 '{item.name}' 已删除", "success")
                else:
                    flash("事项模板不存在", "danger")
                return redirect(url_for('manage_special_items'))

            name = request.form.get('name')
            category = request.form.get('category', '')
            base_amount = float(request.form.get('base_amount', 0))
            formula = request.form.get('formula', '')
            frequency = request.form.get('frequency', 'monthly')
            is_active = 'is_active' in request.form
            calculation_type = request.form.get('calculation_type', 'fixed')
            extra_fields = request.form.get('extra_fields', '')
            if not extra_fields or extra_fields == '[]':
                # 兼容旧版复选框（如果没有传递 extra_fields，则使用旧逻辑）
                extra_fields_list = []
                if request.form.get('extra_work_months'):
                    extra_fields_list.append({"name": "工作月数", "label": "工作月数", "type": "number", "default": 12})
                if request.form.get('extra_unused_days'):
                    extra_fields_list.append({"name": "未休天数", "label": "未休天数", "type": "number", "default": 0})
                if request.form.get('extra_coefficient'):
                    extra_fields_list.append(
                        {"name": "考核系数", "label": "考核系数", "type": "number", "default": 1.0})
                extra_fields = json.dumps(extra_fields_list, ensure_ascii=False)
            salary_items = request.form.get('salary_items', '')
            new_template = SpecialItemTemplate(
                name=name,
                category=category,
                base_amount=base_amount,
                formula=formula,
                frequency=frequency,
                is_active=is_active,
                calculation_type=calculation_type,
                salary_items=salary_items,
                extra_fields=extra_fields
            )
            db.session.add(new_template)
            db.session.commit()
            flash(f"特殊事项模板 '{name}' 添加成功", "success")
            return redirect(url_for('manage_special_items'))

        templates = SpecialItemTemplate.query.order_by(SpecialItemTemplate.order).all()
        units = Unit.query.all()
        current_month = datetime.now().strftime('%Y-%m')
        active_items = SpecialItemTemplate.query.filter_by(is_active=True).all()
        grants = EmployeeSpecialGrant.query.all()
        all_employees = Employee.query.all()
        return render_template('special_items.html',  # <-- 修改此处
                               templates=templates,
                               units=units,
                               active_items=active_items,
                               current_month=current_month,
                               now=datetime.now(),
                               grants=grants,
                               all_employees=all_employees)

    @app.route('/special_items/edit/<int:item_id>', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def edit_special_item(item_id):
        item = SpecialItemTemplate.query.get_or_404(item_id)

        if request.method == 'POST':
            # 处理表单提交，更新模板字段
            item.name = request.form.get('name')
            item.category = request.form.get('category', '')
            base_amount_str = request.form.get('base_amount', '0')
            item.base_amount = float(base_amount_str) if base_amount_str.strip() else 0.0
            item.formula = request.form.get('formula', '')
            item.frequency = request.form.get('frequency', 'monthly')
            item.calculation_type = request.form.get('calculation_type', 'fixed')
            item.is_active = 'is_active' in request.form
            item.salary_items = request.form.get('salary_items', '')  # 新增字段
            extra_fields = request.form.get('extra_fields', '')
            if not extra_fields or extra_fields == '[]':
                # 兼容旧版复选框
                extra_fields_list = []
                if request.form.get('extra_work_months'):
                    extra_fields_list.append({"name": "工作月数", "label": "工作月数", "type": "number", "default": 12})
                if request.form.get('extra_unused_days'):
                    extra_fields_list.append({"name": "未休天数", "label": "未休天数", "type": "number", "default": 0})
                if request.form.get('extra_coefficient'):
                    extra_fields_list.append(
                        {"name": "考核系数", "label": "考核系数", "type": "number", "default": 1.0})
                extra_fields = json.dumps(extra_fields_list, ensure_ascii=False)
            item.extra_fields = extra_fields

            # 条件表达式（如果模板中有此字段）
            if 'condition' in request.form:
                item.condition = request.form.get('condition', '')

            try:
                db.session.commit()
                flash(f"特殊事项模板 '{item.name}' 更新成功", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"更新失败: {str(e)}", "danger")
                current_app.logger.error(f"编辑特殊事项模板失败: {e}", exc_info=True)

            return redirect(url_for('manage_special_items'))

        # ========== GET 请求：准备数据 ==========
        # 1. 获取所有工资项，用于“纳入上年合计”多选框
        all_salary_items = SalaryItem.query.order_by(SalaryItem.order).all()

        # 2. 将当前模板的 salary_items (JSON 字符串) 转为列表，便于模板判断选中状态
        salary_items_list = []
        if item.salary_items:
            try:
                salary_items_list = json.loads(item.salary_items)
            except:
                salary_items_list = []

        return render_template(
            'edit_special_item.html',
            item=item,
            all_salary_items=all_salary_items,
            salary_items_list=salary_items_list
        )

    @app.route('/assign_special_items', methods=['POST'])
    @login_required
    @admin_required
    def assign_special_items():
        employee_ids = request.form.getlist('employee_ids')
        special_item_id = request.form['special_item_id']
        month = request.form['month']
        assessment_value = request.form.get('assessment_value', 0)
        remark = request.form.get('remark', '')
        count = 0
        special_item = db.session.get(SpecialSalaryItem, special_item_id)
        if not special_item:
            flash("特殊事项不存在", "danger")
            return redirect(url_for('manage_special_items') + '#assignment')
        for emp_id in employee_ids:
            existing = EmployeeSpecialItem.query.filter_by(employee_id=emp_id, special_item_id=special_item_id, month=month).first()
            if existing:
                existing.assessment_value = float(assessment_value)
                existing.calculated_amount = calculate_special_amount(special_item, float(assessment_value))
                existing.remark = remark
                count += 1
            else:
                calculated_amount = calculate_special_amount(special_item, float(assessment_value))
                new_assignment = EmployeeSpecialItem(employee_id=emp_id, special_item_id=special_item_id, month=month, assessment_value=float(assessment_value), calculated_amount=calculated_amount, remark=remark)
                db.session.add(new_assignment)
                count += 1
        db.session.commit()
        flash(f"成功为 {count} 名员工分配特殊事项 '{special_item.name}'", "success")
        return redirect(url_for('manage_special_items') + '#assignment')

    @app.route('/batch_import_special_items', methods=['POST'])
    @login_required
    @admin_required
    def batch_import_special_items():
        if 'file' not in request.files:
            flash('没有选择文件', 'danger')
            return redirect(url_for('manage_special_items') + '#batch')
        file = request.files['file']
        if file.filename == '':
            flash('没有选择文件', 'danger')
            return redirect(url_for('manage_special_items') + '#batch')
        special_item_id = request.form['special_item_id']
        month = request.form['month']
        special_item = db.session.get(SpecialSalaryItem, special_item_id)
        if not special_item:
            flash("特殊事项不存在", "danger")
            return redirect(url_for('manage_special_items') + '#batch')
        try:
            df = pd.read_excel(file)
            required_columns = ['employee_id', 'assessment_value']
            for col in required_columns:
                if col not in df.columns:
                    flash(f"缺少必要列: {col}", "danger")
                    return redirect(url_for('manage_special_items') + '#batch')
            count = 0
            errors = []
            for index, row in df.iterrows():
                try:
                    employee_id = int(row['employee_id'])
                    assessment_value = float(row['assessment_value'])
                    remark = str(row.get('remark', ''))
                    employee = db.session.get(Employee, employee_id)
                    if not employee:
                        errors.append(f"行 {index + 2}: 员工ID {employee_id} 不存在")
                        continue
                    calculated_amount = calculate_special_amount(special_item, assessment_value)
                    existing = EmployeeSpecialItem.query.filter_by(employee_id=employee_id, special_item_id=special_item_id, month=month).first()
                    if existing:
                        existing.assessment_value = assessment_value
                        existing.calculated_amount = calculated_amount
                        existing.remark = remark
                    else:
                        new_assignment = EmployeeSpecialItem(employee_id=employee_id, special_item_id=special_item_id, month=month, assessment_value=assessment_value, calculated_amount=calculated_amount, remark=remark)
                        db.session.add(new_assignment)
                    count += 1
                except Exception as e:
                    errors.append(f"行 {index + 2}: 处理错误 - {str(e)}")
            db.session.commit()
            if errors:
                flash(f"成功导入 {count} 条记录，但有 {len(errors)} 个错误: {'; '.join(errors[:5])}" + ("..." if len(errors) > 5 else ""), 'warning')
            else:
                flash(f"成功导入 {count} 条特殊事项分配记录", "success")
        except Exception as e:
            flash(f"导入失败: {str(e)}", "danger")
        return redirect(url_for('manage_special_items') + '#batch')

    @app.route('/export_special_items_template')
    @login_required
    @admin_required
    def export_special_items_template():
        export_dir = os.path.join(current_dir, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = "特殊事项批量导入模板.xlsx"
        filepath = os.path.join(export_dir, filename)

        # 中文列名，只用姓名，不用身份证
        columns = ['员工姓名', '金额', '备注', '工作月数', '未休天数', '考核系数']
        data = [{'员工姓名': '张三', '金额': 1000, '备注': '示例', '工作月数': 12, '未休天数': 0, '考核系数': 1.0}]
        df = pd.DataFrame(data, columns=columns)
        df.to_excel(filepath, index=False)
        return send_file(filepath, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/export_quarter_assessment_template')
    @login_required
    @admin_required
    def export_quarter_assessment_template():
        template_id = request.args.get('template_id')
        quarter = request.args.get('quarter', 'Q1')  # 默认为Q1
        if not template_id:
            flash('请指定模板ID', 'danger')
            return redirect(url_for('manage_special_items'))

        template = db.session.get(SpecialItemTemplate, template_id)
        if not template:
            flash('模板不存在', 'danger')
            return redirect(url_for('manage_special_items'))

        # 季度月份映射
        quarter_months = {
            'Q1': [1, 2, 3],
            'Q2': [4, 5, 6],
            'Q3': [7, 8, 9],
            'Q4': [10, 11, 12]
        }
        months = quarter_months.get(quarter, [1, 2, 3])
        month_names = [f'{m}月考核结果' for m in months]

        export_dir = os.path.join(current_dir, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = f"{template.name}_{quarter}_考核导入模板.xlsx"
        filepath = os.path.join(export_dir, filename)

        columns = ['员工姓名'] + month_names + ['备注']
        example = {'员工姓名': '张三'}
        for mn in month_names:
            example[mn] = '好'
        example['备注'] = ''
        df = pd.DataFrame([example], columns=columns)
        df.to_excel(filepath, index=False)
        return send_file(filepath, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    def calculate_special_amount(special_item, assessment_value):
        if special_item.assessment_type == 'fixed':
            option = db.session.get(AssessmentOption, int(assessment_value))
            if option and option.formula:
                try:
                    return simple_eval(option.formula, names={'base_amount': special_item.amount})
                except:
                    return special_item.amount
            return special_item.amount
        elif special_item.assessment_type == 'formula':
            if special_item.formula:
                try:
                    return simple_eval(special_item.formula, names={'base_amount': special_item.amount, 'assessment_value': assessment_value})
                except:
                    return special_item.amount
            return special_item.amount
        else:
            return special_item.amount



    # ========== 新增：导出月度考核模板 ==========
    @app.route('/export_monthly_assessment_template')
    @login_required
    @admin_required
    def export_monthly_assessment_template():
        """导出月度考核导入模板（根据选中的特殊事项模板）"""
        template_id = request.args.get('template_id', type=int)
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)

        if not template_id:
            flash('请指定模板ID', 'danger')
            return redirect(url_for('manage_special_items'))

        template = db.session.get(SpecialItemTemplate, template_id)
        if not template:
            flash('模板不存在', 'danger')
            return redirect(url_for('manage_special_items'))

        # 解析自定义字段
        extra_fields = []
        if template.extra_fields:
            try:
                extra_fields = json.loads(template.extra_fields)
            except:
                pass

        # 获取所有在职员工（按单位排序）
        employees = Employee.query.filter_by(active_for_payroll=True).order_by(Employee.unit_id, Employee.name).all()

        # 构建列名
        columns = ['员工姓名', '单位']
        for field in extra_fields:
            columns.append(field.get('label') or field.get('name'))
        columns.extend(['考核结果', '金额(元)', '备注'])

        # 创建 DataFrame（仅表头，不填充数据）
        import pandas as pd
        df = pd.DataFrame(columns=columns)

        # 可选：添加一行示例数据
        # example_row = {col: '' for col in columns}
        # example_row['员工姓名'] = '张三'
        # example_row['考核结果'] = '好'
        # df = pd.DataFrame([example_row])

        export_dir = os.path.join(current_dir, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = f"{template.name}_{year}年{month}月考核导入模板.xlsx"
        filepath = os.path.join(export_dir, filename)
        df.to_excel(filepath, index=False)

        return send_file(filepath, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # 工资录入
    @app.route('/salary_entry', methods=['GET', 'POST'])
    @login_required
    def salary_entry():
        if request.method == 'POST':
            if not current_user.is_admin:
                flash('您没有权限修改工资数据', 'danger')
                return redirect(url_for('salary_entry'))
        today = datetime.now()
        first_of_month = today.replace(day=1)
        last_month = (first_of_month - timedelta(days=1)).strftime('%Y-%m')
        month = request.args.get('month', today.strftime('%Y-%m'))

        if request.method == 'POST':
            if 'cancel' in request.form:
                # 从 session 中获取上次工资查询的条件
                last_query = session.get('last_salary_query', {})
                query_params = []
                if last_query.get('start_month'):
                    query_params.append(f"start_month={last_query['start_month']}")
                if last_query.get('end_month'):
                    query_params.append(f"end_month={last_query['end_month']}")
                if last_query.get('unit_id'):
                    query_params.append(f"unit_id={last_query['unit_id']}")
                if last_query.get('employee_id'):
                    query_params.append(f"employee_id={last_query['employee_id']}")
                for item in last_query.get('items', []):
                    query_params.append(f"items={item}")
                redirect_url = url_for('salary_query')
                if query_params:
                    redirect_url += '?' + '&'.join(query_params)
                return redirect(redirect_url)
            current_app.logger.info("=" * 50)
            current_app.logger.info("接收到POST请求，表单数据：")
            # for key, value in request.form.items():
            #     print(f"  {key}: {value}")
            current_app.logger.info("=" * 50)

            if 'manual_entry' not in request.form:
                flash("无效的提交", "danger")
                return redirect(url_for('salary_entry'))

            employee_id = request.form.get('employee_id')
            month = request.form.get('month')
            if not month:
                month = request.args.get('month', last_month)
            if not employee_id:
                flash("员工ID不能为空", "danger")
                return redirect(url_for('salary_entry'))

            remark = request.form.get('remark', '')
            current_app.logger.info(f"准备保存: employee_id={employee_id}, month={month}, remark={remark}")

            record = SalaryRecord.query.filter_by(employee_id=employee_id, month=month).first()
            employee = db.session.get(Employee, employee_id)
            if not employee or not employee.active_for_payroll:
                flash("该员工当前不参与工资核算", "danger")
                return redirect(url_for('salary_entry'))
            if record:
                 current_app.logger.info(f"更新已有记录 ID={record.id}")
            else:
                record = SalaryRecord(employee_id=employee_id, month=month, details={}, total=0, remark=remark)
                db.session.add(record)
                current_app.logger.info(f"创建新记录")

            details = {}
            all_salary_items = SalaryItem.query.all()
            for item in all_salary_items:
                field_name = f'item_{item.id}'
                value_str = request.form.get(field_name, '0')
                try:
                    value = round(float(value_str), 2)
                except (ValueError, TypeError):
                    value = 0.0

                details[item.name.strip()] = value
                current_app.logger.info(f"系统项 {item.name}: {value}")

            temp_names = request.form.getlist('temp_item_name')
            temp_values = request.form.getlist('temp_item_value')
            for name, val_str in zip(temp_names, temp_values):
                if name and name.strip():
                    try:
                        val = float(val_str) if val_str else 0.0

                        details[name.strip()] = val
                        current_app.logger.info(f"临时项 {name}: {val}")
                    except ValueError:
                        continue

            all_items = SalaryItem.query.all()
            item_dict = {item.name: item for item in all_items}

            gross = 0.0
            deduct = 0.0
            for key, val in details.items():
                if key in ['应发工资', '扣款合计', '实发工资']:
                    continue
                item_info = SalaryItem.query.filter_by(name=key).first()
                if item_info:
                    if item_info.item_type == 'income':
                        gross += val
                    elif item_info.item_type == 'deduction':
                        deduct += val
                    elif item_info.item_type == 'calculation':
                        if getattr(item_info, 'calc_direction', 'income') == 'income':
                            gross += val
                        else:
                            deduct += val
                else:
                    # 未定义项，均计为收入（保留原符号，负数自动减少应发）
                    gross += val
            net = gross - deduct
            details['应发工资'] = round(gross, 2)
            details['扣款合计'] = round(deduct, 2)
            details['实发工资'] = round(net, 2)

            record.details = details
            record.total = round(net, 2)
            record.remark = remark
            # 新增分项备注保存
            item_remarks_str = request.form.get('item_remarks', '{}')
            try:
                item_remarks = json.loads(item_remarks_str)
            except:
                item_remarks = {}
            record.item_remarks = json.dumps(item_remarks, ensure_ascii=False)

            current_app.logger.info(f"最终保存的 details: {record.details}")
            current_app.logger.info(f"total: {record.total}")

            try:
                db.session.commit()
                current_app.logger.info("数据库提交成功")
                flash(f"工资录入成功！实发工资: {net:.2f}元", "success")
                # 获取上次查询条件
                last_query = session.get('last_salary_query', {})
                # 构造重定向 URL
                query_params = []
                if last_query.get('start_month'):
                    query_params.append(f"start_month={last_query['start_month']}")
                if last_query.get('end_month'):
                    query_params.append(f"end_month={last_query['end_month']}")
                if last_query.get('unit_id'):
                    query_params.append(f"unit_id={last_query['unit_id']}")
                if last_query.get('employee_id'):
                    query_params.append(f"employee_id={last_query['employee_id']}")
                for item in last_query.get('items', []):
                    query_params.append(f"items={item}")
                redirect_url = url_for('salary_query')
                if query_params:
                    redirect_url += '?' + '&'.join(query_params)
                return redirect(redirect_url)
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"保存工资记录失败: {str(e)}", exc_info=True)
                flash(f"保存失败: {str(e)}", "danger")
                current_app.logger.info(f"保存失败，异常: {str(e)}")
                return redirect(url_for('salary_entry', employee_id=employee_id, month=month))

        # GET 请求
        units = Unit.query.all()
        employees = Employee.query.filter_by(active_for_payroll=True).all()
        salary_items = SalaryItem.query.order_by(SalaryItem.order).all()
        # 正确处理 employee_id 参数，转为整数或 None
        selected_employee_id_raw = request.args.get('employee_id')
        selected_employee_id = None
        selected_employee = None
        existing_record = None
        existing_details = {}
        existing_remark = ''
        existing_item_remarks = {}

        if selected_employee_id_raw and selected_employee_id_raw.isdigit():
            selected_employee_id = int(selected_employee_id_raw)
            selected_employee = db.session.get(Employee, selected_employee_id)
            # 当员工存在时，查询该员工该月份的工资记录
            if selected_employee:
                existing_record = SalaryRecord.query.filter_by(
                    employee_id=selected_employee_id, month=month
                ).first()

        if existing_record:
            if isinstance(existing_record.details, str):
                try:
                    existing_record.details = json.loads(existing_record.details)
                    db.session.commit()
                except:
                    existing_record.details = {}
            if not isinstance(existing_record.details, dict):
                existing_record.details = {}
            all_item_names = [item.name for item in salary_items]
            existing_details = {k: v for k, v in existing_record.details.items() if k in all_item_names}
            existing_remark = getattr(existing_record, 'remark', '')
            if existing_record.item_remarks:
                try:
                    existing_item_remarks = json.loads(existing_record.item_remarks)
                except:
                    existing_item_remarks = {}
        salary_items_json = [
            {
                'id': item.id,
                'name': item.name,
                'item_type': item.item_type,
                'formula': item.formula or '',
                'order': item.order or 0,
                'calc_direction': getattr(item, 'calc_direction', 'income')
            }
            for item in salary_items
        ]
        return render_template('salary_entry.html',
                               units=units,
                               employees=employees,
                               salary_items=salary_items,
                               current_month=month,
                               now=datetime.now(),
                               selected_employee=selected_employee,
                               selected_employee_id=selected_employee_id,
                               existing_details=existing_details,
                               existing_remark=existing_remark,
                               existing_item_remarks=existing_item_remarks,  # 新增
                               salary_items_json=json.dumps(salary_items_json, ensure_ascii=False))

    # 工资查询
    @app.route('/salary_query', methods=['GET', 'POST'])
    @login_required
    def salary_query():
        start_month = request.args.get('start_month') or request.form.get('start_month')
        end_month = request.args.get('end_month') or request.form.get('end_month')
        unit_id = request.args.get('unit_id') or request.form.get('unit_id')
        employee_id = request.args.get('employee_id') or request.form.get('employee_id')
        items = request.args.getlist('items') or request.form.getlist('items')
        if request.method == 'GET':
            session['last_salary_query'] = {
                'start_month': start_month,
                'end_month': end_month,
                'unit_id': unit_id,
                'employee_id': employee_id,
                'items': items
            }
        if not items:
            items = [item.name for item in SalaryItem.query.all()]

        fixed_columns = ['应发工资', '扣款合计', '实发工资']
        original_items = items
        items = [name for name in items if name not in fixed_columns]
        if len(items) != len(original_items):
            flash("已自动过滤'应发工资/扣款合计/实发工资'等固定列", 'info')

        query = SalaryRecord.query
        if start_month and end_month:
            query = query.filter(SalaryRecord.month >= start_month, SalaryRecord.month <= end_month)
        elif start_month:
            query = query.filter(SalaryRecord.month >= start_month)
        elif end_month:
            query = query.filter(SalaryRecord.month <= end_month)
        if unit_id:
            query = query.join(Employee, SalaryRecord.employee_id == Employee.id).filter(Employee.unit_id == unit_id)
        if employee_id:
            query = query.filter(SalaryRecord.employee_id == employee_id)

        records = query.all()
        records = [r for r in records if r.employee is not None]

        # ========== 动态构建扣款项关键词列表 ==========
        deduction_items = SalaryItem.query.filter_by(item_type='deduction').all()
        deduction_keywords = [item.name for item in deduction_items]
        extra_keywords = ['扣', '险', '金', '税', '费', '养老', '医疗', '失业', '公积', '年金', '工会']
        deduction_keywords.extend(extra_keywords)
        # ===========================================

        all_salary_items = {item.name.strip(): item for item in SalaryItem.query.all()}
        summary_keys = ['应发工资', '扣款合计', '实发工资']
        salary_item_info = {}
        salary_item_info = {}
        for item in SalaryItem.query.all():
            name = item.name.strip()
            salary_item_info[name] = {
                'item_type': item.item_type,
                'calc_direction': getattr(item, 'calc_direction', 'income')
            }

        # ========== 遍历所有工资记录 ==========
        for record in records:
            # 解析 details
            if isinstance(record.details, str):
                try:
                    record.details = json.loads(record.details)
                except:
                    record.details = {}
            if not isinstance(record.details, dict):
                record.details = {}

            # 标准化键名（去除空格）
            normalized = {}
            for key, value in record.details.items():
                std_key = key.strip()
                normalized[std_key] = value
            record.details = normalized

            # ========== 重新计算应发、扣款、实发 ==========
            gross = 0.0
            deduct = 0.0
            for item_name, val in record.details.items():
                if item_name in ['应发工资', '扣款合计', '实发工资']:
                    continue
                try:
                    val = float(val)
                except:
                    val = 0.0

                info = salary_item_info.get(item_name)
                if info:
                    if info['item_type'] == 'income':
                        gross += val
                    elif info['item_type'] == 'deduction':
                        deduct += val  # 修复点：保留符号
                    elif info['item_type'] == 'calculation':
                        if info['calc_direction'] == 'income':
                            gross += val
                        else:
                            deduct += val
                else:
                    # 未定义项：根据关键词判断类型，但符号保留
                    if any(kw in item_name for kw in deduction_keywords):
                        deduct += val
                    else:
                        if val >= 0:
                            gross += val
                        else:
                            deduct += val

            net = gross - deduct
            record.gross_salary = round(gross, 2)
            record.deductions = round(deduct, 2)
            record.net_salary = round(net, 2)
            record.total = round(net, 2)
            record.details['应发工资'] = record.gross_salary
            record.details['扣款合计'] = record.deductions
            record.details['实发工资'] = record.net_salary

        selected_item_objs = SalaryItem.query.filter(SalaryItem.name.in_(items)).order_by(SalaryItem.order).all()
        income_selected = [item for item in selected_item_objs if item.item_type == 'income']
        deduction_selected = [item for item in selected_item_objs if item.item_type == 'deduction']
        calc_income_selected = [item for item in selected_item_objs if
                                item.item_type == 'calculation' and getattr(item, 'calc_direction',
                                                                            'income') == 'income']
        calc_deduction_selected = [item for item in selected_item_objs if
                                   item.item_type == 'calculation' and getattr(item, 'calc_direction',
                                                                               'income') == 'deduction']

        selected_items_sorted = selected_item_objs  # <-- 新增这一行
        calculation_selected = calc_income_selected + calc_deduction_selected
        item_totals = {}
        for item_name in items:
            total = 0.0
            for r in records:
                val = r.details.get(item_name, 0)
                total += val
            item_totals[item_name] = round(total, 2)

        gross_total = sum(r.gross_salary for r in records)
        deduction_total = sum(r.deductions for r in records)
        net_total = sum(r.net_salary for r in records)

        income_items = SalaryItem.query.filter_by(item_type='income').order_by(SalaryItem.order).all()
        deduction_items = SalaryItem.query.filter_by(item_type='deduction').order_by(SalaryItem.order).all()
        calculation_items = SalaryItem.query.filter_by(item_type='calculation').order_by(SalaryItem.order).all()

        units = Unit.query.all()
        employees = Employee.query.filter_by(active_for_payroll=True).all()
        all_employees = Employee.query.filter_by(active_for_payroll=True).all()

        return render_template('salary_query.html', records=records, units=units, employees=employees,
                               income_items=income_items, deduction_items=deduction_items,
                               calculation_items=calculation_items,
                               income_selected=income_selected, deduction_selected=deduction_selected,
                               calc_income_selected=calc_income_selected,
                               calc_deduction_selected=calc_deduction_selected,
                               calculation_selected=calculation_selected,
                               selected_items_sorted=selected_items_sorted,  # 新增这一行
                               selected_item_names=items, item_totals=item_totals, gross_total=gross_total,
                               deduction_total=deduction_total, net_total=net_total, start_month=start_month,
                               end_month=end_month, unit_id=unit_id, employee_id=employee_id,
                               all_employees=all_employees)

    # 工资导入（单独路由，此处保留原有逻辑）
    @app.route('/salary_import', methods=['GET', 'POST'])
    @login_required
    def salary_import():
        if request.method == 'POST':
            if not current_user.is_admin:
                flash('您没有权限导入工资数据', 'danger')
                return redirect(url_for('salary_import'))

            if 'file' not in request.files:
                flash('没有选择文件', 'danger')
                return redirect(request.url)
            file = request.files['file']
            if file.filename == '':
                flash('没有选择文件', 'danger')
                return redirect(request.url)

            try:
                df = pd.read_excel(file, dtype={'银行卡号': str})  # 员工ID可不要求字符串
                df.columns = df.columns.str.strip()
                # 必须的列：员工姓名、月份、单位
                required_columns = ['员工姓名', '月份', '单位']
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    flash(f'缺少必要列: {", ".join(missing_columns)}', 'danger')
                    return redirect(request.url)

                system_items = {item.name: item for item in SalaryItem.query.all()}
                deduction_keywords = ['养老', '年金', '失业', '医疗', '公积', '工会']
                count = 0
                errors = []
                imported_employees = set()  # 用于去重检查 (employee_id, month)

                for index, row in df.iterrows():
                    row_num = index + 2
                    try:
                        # 解析月份
                        month_raw = row.get('月份')
                        if pd.isna(month_raw):
                            errors.append(f"行 {index + 2}: 月份为空")
                            continue

                        if isinstance(month_raw, (datetime, pd.Timestamp)):
                            month = month_raw.strftime('%Y-%m')
                        else:
                            month = str(month_raw).strip()

                        if not re.match(r'^\d{4}-\d{2}$', month):
                            errors.append(f"行 {row_num}: 月份格式错误（{month}），应为YYYY-MM")
                            continue

                        # 获取姓名和单位
                        name = str(row.get('员工姓名', '')).strip() if pd.notna(row.get('员工姓名')) else ''
                        unit_name = str(row.get('单位', '')).strip() if pd.notna(row.get('单位')) else ''
                        if not name or not unit_name:
                            errors.append(f"行 {row_num}: 员工姓名或单位为空")
                            continue

                        # 查找员工（姓名+单位唯一匹配）
                        unit = Unit.query.filter_by(name=unit_name).first()
                        if not unit:
                            # 尝试模糊匹配
                            units = Unit.query.filter(Unit.name.like(f'%{unit_name}%')).all()
                            if len(units) == 1:
                                unit = units[0]
                            elif len(units) > 1:
                                errors.append(f"行 {row_num}: 单位名称 '{unit_name}' 匹配到多个单位，请使用精确名称")
                                continue
                            else:
                                errors.append(f"行 {row_num}: 单位 '{unit_name}' 不存在")
                                continue

                        employees_found = Employee.query.filter_by(name=name, unit_id=unit.id).all()
                        if len(employees_found) == 0:
                            errors.append(f"行 {row_num}: 未找到员工 '{name}' (单位: {unit.name})")
                            continue
                        elif len(employees_found) > 1:
                            errors.append(f"行 {row_num}: 单位 '{unit.name}' 中存在多名员工叫 '{name}'，请手动处理")
                            continue
                        employee = employees_found[0]

                        # 去重检查（同员工同月份只能出现一次）
                        key = (employee.id, month)
                        if key in imported_employees:
                            errors.append(f"行 {row_num}: 员工 {employee.name} 的 {month} 月数据已导入")
                            continue

                        # 获取或创建原有工资记录
                        record = SalaryRecord.query.filter_by(employee_id=employee.id, month=month).first()
                        if record:
                            old_details = record.details
                            if isinstance(old_details, str):
                                try:
                                    old_details = json.loads(old_details)
                                except:
                                    old_details = {}
                            elif old_details is None:
                                old_details = {}
                        else:
                            old_details = {}

                        for col in row.index:
                            # 跳过非工资项列
                            if col in ['员工姓名', '单位', '月份', '备注', '员工ID']:
                                continue
                            value = row[col]
                            if pd.isna(value):
                                continue  # 空单元格保留原值
                            try:
                                value = float(value)
                            except ValueError:
                                errors.append(f"行 {row_num}: 列 '{col}' 的值 '{value}' 不是有效数字，已跳过")
                                continue
                            old_details[col] = value

                        # 重新计算计算项（3轮）
                        for _ in range(3):
                            changed = False
                            for item in system_items.values():
                                if item.item_type != 'calculation' or not item.formula:
                                    continue
                                try:
                                    new_val = evaluate_formula(item.formula, old_details)
                                    new_val = round(new_val, 2)
                                    old_val = old_details.get(item.name, 0)
                                    if abs(new_val - old_val) > 0.01:
                                        old_details[item.name] = new_val
                                        changed = True
                                except Exception as e:
                                    current_app.logger.warning(f"计算公式失败: {item.name}, {e}")
                            if not changed:
                                break

                        # 重新计算应发、扣款、实发
                        gross = 0.0
                        deduct = 0.0
                        for item_name, val in old_details.items():
                            if item_name in ['应发工资', '扣款合计', '实发工资']:
                                continue
                            try:
                                val = float(val)
                            except:
                                val = 0.0

                            item_info = system_items.get(item_name)
                            if item_info:
                                if item_info.item_type == 'income':
                                    gross += val
                                elif item_info.item_type == 'deduction':
                                    deduct += val
                                elif item_info.item_type == 'calculation':
                                    calc_dir = getattr(item_info, 'calc_direction', 'income')
                                    if calc_dir == 'income':
                                        gross += val
                                    else:
                                        deduct += val
                            else:
                                if val >= 0:
                                    gross += val
                                else:
                                    deduct += val

                        net = gross - deduct
                        old_details['应发工资'] = round(gross, 2)
                        old_details['扣款合计'] = round(deduct, 2)
                        old_details['实发工资'] = round(net, 2)

                        # 读取备注（如果Excel中有备注列）
                        remark = ''
                        if '备注' in row.index and pd.notna(row['备注']):
                            remark = str(row['备注']).strip()
                        elif record:
                            remark = record.remark if record.remark else ''

                        # 保存记录
                        if record:
                            record.details = old_details
                            flag_modified(record, "details")   # ← 加入这一行
                            record.total = net
                            record.remark = remark
                        else:
                            record = SalaryRecord(
                                employee_id=employee.id,
                                month=month,
                                details=old_details,
                                total=net,
                                remark=remark,
                                item_remarks='{}'
                            )
                            db.session.add(record)

                        imported_employees.add(key)
                        count += 1

                    except Exception as e:
                        errors.append(f"行 {row_num}: 处理错误 - {str(e)}")
                        continue

                db.session.commit()

                # 统计导入的月份
                months_imported = set()
                for (emp_id, month) in imported_employees:
                    months_imported.add(month)
                month_str = ', '.join(sorted(months_imported)) if months_imported else '无'

                if errors:
                    error_msg = f'成功导入 {count} 条记录，涉及月份：{month_str}，但有 {len(errors)} 个错误'
                    if len(errors) > 5:
                        error_msg += f"(显示前5条): {'; '.join(errors[:5])}..."
                    else:
                        error_msg += f": {'; '.join(errors)}"
                    flash(error_msg, 'warning')
                else:
                    flash(f'成功导入 {count} 条工资记录，涉及月份：{month_str}', 'success')

                if imported_employees:
                    first_month = next(iter(imported_employees))[1]
                    return redirect(url_for('salary_query', start_month=first_month, end_month=first_month))
                else:
                    return redirect(url_for('salary_query'))

            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"工资导入失败: {str(e)}", exc_info=True)
                flash(f'导入失败: {str(e)}', 'danger')

        return render_template('salary_import.html')

    # 导出查询结果
    @app.route('/export_query_results')
    @login_required
    def export_query_results():
        start_month = request.args.get('start_month')
        end_month = request.args.get('end_month')
        unit_id = request.args.get('unit_id')
        employee_id = request.args.get('employee_id')
        items = request.args.getlist('items')
        fixed_columns = ['应发工资', '扣款合计', '实发工资']
        items = [name for name in items if name not in fixed_columns]
        query = SalaryRecord.query.join(Employee)
        if start_month and end_month:
            query = query.filter(SalaryRecord.month >= start_month, SalaryRecord.month <= end_month)
        elif start_month:
            query = query.filter(SalaryRecord.month >= start_month)
        elif end_month:
            query = query.filter(SalaryRecord.month <= end_month)
        if unit_id:
            query = query.filter(Employee.unit_id == unit_id)
        if employee_id:
            query = query.filter(SalaryRecord.employee_id == employee_id)
        records = query.all()
        if not records:
            flash("没有数据可导出", "warning")
            return redirect(url_for('salary_query'))

        # ========== ★ 修复：构建工资项类型映射 ★ ==========
        salary_item_info = {}
        for item in SalaryItem.query.all():
            name = item.name.strip()
            salary_item_info[name] = {
                'item_type': item.item_type,
                'calc_direction': getattr(item, 'calc_direction', 'income')
            }

        all_income_names = [item.name for item in SalaryItem.query.filter_by(item_type='income').all()]
        all_deduction_names = [item.name for item in SalaryItem.query.filter_by(item_type='deduction').all()]
        selected_item_objs = SalaryItem.query.filter(SalaryItem.name.in_(items)).order_by(SalaryItem.order).all()
        calculation_selected = [item for item in selected_item_objs if item.item_type == 'calculation']
        income_selected = [item for item in selected_item_objs if item.item_type == 'income']
        deduction_selected = [item for item in selected_item_objs if item.item_type == 'deduction']
        headers = ['月份', '员工姓名']
        headers.extend([item.name for item in income_selected])
        headers.append('应发工资')
        headers.extend([item.name for item in deduction_selected])
        headers.append('扣款合计')
        headers.append('实发工资')
        headers.append('备注')  # 添加备注列
        data = []
        total_row = {h: 0 for h in headers}
        for record in records:
            employee = record.employee
            # ========== 重新计算应发、扣款（统一符号逻辑） ==========
            gross = 0.0
            deduct = 0.0
            for item_name, val in record.details.items():
                if item_name in ['应发工资', '扣款合计', '实发工资']:
                    continue
                try:
                    val = float(val)
                except:
                    val = 0.0
                info = salary_item_info.get(item_name)
                if info:
                    if info['item_type'] == 'income':
                        gross += val
                    elif info['item_type'] == 'deduction':
                        deduct += val  # 保留符号
                    elif info['item_type'] == 'calculation':
                        if info['calc_direction'] == 'income':
                            gross += val
                        else:
                            deduct += val
                else:
                    if val >= 0:
                        gross += val
                    else:
                        deduct += val
            net = gross - deduct
            row = {'月份': record.month, '员工姓名': employee.name}
            row['备注'] = record.remark or ''  # 添加这一行
            for item in income_selected:
                val = record.details.get(item.name, 0)
                row[item.name] = round(val, 2)
                total_row[item.name] += val
            row['应发工资'] = round(gross, 2)
            total_row['应发工资'] += gross
            for item in deduction_selected:
                val = record.details.get(item.name, 0)  # 直接取原值
                row[item.name] = round(val, 2)
                total_row[item.name] += val
            row['扣款合计'] = round(deduct, 2)
            total_row['扣款合计'] += deduct
            row['实发工资'] = round(net, 2)
            total_row['实发工资'] += net
            data.append(row)
        summary_row = {h: '' for h in headers}
        summary_row['月份'] = '合计'
        for k, v in total_row.items():
            if k not in ['月份', '员工姓名']:
                summary_row[k] = round(v, 2)
        data.append(summary_row)
        export_dir = os.path.join(current_dir, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = f"工资查询结果_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        filepath = os.path.join(export_dir, filename)
        df = pd.DataFrame(data, columns=headers)

        # 获取所有工资项名称
        salary_item_names = {item.name for item in SalaryItem.query.all()}
        fixed_number_cols = {'应发工资', '扣款合计', '实发工资'}
        number_cols = salary_item_names.union(fixed_number_cols)

        for col in df.columns:
            if col in number_cols:
                df[col] = pd.to_numeric(df[col], errors='coerce').round(2)

        df.to_excel(filepath, index=False, float_format='%.2f')
        return send_file(filepath, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/api/get_assessments', methods=['GET'])
    @login_required
    def get_assessments():
        import traceback
        try:
            template_id = request.args.get('template_id', type=int)
            year = request.args.get('year', type=int)
            month_param = request.args.get('month')
            unit_id = request.args.get('unit_id', type=int)

            if not template_id or not year:
                return jsonify([])

            # 使用显式 filter 而非 filter_by 以避免实体歧义
            query = EmployeeSpecialGrant.query.filter(
                EmployeeSpecialGrant.template_id == template_id,
                EmployeeSpecialGrant.year == year
            )
            # 只取启用年度事项的员工记录
            query = query.join(Employee).filter(Employee.active_for_annual_special == True)

            if month_param:
                try:
                    month_int = int(month_param)
                except ValueError:
                    if '-' in month_param:
                        month_int = int(month_param.split('-')[1])
                    else:
                        month_int = None
                if month_int is not None:
                    # 修改点：显式指定 EmployeeSpecialGrant.month
                    query = query.filter(EmployeeSpecialGrant.month == month_int)

            if unit_id:
                query = query.filter(Employee.unit_id == unit_id)

            grants = query.all()
            result = []
            for g in grants:
                extra = json.loads(g.extra_data or '{}')
                result.append({
                    'employee_id': g.employee_id,
                    'assessment_value': extra.get('assessment_value', ''),
                    'amount': g.amount,
                    'remark': g.remark,
                    'month': f"{g.year}-{str(g.month).zfill(2)}" if g.month else None,
                    'extra_data': extra
                })
            return jsonify(result)
        except Exception as e:
            current_app.logger.error(f"get_assessments 错误: {e}\n{traceback.format_exc()}")
            return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

    @app.route('/api/import_quarter_assessments', methods=['POST'])
    @login_required
    @admin_required
    def import_quarter_assessments():
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '未上传文件'}), 400
        file = request.files['file']
        template_id = request.form.get('template_id')
        year = int(request.form.get('year'))
        quarter = request.form.get('quarter')

        if not template_id:
            return jsonify({'success': False, 'error': '缺少模板ID'}), 400

        template = db.session.get(SpecialItemTemplate, template_id)
        if not template:
            return jsonify({'success': False, 'error': '模板不存在'}), 404

        quarter_months = {
            'Q1': [1, 2, 3],
            'Q2': [4, 5, 6],
            'Q3': [7, 8, 9],
            'Q4': [10, 11, 12]
        }
        months = quarter_months.get(quarter)
        if not months:
            return jsonify({'success': False, 'error': '无效的季度'}), 400

        try:
            df = pd.read_excel(file)
            df.columns = df.columns.str.strip()
        except Exception as e:
            return jsonify({'success': False, 'error': f'读取Excel失败: {str(e)}'}), 400

        # 根据实际列名动态匹配，不再固定为1/2/3月
        month_cols = []
        for m in months:
            col_name = f'{m}月考核结果'
            if col_name in df.columns:
                month_cols.append(col_name)
            else:
                # 尝试其他可能的列名（兼容性）
                found = False
                for col in df.columns:
                    if f'{m}月' in col and '考核' in col:
                        month_cols.append(col)
                        found = True
                        break
                if not found:
                    return jsonify({'success': False, 'error': f'缺少必要列: {col_name}'}), 400

        required_cols = ['员工姓名'] + month_cols
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            return jsonify({'success': False, 'error': f'缺少必要列: {", ".join(missing)}'}), 400

        count = 0
        overwritten = 0
        errors = []
        for idx, row in df.iterrows():
            row_num = idx + 2
            try:
                name = str(row.get('员工姓名', '')).strip() if pd.notna(row.get('员工姓名')) else ''
                if not name:
                    errors.append(f"行{row_num}: 员工姓名为空")
                    continue

                employees = Employee.query.filter_by(name=name, active_for_payroll=True).all()
                if len(employees) == 0:
                    errors.append(f"行{row_num}: 未找到员工 '{name}'")
                    continue
                elif len(employees) > 1:
                    errors.append(f"行{row_num}: 存在多个同名员工 '{name}'，请手动处理")
                    continue
                employee = employees[0]

                remark = str(row.get('备注', '')) if pd.notna(row.get('备注')) else ''

                for i, month in enumerate(months):
                    col = month_cols[i]
                    assess_val = str(row.get(col, '')).strip() if pd.notna(row.get(col)) else ''
                    if not assess_val:
                        continue

                    extra = {'assessment_value': assess_val}
                    if template.calculation_type == 'fixed':
                        amount = template.base_amount
                    elif template.calculation_type == 'formula' and template.formula:
                        vars_with_base = {'base_amount': template.base_amount, '考核': assess_val, **extra}
                        amount = evaluate_formula_with_context(
                            template.formula, employee.id, year, vars_with_base,
                            salary_item_names=json.loads(template.salary_items) if template.salary_items else None
                        )
                    else:
                        amount = 0.0

                    grant = EmployeeSpecialGrant.query.filter_by(
                        employee_id=employee.id,
                        template_id=template.id,
                        year=year,
                        month=month
                    ).first()

                    if grant:
                        # 记录覆盖
                        overwritten += 1
                        grant.amount = amount
                        grant.remark = remark
                        old_extra = json.loads(grant.extra_data or '{}')
                        old_extra.update(extra)
                        grant.extra_data = json.dumps(old_extra, ensure_ascii=False)
                    else:
                        grant = EmployeeSpecialGrant(
                            employee_id=employee.id,
                            template_id=template.id,
                            year=year,
                            month=month,
                            amount=amount,
                            remark=remark,
                            extra_data=json.dumps(extra, ensure_ascii=False),
                            grant_date=datetime.now().date()
                        )
                        db.session.add(grant)
                    count += 1

            except Exception as e:
                errors.append(f"行{row_num}: 处理错误 - {str(e)}")

        try:
            db.session.commit()
            message = f'成功导入 {count} 条记录。'
            if overwritten > 0:
                message += f' 其中覆盖已有记录 {overwritten} 条。'
            return jsonify(
                {'success': True, 'count': count, 'overwritten': overwritten, 'errors': errors, 'message': message})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/api/special_template_field_usage', methods=['GET'])
    @login_required
    @admin_required
    def check_field_usage():
        template_id = request.args.get('template_id', type=int)
        field_name = request.args.get('field_name')
        if not template_id or template_id <= 0 or not field_name:
            return jsonify({'error': '缺少有效参数'}), 400

        # 确保模板存在
        template = SpecialItemTemplate.query.get(template_id)
        if not template:
            return jsonify({'error': '模板不存在'}), 404

        grants = EmployeeSpecialGrant.query.filter_by(template_id=template_id).all()
        used = False
        count = 0
        for g in grants:
            if g.extra_data:
                try:
                    extra = json.loads(g.extra_data)
                    if field_name in extra:
                        used = True
                        count += 1
                        if count >= 10:  # 最多统计10条，避免性能问题
                            break
                except:
                    pass
        return jsonify({'used': used, 'count': count})
    @app.route('/api/save_assessments', methods=['POST'])
    @login_required
    @admin_required
    def save_assessments():
        data = request.get_json()
        template_id = data.get('template_id')
        year = data.get('year')
        assessments = data.get('assessments', [])

        template = db.session.get(SpecialItemTemplate, template_id)
        if not template:
            return jsonify({'success': False, 'error': '模板不存在'}), 404

        salary_items_list = None
        if template.salary_items:
            try:
                salary_items_list = json.loads(template.salary_items)
            except:
                pass

        saved_count = 0
        errors = []
        for item in assessments:
            try:
                emp_id = item['employee_id']
                emp = db.session.get(Employee, emp_id)
                if not emp:
                    errors.append(f"员工ID {emp_id} 不存在，跳过")
                    continue
                if not emp.active_for_annual_special:
                    errors.append(f"员工 {emp.name} 已停用年度事项，跳过保存")
                    continue

                assessment_val = str(item.get('assessment_value', '')).strip()
                remark = item.get('remark', '')
                month_str = item.get('month')
                extra = item.get('extra_data', {})
                if isinstance(extra, str):
                    try:
                        extra = json.loads(extra)
                    except:
                        extra = {}
                if not isinstance(extra, dict):
                    extra = {}

                def convert_numeric_in_dict(d):
                    for k, v in list(d.items()):
                        if isinstance(v, str) and v.replace('.', '', 1).isdigit():
                            d[k] = float(v) if '.' in v else int(v)
                    return d

                extra = convert_numeric_in_dict(extra)

                if not month_str:
                    errors.append(f"员工ID {emp_id}: 缺少月份")
                    continue

                if len(month_str) == 4:
                    year_val = int(month_str)
                    month_val = None
                else:
                    try:
                        year_val = int(month_str[:4])
                        month_val = int(month_str[5:7])
                    except:
                        errors.append(f"员工ID {emp_id}: 月份格式错误")
                        continue

                context_vars = {
                    'employee_id': emp_id,
                    'year': year_val,
                    'base_amount': template.base_amount,
                    '考核': assessment_val,
                }
                if month_val is not None:
                    context_vars['month'] = month_val

                extra_field_names = [f['name'] for f in json.loads(template.extra_fields or '[]')]
                filtered_extra = {k: v for k, v in extra.items() if k in extra_field_names}
                try:
                    full_extra_data = evaluate_extra_fields(template, filtered_extra, context_vars)
                except Exception as e:
                    errors.append(f"员工ID {emp_id}: 自定义字段计算失败 - {str(e)}")
                    continue

                # 如果前端传递了amount，优先使用；否则重新计算
                amount_from_request = item.get('amount')
                if amount_from_request is not None:
                    amount = float(amount_from_request)
                else:
                    if template.calculation_type == 'fixed':
                        amount = template.base_amount
                    elif template.calculation_type == 'formula' and template.formula:
                        all_vars = context_vars.copy()
                        all_vars.update(full_extra_data)
                        amount = evaluate_formula_with_context(
                            template.formula,
                            emp_id,
                            year_val,
                            all_vars,
                            salary_item_names=salary_items_list,
                            month=month_val
                        )
                    else:
                        amount = 0.0

                grant = EmployeeSpecialGrant.query.filter_by(
                    employee_id=emp_id,
                    template_id=template_id,
                    year=year_val,
                    month=month_val
                ).first()

                full_extra_data['assessment_value'] = assessment_val

                if not grant:
                    grant = EmployeeSpecialGrant(
                        employee_id=emp_id,
                        template_id=template_id,
                        year=year_val,
                        month=month_val,
                        amount=amount,
                        remark=remark,
                        extra_data=json.dumps(full_extra_data, ensure_ascii=False),
                        grant_date=datetime.now().date()
                    )
                    db.session.add(grant)
                else:
                    grant.amount = amount
                    grant.remark = remark
                    old_extra = json.loads(grant.extra_data or '{}')
                    old_extra.update(full_extra_data)
                    grant.extra_data = json.dumps(old_extra, ensure_ascii=False)

                saved_count += 1
            except Exception as e:
                errors.append(f"员工ID {item.get('employee_id', '未知')}: {str(e)}")
                current_app.logger.error(f"保存考核失败: {e}", exc_info=True)

        try:
            db.session.commit()
            return jsonify({'success': True, 'message': f'成功保存 {saved_count} 条记录', 'errors': errors})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/api/preview_amount', methods=['POST'])
    @login_required
    def preview_amount():
        data = request.get_json()
        template_id = data.get('template_id')
        formula = data.get('formula')
        assessment_value = data.get('assessment_value', '')
        employee_id = data.get('employee_id')
        year = data.get('year')
        month = data.get('month')  # 新增：接收月份参数
        extra_vars = data.get('extra_vars', {})

        def convert_numeric(value):
            if isinstance(value, str):
                cleaned = value.strip().strip('()').replace(',', '').replace('，', '').strip()
                try:
                    return float(cleaned)
                except ValueError:
                    pass
            return value

        # 如果没有模板ID，按临时公式测试
        if not template_id and formula:
            if not formula or formula.strip() == '':
                return jsonify({'amount': 0.0})
            variables = {
                '考核': assessment_value,
                **{k: convert_numeric(v) for k, v in extra_vars.items()},
                'base_amount': 0
            }
            amount = evaluate_formula(formula, variables)
            return jsonify({'amount': round(amount, 2)})

        template = db.session.get(SpecialItemTemplate, template_id)
        if not template:
            return jsonify({'amount': 0.0})

        context_vars = {
            'employee_id': employee_id,
            'year': year,
            'base_amount': template.base_amount,
            '考核': assessment_value,
        }
        if month is not None:
            context_vars['month'] = month
        for k, v in extra_vars.items():
            context_vars[k] = convert_numeric(v)

        # 计算自定义字段（含计算字段）
        try:
            extra_field_names = [f['name'] for f in json.loads(template.extra_fields or '[]')]
            input_for_extra = {k: v for k, v in context_vars.items() if k in extra_field_names}
            full_extra_data = evaluate_extra_fields(template, input_for_extra, context_vars)
        except Exception as e:
            current_app.logger.error(f"预览计算自定义字段失败: {e}")
            return jsonify({'amount': 0.0, 'error': str(e)}), 400

        all_vars = context_vars.copy()
        all_vars.update(full_extra_data)

        # 计算最终金额
        if template.calculation_type == 'fixed':
            amount = template.base_amount
            return jsonify({'amount': round(amount, 2), 'extra_data': full_extra_data})

        if template.calculation_type == 'formula' and template.formula:
            salary_items_list = None
            if template.salary_items:
                try:
                    salary_items_list = json.loads(template.salary_items)
                except:
                    pass

            if employee_id and year:
                amount = evaluate_formula_with_context(
                    template.formula,
                    employee_id,
                    year,
                    all_vars,
                    salary_item_names=salary_items_list,
                    month=month  # 传递月份
                )
            else:
                amount = evaluate_formula(template.formula, all_vars)

            return jsonify({'amount': round(amount, 2), 'extra_data': full_extra_data})

        return jsonify({'amount': 0.0, 'extra_data': full_extra_data})

    @app.route('/api/delete_grant/<int:grant_id>', methods=['DELETE'])
    @login_required
    @admin_required
    def delete_grant(grant_id):
        grant = db.session.get(EmployeeSpecialGrant, grant_id)
        if not grant:
            return jsonify({'error': '记录不存在'}), 404
        db.session.delete(grant)
        db.session.commit()
        return jsonify({'success': True})

    # 其他功能路由（报告生成、系统设置等）
    @app.route('/report_generation')
    @login_required
    def report_generation():
        salary_items = SalaryItem.query.order_by(SalaryItem.order).all()
        special_templates = SpecialItemTemplate.query.filter_by(is_active=True).all()
        units = Unit.query.all()
        current_month = datetime.now().strftime('%Y-%m')
        return render_template('report_generation.html',
                               salary_items=salary_items,
                               special_templates=special_templates,
                               units=units,
                               current_month=current_month)

    @app.route('/system_settings', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def system_settings():
        from models import EmailConfig
        config = EmailConfig.query.first()
        if not config:
            config = EmailConfig()
            db.session.add(config)
            db.session.commit()
        if request.method == 'POST':
            config.smtp_server = request.form.get('smtp_server', 'smtp.qq.com')
            config.smtp_port = int(request.form.get('smtp_port', 465))
            config.sender_email = request.form.get('sender_email')
            config.sender_password = request.form.get('sender_password')
            config.use_ssl = 'use_ssl' in request.form
            config.send_interval = int(request.form.get('send_interval', 10))
            db.session.commit()
            flash('邮件配置已保存', 'success')
            return redirect(url_for('system_settings'))
        return render_template('system_settings.html', config=config)

    @app.route('/help')
    @login_required
    def help_page():
        """使用帮助页面"""
        return render_template('help.html')
    @app.route('/api/clear_all_grants', methods=['POST'])
    @login_required
    @admin_required
    def clear_all_grants():
        data = request.get_json()
        template_id = data.get('template_id')
        year = data.get('year')
        unit_id = data.get('unit_id')
        employee_type = data.get('employee_type')
        is_veteran = data.get('is_veteran')

        if not template_id or not year:
            return jsonify({'success': False, 'error': '缺少模板ID或年份'}), 400

        try:
            sql = """
                DELETE FROM employee_special_grant 
                WHERE template_id = :tid AND year = :yr AND month IS NULL
            """
            params = {'tid': template_id, 'yr': int(year)}

            # 构建 employee 子查询的额外条件
            emp_conditions = []
            # 新增：只删除启用年度事项的员工记录
            emp_conditions.append("active_for_annual_special = 1")

            if unit_id:
                emp_conditions.append("unit_id = :uid")
                params['uid'] = unit_id
            if employee_type:
                if isinstance(employee_type, list):
                    if employee_type:
                        placeholders = ','.join([f':etype_{i}' for i in range(len(employee_type))])
                        emp_conditions.append(f"employee_type IN ({placeholders})")
                        for i, et in enumerate(employee_type):
                            params[f'etype_{i}'] = et
                else:
                    emp_conditions.append("employee_type = :etype")
                    params['etype'] = employee_type
            if is_veteran is not None:
                emp_conditions.append("is_veteran = :veteran")
                params['veteran'] = is_veteran

            if emp_conditions:
                sql += " AND employee_id IN (SELECT id FROM employee WHERE " + " AND ".join(emp_conditions) + ")"

            result = db.session.execute(text(sql), params)
            db.session.commit()
            count = result.rowcount
            return jsonify({'success': True, 'deleted': count})
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"清空失败: {str(e)}")
            return jsonify({'success': False, 'error': f'数据库错误: {str(e)}'}), 500

    @app.route('/api/copy_last_year_special', methods=['POST'])
    @login_required
    @admin_required
    def copy_last_year_special():
        data = request.get_json()
        template_id = data.get('template_id')
        source_year = data.get('source_year')
        target_year = data.get('target_year')
        unit_id = data.get('unit_id')
        employee_types = data.get('employee_type', [])
        veteran_only = data.get('veteran_only', False)

        if not template_id or not source_year or not target_year:
            return jsonify({'success': False, 'error': '缺少参数'}), 400

        # 获取目标年启用员工的 ID 列表
        active_emp_ids = db.session.query(Employee.id).filter(
            Employee.active_for_annual_special == True
        ).subquery()

        query = EmployeeSpecialGrant.query.filter_by(template_id=template_id, year=source_year)
        query = query.join(Employee).filter(Employee.id.in_(active_emp_ids))

        if unit_id:
            query = query.filter(Employee.unit_id == unit_id)
        if employee_types:
            if isinstance(employee_types, list) and employee_types:
                query = query.filter(Employee.employee_type.in_(employee_types))
            elif isinstance(employee_types, str) and employee_types:
                query = query.filter(Employee.employee_type == employee_types)
        if veteran_only:
            query = query.filter(Employee.is_veteran == True)

        grants = query.all()
        if not grants:
            return jsonify({'success': False, 'error': '未找到上年数据'}), 404

        count = 0
        for g in grants:
            exists = EmployeeSpecialGrant.query.filter_by(
                employee_id=g.employee_id,
                template_id=template_id,
                year=target_year,
                month=g.month
            ).first()
            if exists:
                continue
            new = EmployeeSpecialGrant(
                employee_id=g.employee_id,
                template_id=g.template_id,
                year=target_year,
                month=g.month,
                amount=g.amount,
                remark=g.remark,
                extra_data=g.extra_data,
                grant_date=datetime.now().date()
            )
            db.session.add(new)
            count += 1
        db.session.commit()
        return jsonify({'success': True, 'copied': count, 'skipped': len(grants) - count})
    # ==================== 用户管理 API ====================
    @app.route('/api/users')
    @login_required
    @admin_required
    def api_users():
        """获取所有用户列表（仅管理员）"""
        users = User.query.all()
        return jsonify([{
            'id': u.id,
            'username': u.username,
            'is_admin': u.is_admin
        } for u in users])

    @app.route('/api/users', methods=['POST'])
    @login_required
    @admin_required
    def create_user():
        """创建新用户"""
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        is_admin = data.get('is_admin', False)

        if not username or not password:
            return jsonify({'success': False, 'error': '用户名和密码不能为空'}), 400

        if User.query.filter_by(username=username).first():
            return jsonify({'success': False, 'error': '用户名已存在'}), 400

        user = User(username=username, is_admin=is_admin)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        return jsonify({'success': True, 'id': user.id})

    @app.route('/api/users/<int:user_id>', methods=['DELETE'])
    @login_required
    @admin_required
    def delete_user(user_id):
        """删除用户（不能删除自己）"""
        user = User.query.get_or_404(user_id)
        if user.id == current_user.id:
            return jsonify({'success': False, 'error': '不能删除当前登录的管理员账户'}), 400
        db.session.delete(user)
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/users/<int:user_id>/toggle_admin', methods=['POST'])
    @login_required
    @admin_required
    def toggle_admin(user_id):
        """切换用户的管理员状态"""
        user = User.query.get_or_404(user_id)
        if user.id == current_user.id:
            # 不允许自己降级（防止误操作把自己踢出管理员）
            return jsonify({'success': False, 'error': '不能修改自己的管理员权限'}), 400
        user.is_admin = not user.is_admin
        db.session.commit()
        return jsonify({'success': True, 'is_admin': user.is_admin})

    @app.route('/api/users/<int:user_id>/reset_password', methods=['POST'])
    @login_required
    @admin_required
    def reset_user_password(user_id):
        """重置用户密码"""
        user = User.query.get_or_404(user_id)
        data = request.get_json()
        new_password = data.get('new_password')
        if not new_password:
            return jsonify({'success': False, 'error': '新密码不能为空'}), 400
        user.set_password(new_password)
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/print_templates')
    @login_required
    @admin_required
    def manage_print_templates():
        """打印模板管理页面"""
        from models import PrintTemplate
        templates = PrintTemplate.query.order_by(PrintTemplate.order, PrintTemplate.name).all()
        return render_template('print_templates.html', templates=templates)

    @app.route('/api/print_templates', methods=['GET', 'POST'])
    @login_required
    def api_print_templates():
        from models import PrintTemplate

        # 新增权限检查：POST 请求必须管理员
        if request.method == 'POST' and not current_user.is_admin:
            abort(403)

        if request.method == 'GET':
            template_type = request.args.get('type')
            query = PrintTemplate.query
            if template_type:
                query = query.filter_by(template_type=template_type)
            templates = query.order_by(PrintTemplate.order, PrintTemplate.name).all()
            return jsonify([{
                'id': t.id,
                'name': t.name,
                'template_type': t.template_type,
                'description': t.description,
                'config': json.loads(t.config) if t.config else {},
                'is_default': t.is_default,
                # 新增字段
                'updated_at': t.updated_at.strftime('%Y-%m-%d %H:%M') if t.updated_at else '',
                'updated_by': t.updater.username if t.updater else ''
            } for t in templates])

        elif request.method == 'POST':
            data = request.get_json()
            template = PrintTemplate(
                name=data['name'],
                template_type=data['template_type'],
                description=data.get('description', ''),
                config=json.dumps(data['config'], ensure_ascii=False),
                is_default=data.get('is_default', False),
                created_by=current_user.id
            )
            db.session.add(template)
            db.session.commit()
            return jsonify({'success': True, 'id': template.id})

    @app.route('/api/print_templates/<int:template_id>', methods=['PUT', 'DELETE'])
    @login_required
    @admin_required
    def api_print_template_detail(template_id):
        from models import PrintTemplate
        template = PrintTemplate.query.get_or_404(template_id)

        if request.method == 'PUT':
            data = request.get_json()
            template.name = data.get('name', template.name)
            template.template_type = data.get('template_type', template.template_type)
            template.description = data.get('description', template.description)
            if 'config' in data:
                template.config = json.dumps(data['config'], ensure_ascii=False)
            template.is_default = data.get('is_default', template.is_default)

            # ===== 新增：记录修改人和时间 =====
            template.updated_by = current_user.id
            # updated_at 会由数据库的 onupdate 自动更新，无需手动赋值

            db.session.commit()
            return jsonify({'success': True})

        elif request.method == 'DELETE':
            db.session.delete(template)
            db.session.commit()
            return jsonify({'success': True})

    @app.route('/api/print_templates/<int:template_id>/set_default', methods=['POST'])
    @login_required
    @admin_required
    def set_default_template(template_id):
        """设为默认模板"""
        from models import PrintTemplate
        template = PrintTemplate.query.get_or_404(template_id)

        # 将同类型的其他模板取消默认
        PrintTemplate.query.filter_by(template_type=template.template_type).update({'is_default': False})
        template.is_default = True
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/print_templates/update_order', methods=['POST'])
    @login_required
    @admin_required
    def update_print_template_order():
        from models import PrintTemplate  # ← 添加这一行导入
        data = request.get_json()
        order_ids = data.get('order_ids', [])
        if not order_ids:
            return jsonify({'success': False, 'error': '缺少排序数据'}), 400

        try:
            for idx, tpl_id in enumerate(order_ids):
                template = db.session.get(PrintTemplate, int(tpl_id))
                if template:
                    template.order = idx
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/api/special_templates/update_order', methods=['POST'])
    @login_required
    @admin_required
    def update_special_template_order():
        data = request.get_json()
        order_ids = data.get('order_ids', [])
        if not order_ids:
            return jsonify({'success': False, 'error': '缺少排序数据'}), 400
        try:
            for idx, tpl_id in enumerate(order_ids):
                template = db.session.get(SpecialItemTemplate, int(tpl_id))
                if template:
                    template.order = idx
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    @app.route('/enhanced_print')
    @login_required
    def enhanced_print():
        """增强打印配置页面"""
        return render_template('enhanced_print.html')

    @app.route('/api/printable_columns')
    @login_required
    def api_printable_columns():
        """获取指定打印模板的所有可用列（系统列 + 工资项/特殊事项列）"""
        template_id = request.args.get('template_id', type=int)
        if not template_id:
            return jsonify({'error': '缺少模板ID'}), 400

        from models import PrintTemplate, SalaryItem, SpecialItemTemplate
        tpl = db.session.get(PrintTemplate, template_id)
        if not tpl:
            return jsonify({'error': '模板不存在'}), 404

        config = json.loads(tpl.config) if tpl.config else {}
        columns = config.get('columns', [])

        # 根据模板类型补充动态列（工资项或特殊事项的自定义字段）
        if tpl.template_type == 'salary_sheet':
            all_items = SalaryItem.query.order_by(SalaryItem.order).all()
            salary_columns = [item.name for item in all_items]
            all_cols = list(dict.fromkeys(
                columns + ['序号', '姓名', '单位', '身份证号', '应发工资', '扣款合计', '实发工资',
                           '备注'] + salary_columns))
        elif tpl.template_type == 'special_item':
            special_tpl_id = config.get('special_template_id')
            extra_fields = []
            if special_tpl_id:
                special_tpl = db.session.get(SpecialItemTemplate, special_tpl_id)
                if special_tpl and special_tpl.extra_fields:
                    try:
                        extra_fields = json.loads(special_tpl.extra_fields)
                    except:
                        pass
            special_cols = [f.get('name') for f in extra_fields if f.get('name')]
            all_cols = list(
                dict.fromkeys(columns + ['序号', '姓名', '单位', '身份证号', '金额(元)', '备注'] + special_cols))
        else:
            all_cols = columns

        return jsonify({
            'columns': all_cols,
            'default_columns': columns
        })

    @app.route('/api/generate_print', methods=['POST'])
    @login_required
    def api_generate_print():
        """根据用户自定义配置生成打印数据（HTML）"""
        data = request.get_json()
        template_id = data.get('template_id')
        unit_id = data.get('unit_id')
        month = data.get('month')
        year = data.get('year')
        quarter = data.get('quarter')
        selected_columns = data.get('selected_columns', [])
        format_options = data.get('format_options', {})

        if not template_id:
            return jsonify({'error': '缺少模板ID'}), 400

        from models import PrintTemplate, Unit
        tpl = db.session.get(PrintTemplate, template_id)
        if not tpl:
            return jsonify({'error': '模板不存在'}), 404

        unit_name = ''
        if unit_id:
            unit = db.session.get(Unit, unit_id)
            unit_name = unit.name if unit else ''

        # 调用内部预览函数（原文件中已存在 _get_preview_data）
        preview_data = _get_preview_data(
            tpl.template_type, template_id, month, year, quarter, unit_id, []
        )
        if preview_data.get('error'):
            return jsonify({'error': preview_data['error']}), 400

        rows = preview_data.get('rows', [])
        meta = preview_data.get('meta', {})

        if not selected_columns:
            selected_columns = preview_data.get('columns', [])

        # 构建表头和数据行
        headers = []
        for col in selected_columns:
            display = preview_data.get('column_display', {}).get(col, col)
            headers.append({'key': col, 'label': display})

        filtered_rows = []
        for row in rows:
            new_row = {}
            for col in selected_columns:
                new_row[col] = row.get(col, '')
            filtered_rows.append(new_row)

        total_row = None
        if format_options.get('show_total_row', True) and rows:
            total_row = {}
            for col in selected_columns:
                if col in ['序号', '姓名', '单位', '备注', '身份证号']:
                    total_row[col] = ''
                else:
                    total = sum(float(r.get(col, 0) or 0) for r in rows)
                    total_row[col] = total

        html = _build_print_html(
            headers=headers,
            rows=filtered_rows,
            total_row=total_row,
            title=format_options.get('title', ''),
            subtitle=format_options.get('subtitle', ''),
            show_unit=format_options.get('show_unit', True),
            show_date=format_options.get('show_date', True),
            unit_name=unit_name,
            month=month,
            year=year,
            quarter=quarter,
            template_name=tpl.name,
            page_margin=format_options.get('page_margin', '20px'),
            font_size=format_options.get('font_size', '12px')
        )

        return jsonify({'html': html, 'meta': meta})

    def _build_print_html(headers, rows, total_row, title, subtitle, show_unit, show_date,
                          unit_name, month, year, quarter, template_name, page_margin, font_size):
        """构建打印 HTML"""
        if not title:
            if quarter:
                title = f"{unit_name} {year}年{quarter} {template_name}"
            elif month:
                title = f"{unit_name} {month}月 {template_name}"
            else:
                title = f"{unit_name} {year}年 {template_name}"

        subtitle_html = f"<p>{subtitle}</p>" if subtitle else ""
        info_parts = []
        if show_unit and unit_name:
            info_parts.append(f"单位：{unit_name}")
        if show_date:
            now = datetime.now().strftime("%Y年%m月%d日")
            info_parts.append(f"打印日期：{now}")
        info_html = f"<p class='print-info'>{' | '.join(info_parts)}</p>" if info_parts else ""

        thead_html = "<thead><tr>" + "".join(f"<th>{h['label']}</th>" for h in headers) + "</tr></thead>"

        tbody_html = ""
        for idx, row in enumerate(rows, 1):
            tbody_html += "<tr>"
            for h in headers:
                val = row.get(h['key'], '')
                if isinstance(val, (int, float)) or (
                        isinstance(val, str) and val.replace('.', '').replace('-', '').isdigit()):
                    tbody_html += f"<td style='text-align:right'>{val}</td>"
                else:
                    tbody_html += f"<td>{val}</td>"
            tbody_html += "</tr>"

        tfoot_html = ""
        if total_row:
            tfoot_html = "<tfoot><tr>"
            for h in headers:
                val = total_row.get(h['key'], '')
                if val == '':
                    tfoot_html += "<td></td>"
                else:
                    tfoot_html += f"<td style='text-align:right; font-weight:bold'>{val:,.2f}</td>"
            tfoot_html += "</tr></tfoot>"

        html = f"""<!DOCTYPE html>
    <html>
    <head>
    <meta charset="UTF-8">
    <title>{title}</title>
    <style>
        body {{
            margin: {page_margin};
            font-family: '微软雅黑', '宋体', Arial, sans-serif;
            font-size: {font_size};
        }}
        .print-container {{
            max-width: 1200px;
            margin: 0 auto;
        }}
        .print-header {{
            text-align: center;
            margin-bottom: 20px;
        }}
        .print-title {{
            font-size: 18px;
            font-weight: bold;
            margin: 10px 0;
        }}
        .print-subtitle {{
            font-size: 14px;
            color: #555;
        }}
        .print-info {{
            font-size: 12px;
            color: #666;
            margin: 8px 0;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 6px 8px;
            vertical-align: middle;
        }}
        th {{
            background-color: #f2f2f2;
            font-weight: bold;
            text-align: center;
        }}
        tfoot td {{
            background-color: #f9f9f9;
        }}
        @media print {{
            body {{
                margin: 0;
                padding: {page_margin};
            }}
            .no-print {{
                display: none;
            }}
        }}
    </style>
    </head>
    <body>
    <div class="print-container">
        <div class="print-header">
            <div class="print-title">{title}</div>
            {subtitle_html}
            {info_html}
        </div>
        <table>
            {thead_html}
            <tbody>
                {tbody_html}
            </tbody>
            {tfoot_html}
        </table>
    </div>
    </body>
    </html>"""
        return html
    @app.route('/api/print_preview/<template_type>', methods=['GET'])
    @login_required
    def print_preview_data(template_type):
        """获取打印预览数据（支持工资发放表和特殊事项发放表）"""
        from models import PrintTemplate, SalaryRecord, Employee, Unit, SalaryItem, SpecialItemTemplate, \
            EmployeeSpecialGrant

        template_id = request.args.get('template_id')
        month = request.args.get('month')  # 格式: YYYY-MM
        year = request.args.get('year')  # 可选，用于年度特殊事项
        unit_id = request.args.get('unit_id')
        employee_id = request.args.get('employee_id')
        empty_mode = request.args.get('empty_mode', 'false').lower() == 'true'

        # 获取打印模板配置
        if template_id:
            template = db.session.get(PrintTemplate, template_id)
            if not template:
                return jsonify({'error': '模板不存在'}), 404
            config = json.loads(template.config) if template.config else {}
        else:
            config = get_default_template_config(template_type)

        # ========== 1. 工资发放表 ==========
        if template_type == 'salary_sheet':
            query = db.session.query(SalaryRecord, Employee, Unit).join(
                Employee, SalaryRecord.employee_id == Employee.id
            ).join(Unit, Employee.unit_id == Unit.id)

            if month:
                query = query.filter(SalaryRecord.month == month)
            if unit_id:
                query = query.filter(Employee.unit_id == unit_id)
            employee_ids = request.args.getlist('employee_ids')
            if employee_ids:
                query = query.filter(SalaryRecord.employee_id.in_(employee_ids))
            elif employee_id:
                query = query.filter(SalaryRecord.employee_id == employee_id)

            results = query.all()

            all_salary_items = SalaryItem.query.all()
            item_dict = {item.name: item for item in all_salary_items}

            display_columns = config.get('columns', [])
            if not display_columns:
                display_columns = [item.name for item in all_salary_items] + ['应发工资', '扣款合计', '实发工资',
                                                                              '备注']

            if empty_mode:
                month = request.args.get('month')
                unit_id_query = request.args.get('unit_id')
                emp_ids_param = request.args.getlist('employee_ids')

                if emp_ids_param:
                    employees = Employee.query.filter(Employee.id.in_(emp_ids_param)).all()
                else:
                    employee_ids_from_records = set()
                    if month:
                        rec_query = db.session.query(SalaryRecord.employee_id).filter(SalaryRecord.month == month)
                        if unit_id_query:
                            rec_query = rec_query.join(Employee, SalaryRecord.employee_id == Employee.id).filter(
                                Employee.unit_id == unit_id_query)
                        employee_ids_from_records = {r[0] for r in rec_query.all()}
                    if employee_ids_from_records:
                        employees = Employee.query.filter(Employee.id.in_(employee_ids_from_records)).all()
                    else:
                        emp_query = Employee.query.filter_by(active_for_payroll=True)
                        if unit_id_query:
                            emp_query = emp_query.filter_by(unit_id=unit_id_query)
                        employees = emp_query.all()

                if not employees:
                    employees = [None]

                rows = []
                for idx, emp in enumerate(employees, 1):
                    row = {}
                    if emp:
                        row['序号'] = idx
                        row['姓名'] = emp.name
                        row['单位'] = emp.unit.name if emp.unit else ''
                        row['身份证号'] = emp.id_card if emp else ''  # 新增
                    else:
                        row['序号'] = 1
                        row['姓名'] = ''
                        row['单位'] = ''
                        row['身份证号'] = ''  # 新增
                    for col in display_columns:
                        if col not in ['序号', '姓名', '单位']:
                            row[col] = ''
                    rows.append(row)

                return jsonify({
                    'config': config,
                    'columns': display_columns,
                    'rows': rows,
                    'meta': {
                        'month': month,
                        'unit_name': db.session.get(Unit, unit_id_query).name if unit_id_query else '',
                        'year': month.split('-')[0] if month else '',
                        'template_name': config.get('title', ''),
                        'empty_mode': True
                    },
                    'empty_mode': True
                })

            # 非空表逻辑（原有）
            rows = []
            for idx, (record, emp, unit) in enumerate(results, 1):
                details = record.details or {}
                if isinstance(details, str):
                    try:
                        details = json.loads(details)
                    except:
                        details = {}
                # ========== 新增：规范化 details 的键名（去掉前后空格） ==========
                details = {k.strip(): v for k, v in details.items()}
                gross = 0.0
                deduct = 0.0
                for item_name, val in details.items():
                    if item_name in ['应发工资', '扣款合计', '实发工资']:
                        continue
                    try:
                        val = float(val)
                    except:
                        val = 0.0

                    item_info = item_dict.get(item_name)
                    if item_info:
                        if item_info.item_type == 'income':
                            gross += val
                        elif item_info.item_type == 'deduction':
                            deduct += val
                        elif item_info.item_type == 'calculation':
                            calc_dir = getattr(item_info, 'calc_direction', 'income')
                            if calc_dir == 'income':
                                gross += val
                            else:
                                deduct += val
                    else:
                        if val >= 0:
                            gross += val
                        else:
                            deduct += val

                net = gross - deduct
                details['应发工资'] = round(gross, 2)
                details['扣款合计'] = round(deduct, 2)
                details['实发工资'] = round(net, 2)

                row = {
                    '序号': idx,
                    '姓名': emp.name,
                    '单位': unit.name,
                    'employee_id': emp.id,
                }
                str(emp.id_card) if emp.id_card else ''
                for col in display_columns:
                    if col in ['序号', '姓名', '单位']:
                        continue
                    row[col] = details.get(col.strip(), 0.0)  # 原来为 details.get(col, 0.0)

                row['备注'] = record.remark or ''
                rows.append(row)
            order = config.get('employee_order', [])
            if order:
                order_map = {eid: idx for idx, eid in enumerate(order)}
                rows.sort(key=lambda r: order_map.get(r.get('employee_id'), 99999))
            return jsonify({
                'config': config,
                'columns': display_columns,
                'rows': rows,
                'meta': {
                    'month': month,
                    'unit_name': db.session.get(Unit, unit_id).name if unit_id else ''
                }
            })

        # ========== 2. 特殊事项发放表 ==========
        elif template_type == 'special_item':
            special_template_id = config.get('special_template_id')
            if not special_template_id:
                return jsonify({'error': '模板未关联特殊事项'}), 400

            special_tpl = db.session.get(SpecialItemTemplate, special_template_id)
            if not special_tpl:
                return jsonify({'error': '关联的特殊事项模板不存在'}), 404

            extra_fields = []
            if special_tpl.extra_fields:
                try:
                    extra_fields = json.loads(special_tpl.extra_fields)
                except:
                    pass

            query = db.session.query(EmployeeSpecialGrant, Employee, Unit).join(
                Employee, EmployeeSpecialGrant.employee_id == Employee.id
            ).join(Unit, Employee.unit_id == Unit.id).filter(
                EmployeeSpecialGrant.template_id == special_template_id
            )

            if month:
                try:
                    y, m = month.split('-')
                    query = query.filter(EmployeeSpecialGrant.year == int(y), EmployeeSpecialGrant.month == int(m))
                except:
                    return jsonify({'error': '月份格式错误'}), 400
            elif year:
                query = query.filter(EmployeeSpecialGrant.year == int(year))
                if special_tpl.frequency == 'quarterly':
                    pass
                else:
                    query = query.filter(EmployeeSpecialGrant.month.is_(None))

            if unit_id:
                query = query.filter(Employee.unit_id == unit_id)

            employee_ids = request.args.getlist('employee_ids')
            if employee_ids:
                query = query.filter(Employee.id.in_(employee_ids))

            grants = query.all()
            column_display = config.get('column_display', {})

            if special_tpl.frequency == 'quarterly' and year:
                quarter = request.args.get('quarter')
                if not quarter:
                    return jsonify({'error': '季度模板需要指定quarter参数'}), 400
                q_months = {'Q1': [1, 2, 3], 'Q2': [4, 5, 6], 'Q3': [7, 8, 9], 'Q4': [10, 11, 12]}[quarter]

                if empty_mode:
                    cols = config.get('columns', [])
                    unit_id_query = request.args.get('unit_id')
                    year = request.args.get('year')
                    emp_ids_param = request.args.getlist('employee_ids')

                    if emp_ids_param:
                        employees = Employee.query.filter(Employee.id.in_(emp_ids_param)).all()
                    else:
                        grant_query = db.session.query(EmployeeSpecialGrant.employee_id).filter(
                            EmployeeSpecialGrant.template_id == special_template_id,
                            EmployeeSpecialGrant.year == int(year) if year else 0
                        )
                        if year:
                            grant_query = grant_query.filter(EmployeeSpecialGrant.month.in_(q_months))
                        if unit_id_query:
                            grant_query = grant_query.join(Employee,
                                                           EmployeeSpecialGrant.employee_id == Employee.id).filter(
                                Employee.unit_id == unit_id_query)
                        emp_ids_from_grant = {r[0] for r in grant_query.all()}
                        if emp_ids_from_grant:
                            employees = Employee.query.filter(Employee.id.in_(emp_ids_from_grant)).all()
                        else:
                            emp_query = Employee.query
                            if unit_id_query:
                                emp_query = emp_query.filter_by(unit_id=unit_id_query)
                            employees = emp_query.all()

                    if not employees:
                        employees = [None]

                    rows = []
                    for idx, emp in enumerate(employees, 1):
                        row = {}
                        if emp:
                            row['序号'] = idx
                            row['姓名'] = emp.name
                            row['单位'] = emp.unit.name if emp.unit else ''
                            row['身份证号'] = emp.id_card if emp else ''
                        else:
                            row['序号'] = 1
                            row['姓名'] = ''
                            row['单位'] = ''
                            row['身份证号'] = emp.id_card if emp else ''
                        for col in cols:
                            if col == '序号':
                                row[col] = row.get('序号')
                            elif col == '姓名':
                                row[col] = row.get('姓名', '')
                            elif col == '单位':
                                row[col] = row.get('单位', '')
                            elif col in [c for c in cols if '事项' in c or '模板' in c or '名称' in c]:
                                row[col] = special_tpl.name
                            else:
                                row[col] = ''
                        rows.append(row)
                    order = config.get('employee_order', [])
                    if order:
                        order_map = {eid: idx for idx, eid in enumerate(order)}
                        rows.sort(key=lambda r: order_map.get(r.get('employee_id'), 99999))
                    return jsonify({
                        'config': config,
                        'columns': cols,
                        'rows': rows,
                        'meta': {
                            'year': year,
                            'quarter': quarter,
                            'unit_name': db.session.get(Unit, unit_id_query).name if unit_id_query else '',
                            'template_name': special_tpl.name,
                            'frequency': 'quarterly',
                            'months': [f'{m}月' for m in q_months],
                            'empty_mode': True
                        },
                        'empty_mode': True
                    })

                # 季度非空表
                emp_data = {}
                for grant, emp, unit in grants:
                    if grant.month in q_months:
                        if emp.id not in emp_data:
                            emp_data[emp.id] = {
                                'employee_name': emp.name,
                                'unit_name': unit.name,
                                'assessments': {},
                                'amounts': {},
                                'remark': grant.remark or '',
                                'extra_data': {}
                            }
                        extra = json.loads(grant.extra_data) if grant.extra_data else {}
                        month_key = f'{grant.month}月'
                        emp_data[emp.id]['assessments'][month_key] = extra.get('assessment_value', '')
                        emp_data[emp.id]['amounts'][month_key] = grant.amount
                        if not emp_data[emp.id]['extra_data']:
                            emp_data[emp.id]['extra_data'] = extra

                rows = []
                for emp_id, data in emp_data.items():
                    emp = db.session.get(Employee, emp_id)  # 新增这一行，获取员工对象
                    id_card = emp.id_card if emp else ''  # 新增
                    row = {
                        'employee_name': data['employee_name'],
                        'unit_name': data['unit_name'],
                        'employee_id': emp_id,
                        '身份证号': id_card,  # 修改
                        'remark': data['remark']
                    }

                    total_amount = 0.0
                    for m in q_months:
                        mk = f'{m}月'
                        row[f'{mk}考核'] = data['assessments'].get(mk, '')
                        amt = data['amounts'].get(mk, 0.0)
                        row[f'{mk}金额(元)'] = amt
                        total_amount += amt
                    row['合计(元)'] = total_amount
                    for field in extra_fields:
                        field_name = field.get('name')
                        row[field_name] = data['extra_data'].get(field_name, '')
                    rows.append(row)
                order = config.get('employee_order', [])
                if order:
                    order_map = {eid: idx for idx, eid in enumerate(order)}
                    rows.sort(key=lambda r: order_map.get(r.get('employee_id'), 99999))
                return jsonify({
                    'config': config,
                    'columns': config.get('columns', []),
                    'rows': rows,
                    'meta': {
                        'year': year,
                        'quarter': quarter,
                        'unit_name': db.session.get(Unit, unit_id).name if unit_id else '',
                        'template_name': special_tpl.name,
                        'frequency': 'quarterly',
                        'months': [f'{m}月' for m in q_months]
                    },
                    'column_display': column_display
                })

            # 非季度（月度/年度）
            else:
                if empty_mode:
                    cols = config.get('columns', [])
                    month = request.args.get('month')
                    year = request.args.get('year')
                    unit_id_query = request.args.get('unit_id')
                    emp_ids_param = request.args.getlist('employee_ids')

                    if emp_ids_param:
                        employees = Employee.query.filter(Employee.id.in_(emp_ids_param)).all()
                    else:
                        grant_query = db.session.query(EmployeeSpecialGrant.employee_id).filter(
                            EmployeeSpecialGrant.template_id == special_template_id
                        )
                        year_int = int(year) if year else None
                        if year_int is not None:
                            grant_query = grant_query.filter(EmployeeSpecialGrant.year == year_int)
                            if month:
                                try:
                                    y, m = month.split('-')
                                    grant_query = grant_query.filter(
                                        EmployeeSpecialGrant.year == int(y),
                                        EmployeeSpecialGrant.month == int(m)
                                    )
                                except:
                                    pass
                            else:
                                if special_tpl.frequency not in ('quarterly',):
                                    grant_query = grant_query.filter(EmployeeSpecialGrant.month.is_(None))
                        elif month:
                            try:
                                y, m = month.split('-')
                                grant_query = grant_query.filter(EmployeeSpecialGrant.year == int(y),
                                                                 EmployeeSpecialGrant.month == int(m))
                            except:
                                pass

                        if unit_id_query:
                            grant_query = grant_query.join(Employee,
                                                           EmployeeSpecialGrant.employee_id == Employee.id).filter(
                                Employee.unit_id == unit_id_query)
                        emp_ids_from_grant = {r[0] for r in grant_query.all()}
                        if emp_ids_from_grant:
                            employees = Employee.query.filter(Employee.id.in_(emp_ids_from_grant)).all()
                        else:
                            emp_query = Employee.query
                            if unit_id_query:
                                emp_query = emp_query.filter_by(unit_id=unit_id_query)
                            employees = emp_query.all()

                    if not employees:
                        employees = [None]

                    rows = []
                    for idx, emp in enumerate(employees, 1):
                        row = {}
                        if emp:
                            row['序号'] = idx
                            row['姓名'] = emp.name
                            row['单位'] = emp.unit.name if emp.unit else ''
                            row['身份证号'] = emp.id_card if emp else ''  # 新增
                        else:
                            row['序号'] = 1
                            row['姓名'] = ''
                            row['单位'] = ''
                            row['身份证号'] = ''  # 新增

                        for col in cols:
                            if col == '序号':
                                row[col] = row.get('序号')
                            elif col == '姓名':
                                row[col] = row.get('姓名', '')
                            elif col == '单位':
                                row[col] = row.get('单位', '')
                            elif col in [c for c in cols if '事项' in c or '模板' in c or '名称' in c]:
                                row[col] = special_tpl.name
                            else:
                                row[col] = ''
                        rows.append(row)
                    order = config.get('employee_order', [])
                    if order:
                        order_map = {eid: idx for idx, eid in enumerate(order)}
                        rows.sort(key=lambda r: order_map.get(r.get('employee_id'), 99999))
                    return jsonify({
                        'config': config,
                        'columns': cols,
                        'rows': rows,
                        'meta': {
                            'year': year,
                            'month': month,
                            'template_name': special_tpl.name,
                            'frequency': special_tpl.frequency,
                            'unit_name': db.session.get(Unit, unit_id_query).name if unit_id_query else '',
                            'empty_mode': True
                        },
                        'empty_mode': True
                    })

                # 非空表逻辑（月度/年度）
                rows = []
                for idx, (grant, emp, unit) in enumerate(grants, 1):
                    extra_data = json.loads(grant.extra_data) if grant.extra_data else {}
                    row = {
                        '序号': idx,
                        '姓名': emp.name,
                        '单位': unit.name,
                        'employee_id': emp.id,  # ★ 新增
                        '身份证号': emp.id_card,  # 新增
                        '金额(元)': grant.amount,
                        '备注': grant.remark or ''
                    }
                    for field in extra_fields:
                        field_name = field.get('name')
                        row[field_name] = extra_data.get(field_name, '')
                    rows.append(row)

                if not rows:
                    example_row = {'序号': 1, '姓名': '示例员工', '单位': '示例单位', '金额(元)': 0.0, '备注': ''}
                    for f in extra_fields:
                        example_row[f.get('name')] = f.get('default', '')
                    rows.append(example_row)
                order = config.get('employee_order', [])
                if order:
                    order_map = {eid: idx for idx, eid in enumerate(order)}
                    rows.sort(key=lambda r: order_map.get(r.get('employee_id'), 99999))
                return jsonify({
                    'config': config,
                    'columns': config.get('columns', []),
                    'rows': rows,
                    'meta': {
                        'month': month,
                        'year': year,
                        'unit_name': db.session.get(Unit, unit_id).name if unit_id else '',
                        'template_name': special_tpl.name,
                        'frequency': special_tpl.frequency
                    },
                    'column_display': column_display
                })

        else:
            return jsonify({'error': f'不支持的模板类型: {template_type}'}), 400

    def get_default_template_config(template_type):
        """获取默认模板配置"""
        if template_type == 'salary_sheet':
            return {
                'title': '{unit_name}{year}年{month}月工资发放表',
                'subtitle': '编制单位：{unit_name}',
                'show_date': True,
                'show_unit': True,
                'columns': ['序号', '姓名', '岗位工资', '薪级工资', '应发工资', '扣款合计', '实发工资', '备注'],
                'show_total_row': True,
                'signatures': [
                    {'label': '负责人：', 'value': ''},
                    {'label': '复核：', 'value': ''},
                    {'label': '制表人：', 'value': ''}
                ]
            }
        return {}

    def is_number_column(col_name, template_type='salary_sheet'):
        if not col_name:
            return False

        # 特殊金额列
        special_amount_cols = ['金额(元)', '本季合计', '合计(元)', '应发工资', '扣款合计', '实发工资']
        if col_name in special_amount_cols:
            return True

        # 匹配 X月金额(元)
        if re.match(r'^\d+月金额\(元\)$', col_name):
            return True

        # 查询工资项表（使用函数属性缓存，避免重复查询）
        if not hasattr(is_number_column, '_salary_item_names'):
            salary_items = SalaryItem.query.all()
            is_number_column._salary_item_names = {item.name for item in salary_items}

        if col_name in is_number_column._salary_item_names:
            return True

        return False

    def _get_preview_data(template_type, template_id, month, year, quarter, unit_id, employee_ids, empty_mode=False):
        """获取打印预览数据（返回字典，与 /api/print_preview/<type> 返回值结构一致）"""
        from models import PrintTemplate, SalaryRecord, Employee, Unit, SalaryItem, SpecialItemTemplate, \
            EmployeeSpecialGrant

        template = db.session.get(PrintTemplate, template_id)
        if not template:
            return {'error': '模板不存在'}

        config = json.loads(template.config) if template.config else {}
        columns = config.get('columns', [])
        column_display = config.get('column_display', {})

        if template_type == 'salary_sheet':
            query = db.session.query(SalaryRecord, Employee, Unit).join(
                Employee, SalaryRecord.employee_id == Employee.id
            ).join(Unit, Employee.unit_id == Unit.id)

            if month:
                query = query.filter(SalaryRecord.month == month)
            if unit_id:
                query = query.filter(Employee.unit_id == unit_id)
            if employee_ids:
                query = query.filter(SalaryRecord.employee_id.in_(employee_ids))

            results = query.all()
            all_salary_items = SalaryItem.query.all()
            item_dict = {item.name: item for item in all_salary_items}

            rows = []
            for record, emp, unit in results:
                details = record.details or {}
                if isinstance(details, str):
                    try:
                        details = json.loads(details)
                    except:
                        details = {}
                # 规范化键名，避免列名与数据库中key不完全匹配
                details = {k.strip(): v for k, v in details.items()}

                gross = 0.0
                deduct = 0.0
                for item_name, val in details.items():
                    if item_name in ['应发工资', '扣款合计', '实发工资']:
                        continue
                    try:
                        val = float(val)
                    except:
                        val = 0.0
                    item_info = item_dict.get(item_name)
                    if item_info:
                        if item_info.item_type == 'income':
                            gross += val
                        elif item_info.item_type == 'deduction':
                            deduct += val
                        elif item_info.item_type == 'calculation':
                            calc_dir = getattr(item_info, 'calc_direction', 'income')
                            if calc_dir == 'income':
                                gross += val
                            else:
                                deduct += val
                    else:
                        if val >= 0:
                            gross += val
                        else:
                            deduct += val

                net = gross - deduct
                details['应发工资'] = round(gross, 2)
                details['扣款合计'] = round(deduct, 2)
                details['实发工资'] = round(net, 2)

                row = {
                    '姓名': emp.name,
                    '单位': unit.name,
                    '身份证号': emp.id_card,  # 新增
                }
                for col in columns:
                    row[col] = details.get(col.strip(), 0.0)
                row['备注'] = record.remark or ''
                rows.append(row)

            meta = {
                'month': month,
                'unit_name': db.session.get(Unit, unit_id).name if unit_id else '',
                'year': month.split('-')[0] if month else '',
                'template_name': template.name
            }
            return {'rows': rows, 'meta': meta}

        elif template_type == 'special_item':
            special_template_id = config.get('special_template_id')
            if not special_template_id:
                return {'error': '模板未关联特殊事项'}

            special_tpl = db.session.get(SpecialItemTemplate, special_template_id)
            if not special_tpl:
                return {'error': '关联的特殊事项模板不存在'}

            extra_fields = []
            if special_tpl.extra_fields:
                try:
                    extra_fields = json.loads(special_tpl.extra_fields)
                except:
                    pass

            query = db.session.query(EmployeeSpecialGrant, Employee, Unit).join(
                Employee, EmployeeSpecialGrant.employee_id == Employee.id
            ).join(Unit, Employee.unit_id == Unit.id).filter(
                EmployeeSpecialGrant.template_id == special_template_id
            )

            if month:
                try:
                    y, m = month.split('-')
                    query = query.filter(EmployeeSpecialGrant.year == int(y), EmployeeSpecialGrant.month == int(m))
                except:
                    return {'error': '月份格式错误'}
            elif year:
                query = query.filter(EmployeeSpecialGrant.year == int(year))
                if special_tpl.frequency != 'quarterly':
                    query = query.filter(EmployeeSpecialGrant.month.is_(None))

            if unit_id:
                query = query.filter(Employee.unit_id == unit_id)
            if employee_ids:
                query = query.filter(Employee.id.in_(employee_ids))

            grants = query.all()

            if special_tpl.frequency == 'quarterly' and year:
                if not quarter:
                    return {'error': '季度模板需要指定quarter参数'}
                q_months = {'Q1': [1, 2, 3], 'Q2': [4, 5, 6], 'Q3': [7, 8, 9], 'Q4': [10, 11, 12]}[quarter]

                emp_data = {}
                for grant, emp, unit in grants:
                    if not emp.active_for_payroll:
                        continue
                    if grant.month in q_months:
                        if emp.id not in emp_data:
                            emp_data[emp.id] = {
                                '姓名': emp.name,
                                '单位': unit.name,
                                'assessments': {},
                                'amounts': {},
                                'remark': grant.remark or '',
                            }
                        extra = json.loads(grant.extra_data or '{}')
                        month_key = f'{grant.month}月'
                        emp_data[emp.id]['assessments'][month_key] = extra.get('assessment_value', '')
                        emp_data[emp.id]['amounts'][month_key] = grant.amount

                rows = []
                for emp_id, data in emp_data.items():
                    emp = db.session.get(Employee, emp_id)
                    row = {
                        '姓名': data['姓名'],
                        '单位': data['单位'],
                        '身份证号': emp.id_card if emp else ''
                    }
                    total = 0.0
                    for m in q_months:
                        mk = f'{m}月'
                        row[f'{mk}考核'] = data['assessments'].get(mk, '')
                        amt = data['amounts'].get(mk, 0.0)
                        row[f'{mk}金额(元)'] = amt
                        total += amt
                    row['合计(元)'] = total
                    for field in extra_fields:
                        row[field.get('name')] = data.get(field.get('name'), '')
                    row['备注'] = data['remark']
                    rows.append(row)

                meta = {
                    'year': year,
                    'quarter': quarter,
                    'unit_name': db.session.get(Unit, unit_id).name if unit_id else '',
                    'template_name': special_tpl.name,
                    'months': [f'{m}月' for m in q_months]
                }
                return {'rows': rows, 'meta': meta}

            # 非季度
            rows = []
            for grant, emp, unit in grants:
                extra_data = json.loads(grant.extra_data or '{}')
                row = {
                    '姓名': emp.name,
                    '单位': unit.name,
                    '身份证号': emp.id_card,
                    '金额(元)': grant.amount,
                    '备注': grant.remark or ''
                }
                for field in extra_fields:
                    row[field.get('name')] = extra_data.get(field.get('name'), '')
                rows.append(row)

            meta = {
                'month': month,
                'year': year,
                'unit_name': db.session.get(Unit, unit_id).name if unit_id else '',
                'template_name': special_tpl.name,
            }
            return {'rows': rows, 'meta': meta}

        return {'error': '不支持的模板类型'}
    @app.route('/assessments', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def manage_assessments():
        if request.method == 'POST':
            if 'delete_id' in request.form:
                item_id = request.form['delete_id']
                item = db.session.get(AssessmentItem, item_id)
                if item:
                    EmployeeAssessment.query.filter_by(assessment_id=item_id).delete()
                    db.session.delete(item)
                    db.session.commit()
                    flash(f"考核项 '{item.name}' 已删除", "success")
                return redirect(url_for('manage_assessments'))
            name = request.form['name']
            option_good = request.form['option_good']
            option_good_value = float(request.form['option_good_value'])
            option_better = request.form['option_better']
            option_better_value = float(request.form['option_better_value'])
            option_blank = request.form['option_blank']
            option_blank_value = float(request.form['option_blank_value'])
            new_item = AssessmentItem(name=name, option_good=option_good, option_good_value=option_good_value, option_better=option_better, option_better_value=option_better_value, option_blank=option_blank, option_blank_value=option_blank_value)
            db.session.add(new_item)
            db.session.commit()
            flash(f"考核项 '{name}' 添加成功", "success")
            return redirect(url_for('manage_assessments'))
        items = AssessmentItem.query.all()
        return render_template('assessments.html', items=items)

    @app.route('/enter_assessments', methods=['GET', 'POST'])
    @login_required
    @admin_required  # ← 新增
    def enter_assessments():
        month = request.args.get('month', datetime.now().strftime('%Y-%m'))
        assessment_id = request.args.get('assessment_id')
        if request.method == 'POST':
            month = request.form['month']
            assessment_id = request.form['assessment_id']
            unit_id = request.form.get('unit_id')
            for emp in Employee.query.filter_by(unit_id=unit_id, active_for_payroll=True).all():
                result_key = f"result_{emp.id}"
                remark_key = f"remark_{emp.id}"
                if result_key in request.form:
                    result = request.form[result_key]
                    remark = request.form.get(remark_key, '')
                    record = EmployeeAssessment.query.filter_by(employee_id=emp.id, assessment_id=assessment_id, month=month).first()
                    if record:
                        record.result = result
                        record.remark = remark
                    else:
                        record = EmployeeAssessment(employee_id=emp.id, assessment_id=assessment_id, month=month, result=result, remark=remark)
                        db.session.add(record)
            db.session.commit()
            flash(f"{month}月考核结果已保存", "success")
            return redirect(url_for('enter_assessments', month=month, assessment_id=assessment_id))
        assessments = AssessmentItem.query.all()
        units = Unit.query.all()
        selected_unit_id = request.args.get('unit_id')
        records = {}
        if assessment_id and selected_unit_id:
            records = {r.employee_id: {'result': r.result, 'remark': r.remark} for r in EmployeeAssessment.query.filter_by(assessment_id=assessment_id, month=month).join(Employee).filter(Employee.unit_id == selected_unit_id).all()}
        return render_template('enter_assessments.html', assessments=assessments, units=units, month=month, assessment_id=assessment_id, selected_unit_id=selected_unit_id, records=records)

    @app.route('/quarter_assessment_report', methods=['GET', 'POST'])
    @login_required
    def quarter_assessment_report():
        if request.method == 'POST':
            quarter = request.form['quarter']
            year = request.form['year']
            assessment_id = request.form['assessment_id']
            unit_id = request.form['unit_id']
            return redirect(url_for('view_quarter_report', quarter=quarter, year=year, assessment_id=assessment_id, unit_id=unit_id))
        current_year = datetime.now().year
        years = [current_year - 1, current_year, current_year + 1]
        assessments = AssessmentItem.query.all()
        units = Unit.query.all()
        return render_template('quarter_assessment_report.html', years=years, assessments=assessments, units=units)

    @app.route('/view_quarter_report')
    @login_required
    def view_quarter_report():
        quarter = request.args['quarter']
        year = int(request.args['year'])
        assessment_id = request.args['assessment_id']
        unit_id = request.args['unit_id']
        if quarter == 'Q1':
            months = [f"{year}-01", f"{year}-02", f"{year}-03"]
        elif quarter == 'Q2':
            months = [f"{year}-04", f"{year}-05", f"{year}-06"]
        elif quarter == 'Q3':
            months = [f"{year}-07", f"{year}-08", f"{year}-09"]
        elif quarter == 'Q4':
            months = [f"{year}-10", f"{year}-11", f"{year}-12"]
        else:
            flash("无效的季度", "danger")
            return redirect(url_for('quarter_assessment_report'))
        assessment = db.session.get(AssessmentItem, assessment_id)
        if not assessment:
            flash("考核项不存在", "danger")
            return redirect(url_for('quarter_assessment_report'))
        unit = db.session.get(Unit, unit_id)
        if not unit:
            flash("单位不存在", "danger")
            return redirect(url_for('quarter_assessment_report'))
        employees = Employee.query.filter_by(unit_id=unit_id, active_for_payroll=True).all()
        report_data = []
        total_quarter = 0
        for emp in employees:
            emp_data = {'id': emp.id, 'name': emp.name, 'months': {}, 'quarter_total': 0, 'remarks': set()}
            for month in months:
                record = EmployeeAssessment.query.filter_by(employee_id=emp.id, assessment_id=assessment_id, month=month).first()
                if record:
                    if record.result == 'good':
                        amount = assessment.option_good_value
                        result_text = assessment.option_good
                    elif record.result == 'better':
                        amount = assessment.option_better_value
                        result_text = assessment.option_better
                    else:
                        amount = assessment.option_blank_value
                        result_text = assessment.option_blank
                    emp_data['months'][month] = {'result': result_text, 'amount': amount, 'remark': record.remark}
                    emp_data['quarter_total'] += amount
                    if record.remark:
                        emp_data['remarks'].add(record.remark)
                else:
                    emp_data['months'][month] = {'result': assessment.option_better, 'amount': assessment.option_better_value, 'remark': None}
                    emp_data['quarter_total'] += assessment.option_better_value
            total_quarter += emp_data['quarter_total']
            emp_data['remarks'] = "; ".join(emp_data['remarks'])
            report_data.append(emp_data)
        report_data.sort(key=lambda x: x['quarter_total'], reverse=True)
        return render_template('view_quarter_report.html', report_data=report_data, months=months, quarter=quarter, year=year, assessment=assessment, unit=unit, total_quarter=total_quarter, print_time=datetime.now().strftime("%Y年%m月%d日"))

    @app.route('/enter_special_assessments', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def enter_special_assessments():
        month = request.args.get('month', datetime.now().strftime('%Y-%m'))
        special_item_id = request.args.get('special_item_id')
        unit_id = request.args.get('unit_id')
        special_item = None
        if special_item_id:
            special_item = db.session.get(SpecialSalaryItem, special_item_id)
        if request.method == 'POST':
            month = request.form['month']
            special_item_id = request.form['special_item_id']
            unit_id = request.form['unit_id']
            special_item = db.session.get(SpecialSalaryItem, special_item_id)
            for emp in Employee.query.filter_by(unit_id=unit_id, active_for_payroll=True).all():
                assessment_key = f"assessment_{emp.id}"
                if assessment_key in request.form:
                    assessment_value = request.form[assessment_key]
                    record = EmployeeSpecialItem.query.filter_by(employee_id=emp.id, special_item_id=special_item_id, month=month).first()
                    calculated_amount = calculate_special_amount(special_item, float(assessment_value))
                    if record:
                        record.assessment_value = assessment_value
                        record.calculated_amount = calculated_amount
                    else:
                        record = EmployeeSpecialItem(employee_id=emp.id, special_item_id=special_item_id, month=month, assessment_value=assessment_value, calculated_amount=calculated_amount)
                        db.session.add(record)
            db.session.commit()
            flash(f"{month}月考核结果已保存", "success")
            return redirect(url_for('enter_special_assessments', month=month, special_item_id=special_item_id, unit_id=unit_id))
        units = Unit.query.all()
        special_items = SpecialSalaryItem.query.all()
        employees = []
        records = {}
        if special_item_id and unit_id:
            employees = Employee.query.filter_by(unit_id=unit_id, active_for_payroll=True).all()
            existing_records = EmployeeSpecialItem.query.filter_by(special_item_id=special_item_id, month=month).filter(EmployeeSpecialItem.employee_id.in_([e.id for e in employees])).all()
            records = {r.employee_id: r for r in existing_records}
        return render_template('enter_special_assessments.html', units=units, month=month, special_item_id=special_item_id, special_item=special_item, selected_unit_id=unit_id, employees=employees, records=records, special_items=special_items)

    # API 路由
    @app.route('/api/employees', methods=['GET'])
    @login_required
    def api_employees_basic():
        unit_id = request.args.get('unit_id')
        active_only = request.args.get('active_only', 'true').lower() == 'true'
        query = Employee.query
        if unit_id:
            query = query.filter_by(unit_id=unit_id)
        if active_only:
            query = query.filter_by(active_for_payroll=True)
        employees = query.all()
        return jsonify([{
            'id': emp.id,
            'name': emp.name,
            'unit': {'id': emp.unit.id, 'name': emp.unit.name}
        } for emp in employees])


    @app.route('/api/salary_items', methods=['GET'])
    @login_required
    def api_salary_items():
        items = SalaryItem.query.all()
        return jsonify([{'id': item.id, 'name': item.name, 'item_type': item.item_type, 'formula': item.formula, 'default_value': item.default_value} for item in items])

    @app.route('/api/salary_records')
    @login_required
    def api_salary_records():
        employee_id = request.args.get('employee_id')
        month = request.args.get('month')
        if not employee_id or not month:
            return jsonify({'error': 'Missing parameters'}), 400
        record = SalaryRecord.query.filter_by(employee_id=employee_id, month=month).first()
        if record:
            return jsonify({'records': [{'id': record.id, 'employee_id': record.employee_id, 'month': record.month, 'details': record.details, 'total': record.total}]})
        else:
            return jsonify({'records': []})

    @app.route('/api/salary_entry', methods=['POST'])
    @login_required
    def api_salary_entry():
        if not current_user.is_admin:
            return jsonify({'success': False, 'error': '无权限'}), 403
        """AJAX 保存工资记录，返回 JSON（增强版）"""
        data = request.get_json()
        employee_id = data.get('employee_id')
        month = data.get('month')
        items = data.get('items', {})
        remark = data.get('remark', '')
        item_remarks = data.get('item_remarks', {})

        if not employee_id or not month:
            return jsonify({'success': False, 'error': '缺少员工ID或月份'}), 400

        # 校验员工是否存在且参与工资计算
        employee = db.session.get(Employee, employee_id)
        if not employee or not employee.active_for_payroll:
            return jsonify({'success': False, 'error': '员工不存在或已停用'}), 400

        record = SalaryRecord.query.filter_by(employee_id=employee_id, month=month).first()
        if not record:
            record = SalaryRecord(employee_id=employee_id, month=month, details={}, total=0, remark=remark)
            db.session.add(record)

        # 构建初始 details（来自前端输入）
        details = {}
        all_salary_items = SalaryItem.query.all()
        for item in all_salary_items:
            value = items.get(item.name, 0)
            try:
                details[item.name] = float(value)
            except:
                details[item.name] = 0.0

        # 处理临时项目
        temp_items = data.get('temp_items', [])
        for temp in temp_items:
            name = temp.get('name')
            value = temp.get('value', 0)
            if name:
                details[name] = float(value)

        # ========== 重算计算项（最多5轮） ==========
        for _ in range(5):
            changed = False
            for item in all_salary_items:
                if item.item_type != 'calculation' or not item.formula:
                    continue
                variables = details.copy()
                try:
                    new_val = evaluate_formula(item.formula, variables)
                    new_val = round(new_val, 2)
                    old_val = details.get(item.name, 0)
                    if abs(new_val - old_val) > 0.01:
                        details[item.name] = new_val
                        changed = True
                except Exception as e:
                    current_app.logger.warning(f"计算项 {item.name} 公式失败: {e}")
            if not changed:
                break

        # ========== 计算应发、扣款、实发 ==========
        gross = 0.0
        deduction = 0.0
        item_dict = {item.name: item for item in all_salary_items}
        for name, val in details.items():
            if name in ['应发工资', '扣款合计', '实发工资']:
                continue
            val = float(val)
            item_info = item_dict.get(name)
            if item_info:
                if item_info.item_type == 'income':
                    gross += val
                elif item_info.item_type == 'deduction':
                    deduction += val
                elif item_info.item_type == 'calculation':
                    calc_dir = getattr(item_info, 'calc_direction', 'income')
                    if calc_dir == 'income':
                        gross += val
                    else:
                        deduction += val
            else:
                if val >= 0:
                    gross += val
                else:
                    deduction += val

        net = gross - deduction
        details['应发工资'] = round(gross, 2)
        details['扣款合计'] = round(deduction, 2)
        details['实发工资'] = round(net, 2)
        from sqlalchemy.orm.attributes import flag_modified  # 请确保文件顶部已导入

        # 保存数据
        record.details = details
        record.total = net
        record.remark = remark
        record.item_remarks = json.dumps(item_remarks, ensure_ascii=False)

        # 强制标记 JSON 字段为已修改（避免 SQLAlchemy 检测不到内部变化）
        flag_modified(record, "details")
        flag_modified(record, "item_remarks")

        try:
            db.session.commit()
            return jsonify({'success': True, 'net_salary': net})
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"保存工资记录失败: {str(e)}", exc_info=True)
            return jsonify({'success': False, 'error': f'数据库错误: {str(e)}'}), 500
    @app.route('/api/units', methods=['GET'])
    @login_required
    def api_units():
        units = Unit.query.all()
        return jsonify([{'id': unit.id, 'name': unit.name} for unit in units])

    @app.route('/api/salary_item_names', methods=['GET'])
    @login_required
    def api_salary_item_names():
        """返回所有工资项的名称列表，供前端插入上年工资项合计变量"""
        items = SalaryItem.query.order_by(SalaryItem.order).all()
        return jsonify([item.name for item in items])

    @app.route('/api/batch_update_remarks', methods=['POST'])
    @login_required
    @admin_required
    def batch_update_remarks():
        """批量修改工资记录的备注"""
        data = request.get_json()
        record_ids = data.get('record_ids', [])
        remark = data.get('remark', '').strip()

        if not record_ids:
            return jsonify({'success': False, 'error': '未选择任何记录'})

        updated = 0
        for rid in record_ids:
            record = db.session.get(SalaryRecord, rid)
            if record:
                record.remark = remark
                updated += 1
        try:
            db.session.commit()
            return jsonify({'success': True, 'updated': updated})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)})

    @app.route('/api/employees_detail', methods=['GET'])
    @login_required
    def api_employees_detail():
        employee_id = request.args.get('employee_id')
        unit_id = request.args.get('unit_id')
        active_only = request.args.get('active_only', 'true').lower() == 'true'
        employee_type = request.args.get('employee_type')
        is_veteran = request.args.get('is_veteran')
        annual_special_active = request.args.get('annual_special_active', 'false').lower() == 'true'  # 新增

        # 1. 单独查询某个员工（忽略其他过滤条件）
        if employee_id:
            emp = db.session.get(Employee, employee_id)
            if not emp:
                return jsonify({'error': '员工不存在'}), 404
            return jsonify([{
                'id': emp.id,
                'name': emp.name,
                'unit_id': emp.unit.id,
                'unit_name': emp.unit.name,
                'id_card': str(emp.id_card) if emp.id_card else '',   # ← 强制转字符串
                'gender': emp.gender,
                'birth_date': emp.birth_date.strftime('%Y-%m-%d') if emp.birth_date else None,
                'employee_type': emp.employee_type,
                'employee_identity': emp.employee_identity,
                'join_date': emp.join_date.strftime('%Y-%m-%d') if emp.join_date else None,
                'position_level': emp.position_level,
                'bank_account': emp.bank_account,
                'is_veteran': emp.is_veteran,
                'email': emp.email
            }])

        # 2. 通用查询（支持单位、在职状态、人员类型、退役军人筛选）
        query = Employee.query

        if annual_special_active:  # 新增
            query = query.filter(Employee.active_for_annual_special == True)

        if unit_id:
            query = query.filter_by(unit_id=unit_id)

        if active_only:
            query = query.filter_by(active_for_payroll=True)

        if employee_type:
            if ',' in employee_type:
                types = [t.strip() for t in employee_type.split(',')]
                query = query.filter(Employee.employee_type.in_(types))
            else:
                query = query.filter_by(employee_type=employee_type)

        if is_veteran is not None:
            query = query.filter(Employee.is_veteran == (is_veteran.lower() == 'true'))

        employees = query.all()

        return jsonify([{
            'id': emp.id,
            'name': emp.name,
            'unit_id': emp.unit.id,
            'unit_name': emp.unit.name,
            'id_card': str(emp.id_card) if emp.id_card else '',
            'gender': emp.gender,
            'birth_date': emp.birth_date.strftime('%Y-%m-%d') if emp.birth_date else None,
            'employee_type': emp.employee_type,
            'employee_identity': emp.employee_identity,
            'join_date': emp.join_date.strftime('%Y-%m-%d') if emp.join_date else None,
            'position_level': emp.position_level,
            'bank_account': emp.bank_account,
            'is_veteran': emp.is_veteran,
            'email': emp.email
        } for emp in employees])

    @app.route('/api/special_items/<int:item_id>')
    @login_required
    def api_special_item(item_id):
        try:
            item = SpecialSalaryItem.query.get_or_404(item_id)
            options = AssessmentOption.query.filter_by(special_item_id=item_id).all()
            return jsonify({'id': item.id, 'name': item.name, 'amount': item.amount, 'frequency': item.frequency, 'assessment_type': item.assessment_type, 'assessment_options': [{'id': opt.id, 'option_name': opt.option_name, 'formula': opt.formula} for opt in options]})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route('/api/units/<int:unit_id>')
    @login_required
    def api_unit(unit_id):
        try:
            unit = Unit.query.get_or_404(unit_id)
            return jsonify({"id": unit.id, "name": unit.name})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route('/api/save_special_remark', methods=['POST'])
    @login_required
    def api_save_special_remark():
        data = request.get_json()
        employee_id = data.get('employee_id')
        special_item_id = data.get('special_item_id')
        period = data.get('period')
        remark = data.get('remark', '')
        if not employee_id or not special_item_id or not period:
            return jsonify({'success': False, 'error': '缺少必要参数'}), 400
        try:
            if len(period) == 7:
                record = EmployeeSpecialItem.query.filter_by(employee_id=employee_id, special_item_id=special_item_id, month=period).first()
                if record:
                    record.remark = remark
                else:
                    record = EmployeeSpecialItem(employee_id=employee_id, special_item_id=special_item_id, month=period, remark=remark)
                    db.session.add(record)
                db.session.commit()
                return jsonify({'success': True})
            elif len(period) == 4:
                for month in range(1, 13):
                    month_str = f"{period}-{str(month).zfill(2)}"
                    record = EmployeeSpecialItem.query.filter_by(employee_id=employee_id, special_item_id=special_item_id, month=month_str).first()
                    if record:
                        record.remark = remark
                    else:
                        record = EmployeeSpecialItem(employee_id=employee_id, special_item_id=special_item_id, month=month_str, remark=remark)
                        db.session.add(record)
                db.session.commit()
                return jsonify({'success': True})
            else:
                quarter, year = period.split('-')
                quarter_months = {'Q1': [1,2,3], 'Q2': [4,5,6], 'Q3': [7,8,9], 'Q4': [10,11,12]}
                if quarter in quarter_months:
                    for month in quarter_months[quarter]:
                        month_str = f"{year}-{str(month).zfill(2)}"
                        record = EmployeeSpecialItem.query.filter_by(employee_id=employee_id, special_item_id=special_item_id, month=month_str).first()
                        if record:
                            record.remark = remark
                        else:
                            record = EmployeeSpecialItem(employee_id=employee_id, special_item_id=special_item_id, month=month_str, remark=remark)
                            db.session.add(record)
                    db.session.commit()
                    return jsonify({'success': True})
                else:
                    return jsonify({'success': False, 'error': '无效的季度格式'}), 400
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/api/quarter_salary_table')
    @login_required
    def api_quarter_salary_table():
        quarter = request.args.get('quarter')
        year = request.args.get('year')
        item_id = request.args.get('itemId')
        unit_id = request.args.get('unitId')
        employee_ids = request.args.get('employeeIds')

        if not quarter or not year or not item_id:
            return jsonify({'error': '缺少必要参数'}), 400

        try:
            # 季度月份映射
            quarter_months = {
                'Q1': [1, 2, 3], 'Q2': [4, 5, 6],
                'Q3': [7, 8, 9], 'Q4': [10, 11, 12]
            }
            months = quarter_months[quarter]
            month_names = [f'{m}月' for m in months]

            # 获取模板和单位信息
            template = db.session.get(SpecialItemTemplate, item_id)
            if not template:
                return jsonify({'error': '事项模板不存在'}), 404

            unit_name = ''
            if unit_id:
                unit = db.session.get(Unit, unit_id)
                if unit:
                    unit_name = unit.name

            # 查询员工
            if employee_ids:
                emp_id_list = [int(eid) for eid in employee_ids.split(',') if eid.strip()]
                query = Employee.query.filter(Employee.id.in_(emp_id_list))
                if unit_id:
                    query = query.filter_by(unit_id=unit_id)
                employees = query.filter_by(active_for_payroll=True).all()
            elif unit_id:
                employees = Employee.query.filter_by(unit_id=unit_id, active_for_payroll=True).all()
            else:
                employees = Employee.query.filter_by(active_for_payroll=True).all()

            result = {
                'unit_name': unit_name,
                'item_name': template.name,
                'months': month_names,
                'employees': []
            }

            # 导入公式计算函数（需要定义在文件顶部，见后文）
            for emp in employees:
                emp_data = {
                    'id': emp.id,
                    'name': emp.name,
                    'id_card': str(emp.id_card) if emp.id_card else '',
                    'employee_type': emp.employee_type,
                    'assessments': {},
                    'amounts': {},
                    'total': 0.0,
                    'remark': ''
                }

                for idx, month in enumerate(months):
                    month_str = f"{year}-{str(month).zfill(2)}"
                    month_key = month_names[idx]

                    grant = EmployeeSpecialGrant.query.filter_by(
                        employee_id=emp.id,
                        template_id=template.id,
                        year=int(year),
                        month=month
                    ).first()

                    if grant:
                        extra = json.loads(grant.extra_data or '{}')
                        assessment_val = extra.get('assessment_value', '')
                        # 若模板为公式计算且未提供手工金额，则重新计算
                        if template.calculation_type == 'formula' and template.formula:
                            # 从 grant.extra_data 中获取所有自定义字段
                            extra = json.loads(grant.extra_data or '{}')
                            vars_for_formula = {'base_amount': template.base_amount, '考核': assessment_val, **extra}
                            amount = evaluate_formula(template.formula, vars_for_formula)
                        else:
                            amount = grant.amount
                        emp_data['assessments'][month_key] = assessment_val
                        emp_data['amounts'][month_key] = amount
                        emp_data['total'] += amount
                        emp_data['remark'] = grant.remark or ''
                    else:
                        emp_data['assessments'][month_key] = ''
                        emp_data['amounts'][month_key] = 0.0

                result['employees'].append(emp_data)

            return jsonify(result)

        except Exception as e:
            current_app.logger.error(f"生成季度发放表失败: {str(e)}", exc_info=True)
            return jsonify({'error': '服务器内部错误: ' + str(e)}), 500

    @app.route('/api/monthly_special_table')
    @login_required
    def api_monthly_special_table():
        month = request.args.get('month')
        item_id = request.args.get('itemId')
        unit_id = request.args.get('unitId')
        if not month or not item_id or not unit_id:
            return jsonify({'error': '缺少必要参数'}), 400
        try:
            unit = db.session.get(Unit, unit_id)
            if not unit:
                return jsonify({'error': '单位不存在'}), 404

            template = db.session.get(SpecialItemTemplate, item_id)
            if not template or template.frequency != 'monthly':
                return jsonify({'error': '事项模板不存在或非月度模板'}), 404

            employees = Employee.query.filter_by(unit_id=unit_id, active_for_payroll=True).all()
            result = {
                'unit_name': unit.name,
                'item_name': template.name,
                'employees': []
            }

            # 解析月份为整数
            year = int(month[:4])
            month_num = int(month[5:7])

            for emp in employees:
                grant = EmployeeSpecialGrant.query.filter_by(
                    employee_id=emp.id,
                    template_id=template.id,
                    year=year,
                    month=month_num
                ).first()

                if grant:
                    extra = json.loads(grant.extra_data or '{}')
                    # ========== 新增：转换数字字符串 ==========
                    for k, v in extra.items():
                        if isinstance(v, str) and v.replace('.', '', 1).isdigit():
                            extra[k] = float(v) if '.' in v else int(v)
                    # =====================================
                    assessment_val = extra.get('assessment_value', '')
                    if template.calculation_type == 'formula' and template.formula:
                        amount = evaluate_formula(template.formula,
                                                  {'base_amount': template.base_amount, '考核': assessment_val,
                                                   **extra})
                    else:
                        amount = grant.amount
                    remark = grant.remark or ''
                else:
                    if template.calculation_type == 'fixed':
                        amount = template.base_amount
                    else:
                        amount = 0.0
                    assessment_val = ''
                    remark = ''
                    extra = {}  # ✅ 无记录时为空字典

                emp_data = {
                    'id': emp.id,
                    'name': emp.name,
                    'id_card': str(emp.id_card) if emp.id_card else '',
                    'assessment': assessment_val,
                    'amount': amount,
                    'remark': remark,
                    **extra
                }
                result['employees'].append(emp_data)
            return jsonify(result)
        except Exception as e:
            current_app.logger.error(f"生成月度发放表失败: {str(e)}", exc_info=True)
            return jsonify({'error': '服务器内部错误: ' + str(e)}), 500

    @app.route('/api/yearly_special_table')
    @login_required
    def api_yearly_special_table():
        year = request.args.get('year')
        item_id = request.args.get('itemId')
        unit_id = request.args.get('unitId')
        employee_ids = request.args.get('employeeIds')
        employee_type = request.args.get('employee_type')
        include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'  # 新增

        if not year or not item_id:
            return jsonify({'error': '缺少必要参数'}), 400
        try:
            template = db.session.get(SpecialItemTemplate, item_id)
            if not template:
                return jsonify({'error': '事项模板不存在'}), 404

            unit = db.session.get(Unit, unit_id) if unit_id else None
            if unit_id and not unit:
                return jsonify({'error': '单位不存在'}), 404

            base_query = Employee.query
            if not include_inactive:  # 新增
                base_query = base_query.filter(Employee.active_for_annual_special == True)

            if employee_ids:
                emp_id_list = [int(eid) for eid in employee_ids.split(',') if eid.strip()]
                base_query = base_query.filter(Employee.id.in_(emp_id_list))
            if unit_id:
                base_query = base_query.filter_by(unit_id=unit_id)
            if employee_type:
                base_query = base_query.filter_by(employee_type=employee_type)

            employees = base_query.all()

            result = {
                'unit_name': unit.name if unit else '',
                'item_name': template.name,
                'employees': []
            }

            for employee in employees:
                grant = EmployeeSpecialGrant.query.filter_by(
                    employee_id=employee.id,
                    template_id=template.id,
                    year=int(year),
                    month=None
                ).first()

                total = 0.0
                remark = ''
                extra_data_merged = {}

                if grant:
                    total = grant.amount
                    if total == 0:
                        continue
                    remark = grant.remark or ''
                    extra_data_merged = json.loads(grant.extra_data or '{}')
                    for k, v in list(extra_data_merged.items()):
                        if isinstance(v, str) and v.replace('.', '', 1).isdigit():
                            extra_data_merged[k] = float(v) if '.' in v else int(v)
                else:
                    continue

                emp_data = {
                    'id': employee.id,
                    'name': employee.name,
                    'unit_name': employee.unit.name if employee.unit else '',
                    'id_card': str(employee.id_card) if employee.id_card else '',
                    'employee_type': employee.employee_type,
                    'total': round(total, 2),
                    'remark': remark,
                    'grant_id': grant.id if grant else None,
                }
                emp_data.update(extra_data_merged)
                result['employees'].append(emp_data)

            total_amount = sum(emp['total'] for emp in result['employees'])
            result['total_amount'] = round(total_amount, 2)

            return jsonify(result)

        except Exception as e:
            current_app.logger.error(f"生成年度发放表失败: {str(e)}", exc_info=True)
            return jsonify({'error': '服务器内部错误: ' + str(e)}), 500

    # 其他辅助路由
    @app.route('/salary_entry/delete_item/<int:record_id>/<item_name>', methods=['POST'])
    @login_required
    @admin_required  # 建议添加
    def delete_salary_item(record_id, item_name):
        record = SalaryRecord.query.get_or_404(record_id)
        if item_name in record.details:
            record.details.pop(item_name)
            record.total = sum(record.details.values())
            db.session.commit()
            flash(f"已删除工资项: {item_name}", "success")
        else:
            flash("工资项不存在", "danger")
        return redirect(url_for('salary_entry', record_id=record_id))

    @app.route('/employees/export_template')
    @login_required
    def export_employee_template():
        export_dir = os.path.join(current_dir, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = "员工导入模板.xlsx"
        filepath = os.path.join(export_dir, filename)
        data = [{
            '姓名': '张三',
            '单位': '人力资源部',
            '身份证号': '110105199001011234',
            '性别': '男',
            '出生日期': '1990-01-01',
            '人员类型': '在职',
            '人员身份': '管理人员',
            '入职日期': '2012-07-15',
            '退休时间': '',  # 新增
            '银行卡号': '6217000010001234567',
            '是否退役军人': '否',
            '邮箱': 'zhangsan@example.com',
            '电话号码': '13800138000',  # 新增
            '岗位级别': '高级工程师 P7',
            '薪资级别': 'G8',
            '扩展信息': '{"备注":"示例"}'
        }]
        df = pd.DataFrame(data)
        df.to_excel(filepath, index=False)
        return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/export_salary_items_template')
    @login_required
    @admin_required
    def export_salary_items_template():
        export_dir = os.path.join(current_dir, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = "工资项导入模板.xlsx"
        filepath = os.path.join(export_dir, filename)
        data = [{'项目名称': '基本工资', '项目类型': '收入项', '计算公式': '', '默认值': 0.0}]
        df = pd.DataFrame(data)
        df.to_excel(filepath, index=False)
        return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/export_salary_template')
    @login_required
    @admin_required
    def export_salary_template():
        export_dir = os.path.join(current_dir, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = "工资导入模板.xlsx"
        filepath = os.path.join(export_dir, filename)
        salary_items = SalaryItem.query.order_by(SalaryItem.order).all()
        item_names = [item.name for item in salary_items]
        df = pd.DataFrame(columns=['员工ID', '月份', '员工姓名', '单位'] + item_names)
        example_employee = Employee.query.first()
        if example_employee:
            example_data = {'员工ID': example_employee.id, '月份': datetime.now().strftime('%Y-%m'), '员工姓名': example_employee.name, '单位': example_employee.unit.name}
            for item in salary_items:
                example_data[item.name] = 0.0
            df = pd.concat([df, pd.DataFrame([example_data])], ignore_index=True)
        df.to_excel(filepath, index=False)
        return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/export_salary_import_template')
    @login_required
    @admin_required
    def export_salary_import_template():
        salary_items = SalaryItem.query.order_by(SalaryItem.order).all()
        income_items = [item.name for item in salary_items if item.item_type == 'income']
        deduction_items = [item.name for item in salary_items if item.item_type == 'deduction']
        all_item_names = income_items + deduction_items
        columns = ['员工ID', '月份', '员工姓名', '单位', '备注'] + all_item_names
        example_employee = Employee.query.first()
        if example_employee:
            example_row = {
                '员工ID': example_employee.id,
                '月份': datetime.now().strftime('%Y-%m'),
                '员工姓名': example_employee.name,
                '单位': example_employee.unit.name,
                '备注': '例如：病假扣款、奖金等'
            }
            for item in all_item_names:
                example_row[item] = 0.0
            df = pd.DataFrame([example_row])
        else:
            df = pd.DataFrame(columns=columns)
        export_dir = os.path.join(current_dir, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = f"工资导入模板_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(export_dir, filename)
        df.to_excel(filepath, index=False)
        return send_file(filepath, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/backup_database')
    @login_required
    @admin_required
    def backup_database():
        """备份数据库文件（下载）"""
        db_path = current_app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')
        if not os.path.isabs(db_path):
            db_path = os.path.join(current_app.instance_path, os.path.basename(db_path))

        if not os.path.exists(db_path):
            flash('数据库文件不存在', 'danger')
            return redirect(url_for('system_settings'))

        backup_filename = f"hr_salary_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        return send_file(db_path, as_attachment=True, download_name=backup_filename)

    @app.route('/restore_database', methods=['POST'])
    @login_required
    @admin_required
    def restore_database():
        """恢复数据库：上传备份文件并替换当前数据库"""
        if 'backup_file' not in request.files:
            return jsonify({'success': False, 'error': '未选择文件'}), 400

        file = request.files['backup_file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '文件名为空'}), 400

        # 限制文件扩展名
        allowed_ext = {'.db', '.sqlite', '.sqlite3'}
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in allowed_ext:
            return jsonify({'success': False, 'error': '只支持 .db, .sqlite, .sqlite3 格式的备份文件'}), 400

        # 获取当前数据库路径
        db_uri = current_app.config['SQLALCHEMY_DATABASE_URI']
        db_path = db_uri.replace('sqlite:///', '')
        if not os.path.isabs(db_path):
            db_path = os.path.join(current_app.instance_path, os.path.basename(db_path))

        # 保存上传的临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        try:
            # 关闭当前所有数据库连接（简单方式：dispose）
            db.engine.dispose()

            # 备份当前数据库（防止恢复失败）
            backup_current = db_path + '.before_restore'
            if os.path.exists(db_path):
                shutil.copy2(db_path, backup_current)

            # 替换数据库文件
            shutil.copy2(tmp_path, db_path)

            # 重新初始化数据库连接（可选，但需要重启应用才完全生效）
            # 这里只是替换了文件，后续请求会重新建立连接
            return jsonify({'success': True, 'message': '数据库已恢复，请重启应用以完全生效'})
        except Exception as e:
            current_app.logger.error(f"数据库恢复失败: {str(e)}")
            return jsonify({'success': False, 'error': str(e)}), 500
        finally:
            # 删除临时文件
            try:
                os.unlink(tmp_path)
            except:
                pass

    @app.route('/maintain_database', methods=['POST'])
    @login_required
    @admin_required
    def maintain_database():
        """执行数据库维护：完整性检查 + VACUUM"""
        try:
            # 执行 PRAGMA integrity_check
            result = db.session.execute(text("PRAGMA integrity_check")).fetchone()
            integrity_ok = (result[0] == 'ok')

            # 执行 VACUUM（使用 db.session）
            db.session.execute(text("VACUUM"))
            db.session.commit()

            return jsonify({
                'success': True,
                'integrity_check': 'ok' if integrity_ok else result[0],
                'vacuum_status': '已完成'
            })
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"数据库维护失败: {str(e)}")
            return jsonify({'success': False, 'error': str(e)}), 500

    # ==================== 工作计划 API ====================
    @app.route('/api/tasks', methods=['GET'])
    @login_required
    def get_tasks():
        tasks = UserTask.query.filter_by(user_id=current_user.id).order_by(
            UserTask.is_completed.asc(),
            UserTask.due_date.asc(),
            UserTask.created_at.desc()
        ).all()
        return jsonify([{
            'id': t.id,
            'title': t.title,
            'due_date': t.due_date.strftime('%Y-%m-%d') if t.due_date else None,
            'is_completed': t.is_completed
        } for t in tasks])

    @app.route('/api/tasks', methods=['POST'])
    @login_required
    def add_task():
        data = request.get_json()
        title = data.get('title', '').strip()
        if not title:
            return jsonify({'error': '标题不能为空'}), 400
        due_date = None
        if data.get('due_date'):
            try:
                due_date = datetime.strptime(data['due_date'], '%Y-%m-%d').date()
            except:
                return jsonify({'error': '日期格式错误，应为YYYY-MM-DD'}), 400
        task = UserTask(user_id=current_user.id, title=title, due_date=due_date)
        db.session.add(task)
        db.session.commit()
        return jsonify({'id': task.id, 'message': '添加成功'}), 201

    @app.route('/api/tasks/<int:task_id>', methods=['PUT'])
    @login_required
    def update_task(task_id):
        task = UserTask.query.filter_by(id=task_id, user_id=current_user.id).first_or_404()
        data = request.get_json()
        if 'title' in data:
            task.title = data['title'].strip()
        if 'due_date' in data:
            if data['due_date']:
                try:
                    task.due_date = datetime.strptime(data['due_date'], '%Y-%m-%d').date()
                except:
                    return jsonify({'error': '日期格式错误'}), 400
            else:
                task.due_date = None
        if 'is_completed' in data:
            task.is_completed = bool(data['is_completed'])
        db.session.commit()
        return jsonify({'message': '更新成功'})

    @app.route('/api/tasks/<int:task_id>', methods=['DELETE'])
    @login_required
    def delete_task(task_id):
        task = UserTask.query.filter_by(id=task_id, user_id=current_user.id).first_or_404()
        db.session.delete(task)
        db.session.commit()
        return jsonify({'message': '删除成功'})
    @app.route('/print_salary', methods=['POST'])
    @login_required
    def print_salary():
        print_title = request.form.get('print_title')
        unit_name = request.form.get('unit_name')
        responsible_person = request.form.get('responsible_person')
        reviewer = request.form.get('reviewer')
        preparer = request.form.get('preparer')
        month = request.form.get('month')
        unit_id = request.form.get('unit_id')
        employee_id = request.form.get('employee_id')
        items = request.form.getlist('items')
        query = SalaryRecord.query
        if month: query = query.filter_by(month=month)
        if unit_id: query = query.join(Employee).filter(Employee.unit_id == unit_id)
        if employee_id: query = query.filter_by(employee_id=employee_id)
        records = query.all()
        if not unit_name and records:
            unit_name = records[0].employee.unit.name if records[0].employee else ""
        return render_template('print_template.html', records=records, print_title=print_title, unit_name=unit_name, responsible_person=responsible_person, reviewer=reviewer, preparer=preparer, items=items)

    # 核心导入/复制/计算路由（支持姓名+单位匹配，自动重算计算项）
    @app.route('/calculate_salary', methods=['GET', 'POST'])
    @login_required
    def run_salary_calculation():
        if request.method == 'POST':
            if not current_user.is_admin:
                flash('您没有权限执行此操作', 'danger')
                return redirect(url_for('run_salary_calculation'))
            action = request.form.get('action')

            # ==================== 导入工资数据 ====================
            if action == 'import':
                try:
                    import_month = request.form['import_month']
                    file = request.files['salary_file']
                    if not file or file.filename == '':
                        flash('没有选择文件', 'danger')
                        return redirect(url_for('run_salary_calculation'))
                    if not (file.filename.lower().endswith('.xlsx') or file.filename.lower().endswith('.xls')):
                        flash('只支持Excel文件(.xlsx, .xls)', 'danger')
                        return redirect(url_for('run_salary_calculation'))
                    filename = secure_filename(file.filename)
                    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
                    file.save(filepath)
                    df = pd.read_excel(filepath)

                    required_columns = ['员工姓名', '单位', '月份']
                    for col in required_columns:
                        if col not in df.columns:
                            flash(f'缺少必要列: {col}', 'danger')
                            return redirect(url_for('run_salary_calculation'))

                    all_salary_items = SalaryItem.query.all()
                    count = 0
                    errors = []

                    for index, row in df.iterrows():
                        try:
                            # ---------- 1. 解析员工、单位、月份 ----------
                            month = str(row['月份']).strip()
                            if not re.match(r'^\d{4}-\d{2}$', month):
                                errors.append(f"行 {index + 2}: 月份格式错误，应为YYYY-MM")
                                continue

                            name = str(row['员工姓名']).strip()
                            unit_name = str(row['单位']).strip()
                            if not name or not unit_name:
                                errors.append(f"行 {index + 2}: 员工姓名或单位为空")
                                continue

                            unit = Unit.query.filter_by(name=unit_name).first()
                            if not unit:
                                units = Unit.query.filter(Unit.name.like(f'%{unit_name}%')).all()
                                if len(units) == 1:
                                    unit = units[0]
                                elif len(units) > 1:
                                    errors.append(
                                        f"行 {index + 2}: 单位名称 '{unit_name}' 匹配到多个单位，请使用更精确的名称")
                                    continue
                                else:
                                    errors.append(f"行 {index + 2}: 单位 '{unit_name}' 不存在")
                                    continue

                            employees_found = Employee.query.filter_by(name=name, unit_id=unit.id).all()
                            if len(employees_found) == 0:
                                errors.append(f"行 {index + 2}: 未找到员工 '{name}' (单位: {unit.name})")
                                continue
                            elif len(employees_found) > 1:
                                errors.append(
                                    f"行 {index + 2}: 单位 '{unit.name}' 中存在多名员工叫 '{name}'，请手动处理")
                                continue
                            else:
                                employee = employees_found[0]

                            # ---------- 2. 获取原有工资记录（如果存在） ----------
                            record = SalaryRecord.query.filter_by(employee_id=employee.id, month=month).first()
                            if record:
                                old_details = record.details
                                if isinstance(old_details, str):
                                    try:
                                        old_details = json.loads(old_details)
                                    except:
                                        old_details = {}
                                elif old_details is None:
                                    old_details = {}
                            else:
                                old_details = {}

                            # ---------- 3. 合并 Excel 数据到 old_details ----------
                            for col in row.index:
                                if col in ['员工姓名', '单位', '月份']:
                                    continue
                                value = row[col]
                                if pd.isna(value):
                                    # 单元格为空：跳过，不修改数据库中原有值
                                    continue
                                try:
                                    value = float(value)
                                except ValueError:
                                    errors.append(f"行 {index + 2}: 列 '{col}' 的值 '{value}' 不是有效数字，已跳过")
                                    continue
                                old_details[col] = value

                            # ---------- 4. 重新计算所有计算项（最多5轮） ----------
                            for _ in range(5):
                                changed = False
                                for item in all_salary_items:
                                    if item.item_type != 'calculation' or not item.formula:
                                        continue
                                    try:
                                        new_val = evaluate_formula(item.formula, old_details)
                                        new_val = round(new_val, 2)
                                        old_val = old_details.get(item.name, 0)
                                        if abs(new_val - old_val) > 0.01:
                                            old_details[item.name] = new_val
                                            changed = True
                                    except Exception as e:
                                        current_app.logger.warning(f"计算公式失败: {item.name}, {e}")
                                if not changed:
                                    break

                            # ---------- 5. 重新计算应发、扣款、实发 ----------
                            gross = 0.0
                            deduct = 0.0
                            item_dict = {item.name: item for item in all_salary_items}
                            for name, val in old_details.items():
                                if name in ['应发工资', '扣款合计', '实发工资']:
                                    continue
                                try:
                                    val = float(val)
                                except:
                                    val = 0.0
                                item_info = item_dict.get(name)
                                if item_info:
                                    if item_info.item_type == 'income':
                                        gross += val
                                    elif item_info.item_type == 'deduction':
                                        deduct += val
                                    elif item_info.item_type == 'calculation':
                                        calc_dir = getattr(item_info, 'calc_direction', 'income')
                                        if calc_dir == 'income':
                                            gross += val
                                        else:
                                            deduct += val
                                else:
                                    if val >= 0:
                                        gross += val
                                    else:
                                        deduct += val
                            net = gross - deduct
                            old_details['应发工资'] = round(gross, 2)
                            old_details['扣款合计'] = round(deduct, 2)
                            old_details['实发工资'] = round(net, 2)

                            # ---------- 6. 读取备注 ----------
                            remark = ''
                            if '备注' in row.index and pd.notna(row['备注']):
                                remark = str(row['备注']).strip()
                            elif record:
                                remark = record.remark if record.remark else ''

                            # ---------- 7. 保存记录 ----------
                            if record:
                                record.details = old_details
                                flag_modified(record, "details")  # ← 必须加上这一行
                                record.total = net
                                record.remark = remark
                            else:
                                record = SalaryRecord(
                                    employee_id=employee.id,
                                    month=month,
                                    details=old_details,
                                    total=net,
                                    remark=remark,
                                    item_remarks='{}'
                                )
                                db.session.add(record)

                            count += 1

                        except Exception as e:
                            errors.append(f"行 {index + 2}: 处理错误 - {str(e)}")
                            continue

                    db.session.commit()
                    if errors:
                        flash(f'成功导入 {count} 条记录，但有 {len(errors)} 个错误: {"；".join(errors[:5])}' + (
                            '...' if len(errors) > 5 else ''), 'warning')
                    else:
                        flash(f'成功导入 {count} 条工资记录', 'success')
                    return redirect(url_for('salary_query', start_month=import_month, end_month=import_month))

                except Exception as e:
                    db.session.rollback()
                    flash(f'导入失败: {str(e)}', 'danger')
                    current_app.logger.error(f"工资导入失败: {str(e)}", exc_info=True)
                    return redirect(url_for('run_salary_calculation'))

            # ==================== 批量操作（复制/计算） ====================
            elif action == 'batch':
                try:
                    source_month = request.form['source_month']
                    target_month = request.form['target_month']
                    batch_action = request.form['batch_action']
                    unit_id = request.form.get('unit_id')

                    if unit_id:
                        employees = Employee.query.filter_by(unit_id=unit_id, active_for_payroll=True).all()
                    else:
                        employees = Employee.query.filter_by(active_for_payroll=True).all()

                    success_count = 0
                    fail_count = 0
                    errors = []

                    for emp in employees:
                        try:
                            if batch_action == 'copy':
                                source_record = SalaryRecord.query.filter_by(employee_id=emp.id,
                                                                             month=source_month).first()
                                if not source_record:
                                    fail_count += 1
                                    errors.append(f"员工 {emp.name} 无 {source_month} 月工资记录，跳过")
                                    continue

                                details = source_record.details
                                if isinstance(details, str):
                                    details = json.loads(details)
                                elif details is None:
                                    details = {}
                                else:
                                    details = details.copy()

                                SalaryRecord.query.filter_by(employee_id=emp.id, month=target_month).delete()

                                new_record = SalaryRecord(
                                    employee_id=emp.id,
                                    month=target_month,
                                    details=details,
                                    total=source_record.total,
                                    remark=getattr(source_record, 'remark', '')
                                )
                                db.session.add(new_record)
                                success_count += 1

                            elif batch_action == 'calculate':
                                # 批量重新计算逻辑（可后续完善，此处暂不实现复杂逻辑）
                                pass

                        except Exception as e:
                            fail_count += 1
                            errors.append(f"员工 {emp.name}: {str(e)}")

                    db.session.commit()

                    batch_result = {
                        'target_month': target_month,
                        'success_count': success_count,
                        'fail_count': fail_count,
                        'errors': errors
                    }
                    units = Unit.query.all()
                    return render_template('calculate_salary.html', units=units, batch_result=batch_result)

                except Exception as e:
                    db.session.rollback()
                    flash(f"批量操作失败: {str(e)}", 'danger')
                    return redirect(url_for('run_salary_calculation'))

            # ==================== 单独计算工资（保留原值，仅重算计算项） ====================
            elif action == 'calculate':
                try:
                    month = request.form['month']
                    unit_id = request.form.get('unit_id')

                    if unit_id:
                        employees = Employee.query.filter_by(unit_id=unit_id, active_for_payroll=True).all()
                    else:
                        employees = Employee.query.filter_by(active_for_payroll=True).all()

                    count = 0
                    errors = []
                    all_salary_items = SalaryItem.query.all()
                    item_dict = {item.name: item for item in all_salary_items}

                    for emp in employees:
                        try:
                            record = SalaryRecord.query.filter_by(employee_id=emp.id, month=month).first()
                            if not record:
                                details = {}
                                for item in all_salary_items:
                                    details[item.name] = item.default_value if item.default_value else 0.0
                                record = SalaryRecord(employee_id=emp.id, month=month, details=details, total=0)
                                db.session.add(record)
                            else:
                                details = record.details
                                if isinstance(details, str):
                                    details = json.loads(details)
                                for item in all_salary_items:
                                    if item.name not in details:
                                        details[item.name] = item.default_value if item.default_value else 0.0

                            # 重新计算计算项（最多5轮）
                            for _ in range(5):
                                changed = False
                                for item in all_salary_items:
                                    if item.item_type != 'calculation' or not item.formula:
                                        continue
                                    try:
                                        new_val = evaluate_formula(item.formula, details)
                                        new_val = round(new_val, 2)
                                        old_val = details.get(item.name, 0)
                                        if abs(new_val - old_val) > 0.01:
                                            details[item.name] = new_val
                                            changed = True
                                    except Exception as e:
                                        current_app.logger.warning(f"计算公式失败: {item.name}, {e}")
                                if not changed:
                                    break

                            # 重新计算应发、扣款、实发
                            gross = 0.0
                            deduct = 0.0
                            for name, val in details.items():
                                if name in ['应发工资', '扣款合计', '实发工资']:
                                    continue
                                try:
                                    val = float(val)
                                except:
                                    val = 0.0
                                info = item_dict.get(name)
                                if info:
                                    if info.item_type == 'income':
                                        gross += val
                                    elif info.item_type == 'deduction':
                                        deduct += val
                                    elif info.item_type == 'calculation':
                                        calc_dir = getattr(info, 'calc_direction', 'income')
                                        if calc_dir == 'income':
                                            gross += val
                                        else:
                                            deduct += val
                                else:
                                    if val >= 0:
                                        gross += val
                                    else:
                                        deduct += val
                            net = gross - deduct
                            details['应发工资'] = round(gross, 2)
                            details['扣款合计'] = round(deduct, 2)
                            details['实发工资'] = round(net, 2)

                            record.details = details
                            record.total = net
                            count += 1

                        except Exception as e:
                            errors.append(f"员工 {emp.name}: {str(e)}")
                            continue

                    db.session.commit()
                    if errors:
                        flash(f"成功计算 {count} 名员工，但有错误：{'；'.join(errors[:5])}", 'warning')
                    else:
                        flash(f"成功重新计算 {count} 名员工工资（已保留原有收入/扣款数据）", 'success')
                    return redirect(url_for('salary_query', month=month))

                except Exception as e:
                    db.session.rollback()
                    flash(f"计算失败: {str(e)}", 'danger')
                    return redirect(url_for('run_salary_calculation'))

        # ==================== GET 请求 ====================
        units = Unit.query.all()
        return render_template('calculate_salary.html', units=units)

    # 其他辅助路由
    @app.route('/get_employees_by_unit')
    @login_required
    def get_employees_by_unit():
        unit_id = request.args.get('unit_id')
        active_only = request.args.get('active_only', 'true').lower() == 'true'
        if unit_id:
            query = Employee.query.filter_by(unit_id=unit_id)
            if active_only:
                query = query.filter_by(active_for_payroll=True)
            employees = query.all()
        else:
            if active_only:
                employees = Employee.query.filter_by(active_for_payroll=True).all()
            else:
                employees = Employee.query.all()
        return jsonify({'employees': [{'id': emp.id, 'name': emp.name.split('(')[0].strip()} for emp in employees]})

    @app.route('/get_employee')
    @login_required
    def get_employee():
        employee_id = request.args.get('employee_id')
        employee = db.session.get(Employee, employee_id)
        if not employee:
            return jsonify({'error': '员工不存在'}), 404
        return jsonify({'employee': {'id': employee.id, 'name': employee.name, 'unit': employee.unit.name, 'id_card': employee.id_card}})

    @app.route('/get_salary_items')
    @login_required
    def get_salary_items():
        items = SalaryItem.query.all()
        return jsonify({'items': [{'id': item.id, 'name': item.name} for item in items]})

    @app.route('/salary_records/<month>')
    @login_required
    def view_salary_records(month):
        page = request.args.get('page', 1, type=int)
        per_page = 20
        records = SalaryRecord.query.join(Employee).filter(SalaryRecord.month == month).order_by(SalaryRecord.id.desc()).paginate(page=page, per_page=per_page)
        return render_template('salary_records.html', records=records, month=month)

    @app.route('/salary_records/edit/<int:record_id>', methods=['GET', 'POST'])
    @login_required
    def edit_salary_record(record_id):
        record = SalaryRecord.query.get_or_404(record_id)
        if request.method == 'POST':
            record.details = request.form.get('details')
            record.total = request.form.get('total')
            db.session.commit()
            flash("工资记录更新成功", "success")
            return redirect(url_for('view_salary_records', month=record.month))
        return render_template('edit_salary_record.html', record=record)

    @app.route('/salary_records/delete/<int:record_id>', methods=['POST'])
    @login_required
    @admin_required  # ← 新增
    def delete_salary_record(record_id):
        record = SalaryRecord.query.get_or_404(record_id)
        month = record.month
        db.session.delete(record)
        db.session.commit()
        flash(f"工资记录已删除", "success")
        return redirect(url_for('salary_query', month=month))

    @app.route('/update_salary_remark/<int:record_id>', methods=['POST'])
    @login_required
    def update_salary_remark(record_id):
        if not current_user.is_admin:
            return jsonify({'success': False, 'error': '无权限'}), 403
        data = request.get_json()
        remark = data.get('remark', '')
        record = SalaryRecord.query.get_or_404(record_id)
        record.remark = remark
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/update_salary_cell/<int:record_id>', methods=['POST'])
    @login_required
    def update_salary_cell(record_id):
        if not current_user.is_admin:
            return jsonify({'success': False, 'error': '无权限'}), 403
        data = request.get_json()
        item_name = data.get('item_name')
        value = data.get('value')
        if not item_name or value is None:
            return jsonify({'success': False, 'error': '缺少参数'}), 400

        # 1. 处理符号：只对系统定义的工资项自动判断，未定义项保留用户输入的符号
        salary_item = SalaryItem.query.filter_by(name=item_name).first()
        if salary_item:
            if salary_item:
                # 不再强制转换符号，保留用户输入的原值
                pass
        else:
            # 未定义项：完全保留用户输入的值（正负由用户决定）
            # 不做任何符号转换
            pass

        record = SalaryRecord.query.get_or_404(record_id)

        # 2. 安全解析 details
        if isinstance(record.details, str):
            try:
                record.details = json.loads(record.details)
            except:
                record.details = {}
        if not isinstance(record.details, dict):
            record.details = {}

        # 3. 更新指定项
        record.details[item_name] = float(value)

        # 4. 获取所有工资项
        all_items = SalaryItem.query.all()
        item_dict = {item.name: item for item in all_items}

        # 5. 重算所有计算项（最多5轮）
        for _ in range(5):
            changed = False
            for item in all_items:
                if item.item_type != 'calculation' or not item.formula:
                    continue
                variables = {}
                for name, val in record.details.items():
                    variables[name] = val
                try:
                    new_val = evaluate_formula(item.formula, variables)
                    new_val = round(new_val, 2)
                    old_val = record.details.get(item.name, 0)
                    if abs(new_val - old_val) > 0.01:
                        record.details[item.name] = new_val
                        changed = True
                except Exception as e:
                    current_app.logger.warning(f"计算项 {item.name} 公式计算失败: {e}")
            if not changed:
                break

        # 6. 重新计算应发、扣款、实发
        gross = 0.0
        deduct = 0.0
        item_dict = {item.name: item for item in SalaryItem.query.all()}
        for name, val in record.details.items():
            if name in ['应发工资', '扣款合计', '实发工资']:
                continue
            try:
                val = float(val)
            except:
                val = 0.0

            item_info = item_dict.get(name)
            if item_info:
                if item_info.item_type == 'income':
                    gross += val
                elif item_info.item_type == 'deduction':
                    deduct += val  # 修复点：保留符号
                elif item_info.item_type == 'calculation':
                    calc_dir = getattr(item_info, 'calc_direction', 'income')
                    if calc_dir == 'income':
                        gross += val
                    else:
                        deduct += val
            else:
                if val >= 0:
                    gross += val
                else:
                    deduct += val

        net = gross - deduct
        record.details['应发工资'] = round(gross, 2)
        record.details['扣款合计'] = round(deduct, 2)
        record.details['实发工资'] = round(net, 2)
        record.total = net
        record.gross_salary = round(gross, 2)
        record.deductions = round(deduct, 2)
        record.net_salary = round(net, 2)

        # 标记 JSON 字段已修改
        flag_modified(record, "details")
        db.session.commit()
        return jsonify({'success': True})
    @app.route('/export_salary/<int:record_id>')
    @login_required
    def export_salary(record_id):
        record = SalaryRecord.query.get_or_404(record_id)
        employee = db.session.get(Employee, record.employee_id)
        unit = db.session.get(Unit, employee.unit_id)
        export_dir = os.path.join(current_dir, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = f"{employee.name}_{record.month}_工资表.xlsx"
        filepath = os.path.join(export_dir, filename)
        signers = [{"name": "财务主管", "title": "财务部"}, {"name": "单位负责人", "title": unit.name}]
        export_salary_table(record, employee, signers, filepath)
        return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def generate_salary_html(record, employee, item_names=None, special_template_ids=None):
        """
        生成工资条 HTML 邮件正文（美化版，每月不同风格）
        """
        from sqlalchemy import or_
        all_items = SalaryItem.query.order_by(SalaryItem.order).all()
        item_dict = {item.name: item for item in all_items}

        # 解析 details
        details = record.details
        if isinstance(details, str):
            try:
                details = json.loads(details)
            except:
                details = {}
        if not isinstance(details, dict):
            details = {}

        gross = float(details.get('应发工资', 0))
        deduct = float(details.get('扣款合计', 0))
        net = float(details.get('实发工资', 0))

        # ------------------- 月度主题定义 -------------------
        def get_month_theme(month_num):
            """根据月份数字返回主题色、背景、图标等"""
            themes = {
                1: {'primary': '#c0392b', 'light': '#f9ebea', 'gradient': 'linear-gradient(135deg, #c0392b, #e74c3c)',
                    'icon': '🎉', 'greeting': '新年快乐！', 'accent': '#e74c3c'},
                2: {'primary': '#e91e63', 'light': '#fce4ec', 'gradient': 'linear-gradient(135deg, #e91e63, #f06292)',
                    'icon': '❤️', 'greeting': '情人节快乐', 'accent': '#f06292'},
                3: {'primary': '#4caf50', 'light': '#e8f5e9', 'gradient': 'linear-gradient(135deg, #4caf50, #81c784)',
                    'icon': '🌸', 'greeting': '春暖花开', 'accent': '#81c784'},
                4: {'primary': '#ff9800', 'light': '#fff3e0', 'gradient': 'linear-gradient(135deg, #ff9800, #ffb74d)',
                    'icon': '🌼', 'greeting': '四月你好', 'accent': '#ffb74d'},
                5: {'primary': '#9c27b0', 'light': '#f3e5f5', 'gradient': 'linear-gradient(135deg, #9c27b0, #ce93d8)',
                    'icon': '🌺', 'greeting': '劳动光荣', 'accent': '#ce93d8'},
                6: {'primary': '#2196f3', 'light': '#e3f2fd', 'gradient': 'linear-gradient(135deg, #2196f3, #64b5f6)',
                    'icon': '☀️', 'greeting': '夏至快乐', 'accent': '#64b5f6'},
                7: {'primary': '#f44336', 'light': '#ffebee', 'gradient': 'linear-gradient(135deg, #f44336, #e57373)',
                    'icon': '🏖️', 'greeting': '暑期愉快', 'accent': '#e57373'},
                8: {'primary': '#ff5722', 'light': '#fbe9e7', 'gradient': 'linear-gradient(135deg, #ff5722, #ff8a65)',
                    'icon': '🍉', 'greeting': '盛夏热情', 'accent': '#ff8a65'},
                9: {'primary': '#8bc34a', 'light': '#f1f8e9', 'gradient': 'linear-gradient(135deg, #8bc34a, #aed581)',
                    'icon': '🍂', 'greeting': '金秋九月', 'accent': '#aed581'},
                10: {'primary': '#ffc107', 'light': '#fff8e1', 'gradient': 'linear-gradient(135deg, #ffc107, #ffd54f)',
                     'icon': '🎃', 'greeting': '国庆快乐', 'accent': '#ffd54f'},
                11: {'primary': '#795548', 'light': '#efebe9', 'gradient': 'linear-gradient(135deg, #795548, #a1887f)',
                     'icon': '🍁', 'greeting': '秋收冬藏', 'accent': '#a1887f'},
                12: {'primary': '#9e9e9e', 'light': '#f5f5f5', 'gradient': 'linear-gradient(135deg, #9e9e9e, #bdbdbd)',
                     'icon': '🎄', 'greeting': '圣诞快乐', 'accent': '#bdbdbd'},
            }
            return themes.get(month_num, themes[1])  # 默认1月风格

        # 获取月份主题
        month_str = record.month  # 格式 YYYY-MM
        month_num = int(month_str[5:7])
        theme = get_month_theme(month_num)

        # 构建工资表格行（收入项）
        salary_rows = []
        # 收入项
        income_items = [item for item in all_items if item.item_type == 'income']
        deduction_items = [item for item in all_items if item.item_type == 'deduction']
        # 如果指定了 item_names，则按指定列表显示（保留原逻辑）
        if item_names is not None:
            # 指定列模式：只显示 item_names 中的项，过滤 0 值
            for name in item_names:
                if name in ['应发工资', '扣款合计', '实发工资']:
                    continue
                val = details.get(name, 0.0)
                try:
                    val = float(val)
                except:
                    val = 0.0
                if abs(val) < 0.005:
                    continue
                salary_rows.append(
                    f'<tr><td style="padding:8px; border-bottom:1px solid #eee;">{name}</td><td style="padding:8px; text-align:right; border-bottom:1px solid #eee;">¥ {val:,.2f}</td></tr>')
        else:
            # 默认模式：显示全部非0收入项和扣款项
            for item in income_items:
                val = details.get(item.name, 0.0)
                try:
                    val = float(val)
                except:
                    val = 0.0
                if abs(val) < 0.005:
                    continue
                salary_rows.append(
                    f'<tr><td style="padding:8px; border-bottom:1px solid #eee;">{item.name}</td><td style="padding:8px; text-align:right; border-bottom:1px solid #eee;">¥ {val:,.2f}</td></tr>')
            for item in deduction_items:
                val = details.get(item.name, 0.0)
                try:
                    val = float(val)
                except:
                    val = 0.0
                if abs(val) < 0.005:
                    continue
                salary_rows.append(
                    f'<tr><td style="padding:8px; border-bottom:1px solid #eee;">{item.name}</td><td style="padding:8px; text-align:right; border-bottom:1px solid #eee;">¥ {val:,.2f}</td></tr>')

        # 汇总行
        summary_rows = f'''
                <tr style="background-color:#f8f9fa; font-weight:bold;">
                    <td style="padding:8px; border-top:2px solid {theme['primary']};">应发工资</td>
                    <td style="padding:8px; text-align:right; border-top:2px solid {theme['primary']};">¥ {gross:,.2f}</td>
                </tr>
                <tr style="background-color:#f8f9fa;">
                    <td style="padding:8px;">扣款合计</td>
                    <td style="padding:8px; text-align:right;">¥ {deduct:,.2f}</td>
                </tr>
                <tr style="background-color:#e9f7ef; font-weight:bold;">
                    <td style="padding:8px; border-top:2px solid {theme['primary']};">实发工资</td>
                    <td style="padding:8px; text-align:right; border-top:2px solid {theme['primary']}; color: #2e7d32;">¥ {net:,.2f}</td>
                </tr>
            '''

        # 仿真进度条（实发/应发占比，仅当应发>0）
        progress_bar = ''
        if gross > 0:
            percent = min(100, int((net / gross) * 100))
            progress_bar = f'''
                    <div style="margin:15px 0; background:#e0e0e0; border-radius:10px; height:12px;">
                        <div style="width:{percent}%; background:{theme['primary']}; border-radius:10px; height:12px;"></div>
                    </div>
                    <p style="text-align:center; font-size:12px; color:#555;">实发占比 {percent}%</p>
                '''

        # 特殊事项表格
        special_html = ""
        if special_template_ids:
            year = int(record.month[:4])
            month = int(record.month[5:7])
            templates = SpecialItemTemplate.query.filter(SpecialItemTemplate.id.in_(special_template_ids)).all()
            all_grants = []
            for tpl in templates:
                query = EmployeeSpecialGrant.query.filter(
                    EmployeeSpecialGrant.employee_id == employee.id,
                    EmployeeSpecialGrant.template_id == tpl.id,
                    EmployeeSpecialGrant.year == year
                )
                if tpl.frequency == 'quarterly':
                    query = query.order_by(EmployeeSpecialGrant.id.desc()).limit(1)
                elif tpl.frequency == 'monthly':
                    query = query.filter(EmployeeSpecialGrant.month == month)
                else:
                    query = query.filter(EmployeeSpecialGrant.month.is_(None))
                grants = query.all()
                all_grants.extend(grants)
            if all_grants:
                total_special = 0.0
                special_rows = []
                for grant in all_grants:
                    template = db.session.get(SpecialItemTemplate, grant.template_id)
                    if not template:
                        continue
                    amount = grant.amount
                    total_special += amount
                    remark = grant.remark or ''
                    special_rows.append(
                        f'<tr><td style="padding:8px; border-bottom:1px solid #eee;">{template.name}</td><td style="padding:8px; text-align:right; border-bottom:1px solid #eee;">¥ {amount:,.2f}</td><td style="padding:8px; border-bottom:1px solid #eee;">{remark}</td></tr>')
                special_html = f'''
                        <div style="margin-top:25px;">
                            <h3 style="font-size:16px; color:{theme['primary']}; border-left:4px solid {theme['primary']}; padding-left:12px;">✨ 特殊事项明细</h3>
                            <table style="width:100%; border-collapse:collapse; margin:12px 0;">
                                <thead><tr style="background:#f2f2f2;"><th style="padding:8px; text-align:left;">事项名称</th><th style="padding:8px; text-align:right;">金额(元)</th><th style="padding:8px; text-align:left;">备注</th></tr></thead>
                                <tbody>{"".join(special_rows)}</tbody>
                                <tfoot><tr style="background:#f9f9f9; font-weight:bold;"><td style="padding:8px;">特殊事项合计</td><td style="padding:8px; text-align:right;">¥ {total_special:,.2f}</td><td style="padding:8px;"></td></tr></tfoot>
                            </table>
                        </div>
                    '''

        # 备注
        remark = getattr(record, 'remark', '') or ''
        remark_html = f'<div style="margin-top:20px; padding:10px; background:{theme["light"]}; border-left:4px solid {theme["primary"]}; border-radius:6px;"><strong>📝 备注：</strong> {remark}</div>' if remark else ''

        # 完整HTML（内联样式，适配邮件客户端）
        html = f"""<!DOCTYPE html>
    <html>
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{record.month} 工资条</title>
    </head>
    <body style="margin:0; padding:20px; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; background:#f4f7fb;">
    <div style="max-width:600px; margin:0 auto; background:white; border-radius:24px; box-shadow:0 8px 20px rgba(0,0,0,0.05); overflow:hidden;">

        <!-- 头部主题区 -->
        <div style="background:{theme['gradient']}; padding:24px 20px; text-align:center; color:white;">
            <div style="font-size:48px; line-height:1;">{theme['icon']}</div>
            <h1 style="margin:12px 0 6px 0; font-size:26px; font-weight:600;">{record.month} 工资条</h1>
            <p style="margin:0; opacity:0.9;">{theme['greeting']}</p>
        </div>

        <!-- 员工信息卡片 -->
        <div style="padding:20px 24px; background:{theme['light']}; border-bottom:1px solid #ddd;">
            <table style="width:100%; border-collapse:collapse;">
                <tr>
                    <td style="padding:4px 0; width:80px;"><strong>姓名</strong></td>
                    <td style="padding:4px 0;">{employee.name}</td>
                </tr>
                <tr>
                    <td style="padding:4px 0;"><strong>部门</strong></td>
                    <td style="padding:4px 0;">{employee.unit.name if employee.unit else ''}</td>
                </tr>
                <tr>
                    <td style="padding:4px 0;"><strong>人员类型</strong></td>
                    <td style="padding:4px 0;">{employee.employee_type or ''}</td>
                </tr>
            </table>
        </div>

        <!-- 工资明细表格 -->
        <div style="padding:20px 24px;">
            <h3 style="font-size:16px; color:{theme['primary']}; border-left:4px solid {theme['primary']}; padding-left:12px; margin:0 0 12px 0;">💰 工资项目明细</h3>
            <table style="width:100%; border-collapse:collapse;">
                <tbody>
                    {''.join(salary_rows) if salary_rows else '<tr><td colspan="2" style="padding:16px; text-align:center;">无明细项目</td></tr>'}
                    {summary_rows}
                </tbody>
            </table>
            {progress_bar}
        </div>

        <!-- 特殊事项 -->
        {special_html}

        <!-- 备注 -->
        {remark_html}

        <!-- 脚注 -->
        <div style="padding:16px 24px; background:#f9f9f9; text-align:center; font-size:12px; color:#888; border-top:1px solid #eee;">
            <p>此邮件为系统自动发送，请勿直接回复<br>如有疑问，请联系办公室</p>
            <p style="margin-top:8px;">发送时间：{datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}</p>
        </div>
    </div>
    </body>
    </html>"""
        return html


    def get_default_salary_sheet_columns():
        """获取默认工资发放表打印模板中配置的工资项列名（过滤掉非工资项）"""
        from models import PrintTemplate
        tpl = PrintTemplate.query.filter_by(template_type='salary_sheet', is_default=True).first()
        if tpl and tpl.config:
            try:
                config = json.loads(tpl.config)
                cols = config.get('columns', [])
                # 获取所有工资项名称
                salary_item_names = {item.name for item in SalaryItem.query.all()}
                # 只保留属于工资项的名称（过滤掉“序号”“姓名”“备注”等）
                return [c for c in cols if c in salary_item_names]
            except:
                pass
        return None

    @app.route('/batch_send_emails', methods=['POST'])
    @login_required
    @admin_required
    def batch_send_emails():
        from models import EmailConfig
        record_ids = request.form.getlist('record_ids')
        month = request.form['month']
        item_names_json = request.form.get('item_names', '')

        # 工资项列表
        item_names = None
        if item_names_json:
            try:
                item_names = json.loads(item_names_json)
            except:
                pass
        if not item_names:
            item_names = get_default_salary_sheet_columns()

        config = EmailConfig.query.first()
        interval = config.send_interval if config else 10
        success_count = 0
        fail_count = 0
        errors = []
        results = []

        for idx, record_id in enumerate(record_ids, 1):
            record = db.session.get(SalaryRecord, record_id)
            if not record or not record.employee:
                fail_count += 1
                errors.append(f"记录ID {record_id} 无效")
                continue
            try:
                if not record.employee.email:
                    raise ValueError(f"员工 {record.employee.name} 未设置邮箱")
                subject = f"{record.month}工资条"
                # 不再包含特殊事项
                html_body = generate_salary_html(record, record.employee, item_names=item_names,
                                                 special_template_ids=None)
                send_email_with_attachment(to_email=record.employee.email, subject=subject, body=html_body,
                                           is_html=True)
                success_count += 1
                results.append(f"成功发送给 {record.employee.name}")
            except Exception as e:
                error_msg = f"{record.employee.name}: {str(e)}"
                errors.append(error_msg)
                fail_count += 1
            if idx < len(record_ids):
                time.sleep(interval)

        return jsonify({'success': True, 'message': f"批量发送完成: 成功 {success_count} 封, 失败 {fail_count} 封",
                        'errors': errors, 'results': results})

    @app.route('/api/send_special_email', methods=['POST'])
    @login_required
    @admin_required
    def send_special_email():
        """根据打印模板配置发送特殊事项邮件给员工"""
        from models import PrintTemplate, EmailConfig
        data = request.get_json()
        template_id = data.get('template_id')
        year = data.get('year')
        month = data.get('month')  # 格式 YYYY-MM
        quarter = data.get('quarter')
        unit_id = data.get('unit_id')
        employee_ids = data.get('employee_ids', [])
        columns = data.get('columns', [])

        if not template_id:
            return jsonify({'success': False, 'error': '缺少模板ID'}), 400

        print_template = db.session.get(PrintTemplate, template_id)
        if not print_template:
            return jsonify({'success': False, 'error': '打印模板不存在'}), 404

        config = json.loads(print_template.config) if print_template.config else {}
        special_template_id = config.get('special_template_id')
        if not special_template_id:
            return jsonify({'success': False, 'error': '打印模板未关联特殊事项模板'}), 400

        special_tpl = db.session.get(SpecialItemTemplate, special_template_id)
        if not special_tpl:
            return jsonify({'success': False, 'error': '关联的特殊事项模板不存在'}), 404

        # 查询 grants 并处理时间参数
        query = EmployeeSpecialGrant.query.join(Employee, EmployeeSpecialGrant.employee_id == Employee.id) \
            .filter(EmployeeSpecialGrant.template_id == special_template_id, Employee.active_for_payroll == True)

        if month:
            try:
                y, m = month.split('-')
                query = query.filter(EmployeeSpecialGrant.year == int(y), EmployeeSpecialGrant.month == int(m))
            except:
                return jsonify({'success': False, 'error': '月份格式错误'}), 400
        elif quarter:
            if quarter not in ['Q1', 'Q2', 'Q3', 'Q4']:
                return jsonify({'success': False, 'error': '无效的季度'}), 400
            q_months_map = {'Q1': [1, 2, 3], 'Q2': [4, 5, 6], 'Q3': [7, 8, 9], 'Q4': [10, 11, 12]}
            query = query.filter(EmployeeSpecialGrant.year == int(year),
                                 EmployeeSpecialGrant.month.in_(q_months_map[quarter]))
        elif year:
            query = query.filter(EmployeeSpecialGrant.year == int(year))
            if special_tpl.frequency not in ('quarterly',):
                query = query.filter(EmployeeSpecialGrant.month.is_(None))
        else:
            return jsonify({'success': False, 'error': '缺少时间参数'}), 400

        if unit_id:
            query = query.filter(Employee.unit_id == unit_id)
        if employee_ids:
            query = query.filter(Employee.id.in_(employee_ids))

        grants = query.all()

        if not grants:
            return jsonify({
                'success': False,
                'error': '未找到符合条件的特殊事项发放记录，请先录入数据或检查筛选条件。'
            }), 404

        # ========== 季度专用发送逻辑 ==========
        if quarter:
            q_months = q_months_map[quarter]
            emp_data = {}
            for g in grants:
                eid = g.employee_id
                if eid not in emp_data:
                    emp_data[eid] = {
                        'assessments': {m: '' for m in q_months},
                        'amounts': {m: 0.0 for m in q_months},
                        'remark': g.remark or ''
                    }
                extra = json.loads(g.extra_data) if g.extra_data else {}
                emp_data[eid]['assessments'][g.month] = extra.get('assessment_value', '')
                emp_data[eid]['amounts'][g.month] = g.amount
                if g.remark:
                    emp_data[eid]['remark'] = g.remark

            config_email = EmailConfig.query.first()
            interval = config_email.send_interval if config_email else 10
            success_count = 0
            fail_count = 0
            errors = []
            for idx, (eid, data) in enumerate(emp_data.items()):
                employee = db.session.get(Employee, eid)
                if not employee or not employee.email:
                    fail_count += 1
                    errors.append(f"{employee.name if employee else '未知员工'} 未设置邮箱")
                    continue
                try:
                    total = sum(data['amounts'].values())
                    html = f"""<html><head><style>
                        body {{ font-family: Arial, sans-serif; line-height: 1.6; }}
                        .container {{ max-width: 800px; margin: 0 auto; padding: 20px; }}
                        .header {{ text-align: center; margin-bottom: 30px; }}
                        .footer {{ margin-top: 30px; font-size: 0.9em; color: #666; }}
                        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                        th, td {{ padding: 8px; text-align: center; border: 1px solid #ddd; }}
                        th {{ background-color: #f2f2f2; }}
                    </style></head><body><div class="container">
                    <div class="header"><h2>{year}年{quarter} {special_tpl.name}发放明细</h2></div>
                    <p>尊敬的{employee.name}：</p>
                    <p>您好！以下是您的 {special_tpl.name} 明细：</p>
                    <table>
                        <thead><tr>
                            <th>1月考核</th><th>1月金额(元)</th>
                            <th>2月考核</th><th>2月金额(元)</th>
                            <th>3月考核</th><th>3月金额(元)</th>
                            <th>合计(元)</th><th>备注</th>
                        </tr></thead>
                        <tbody><tr>
                            <td>{data['assessments'][1]}</td><td>{data['amounts'][1]:,.2f}</td>
                            <td>{data['assessments'][2]}</td><td>{data['amounts'][2]:,.2f}</td>
                            <td>{data['assessments'][3]}</td><td>{data['amounts'][3]:,.2f}</td>
                            <td><strong>{total:,.2f}</strong></td>
                            <td>{data['remark']}</td>
                        </tr></tbody>
                    </table>
                    <div class="footer"><p>注：此邮件为系统自动发送，请勿直接回复</p><p>如有疑问，请联系办公室</p><p>发送时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p></div></div></body></html>"""

                    subject = f"{year}年{quarter} {special_tpl.name}发放明细"
                    send_email_with_attachment(to_email=employee.email, subject=subject, body=html, is_html=True)
                    success_count += 1
                except Exception as e:
                    fail_count += 1
                    errors.append(f"{employee.name}: {str(e)}")

                if idx < len(emp_data) - 1:
                    time.sleep(interval)

            return jsonify({
                'success': True,
                'message': f'发送完成：成功 {success_count} 封，失败 {fail_count} 封',
                'errors': errors
            })

        # ========== 月度和年度发送逻辑（原有逻辑，按员工分组逐条发送） ==========
        emp_grants = defaultdict(list)
        for grant in grants:
            emp_grants[grant.employee_id].append(grant)

        config_email = EmailConfig.query.first()
        interval = config_email.send_interval if config_email else 10
        success_count = 0
        fail_count = 0
        errors = []

        for idx, (emp_id, grant_list) in enumerate(emp_grants.items()):
            employee = db.session.get(Employee, emp_id)
            if not employee or not employee.email:
                fail_count += 1
                errors.append(f"{employee.name if employee else '未知员工'} 未设置邮箱")
                continue
            try:
                html = _build_special_email_html(grant_list, employee, special_tpl, columns, config)
                subject = f"{year}年{month or '年度'} {special_tpl.name}发放明细"
                send_email_with_attachment(to_email=employee.email, subject=subject, body=html, is_html=True)
                success_count += 1
            except Exception as e:
                fail_count += 1
                errors.append(f"{employee.name}: {str(e)}")

            if idx < len(emp_grants) - 1:
                time.sleep(interval)

        return jsonify({
            'success': True,
            'message': f'发送完成：成功 {success_count} 封，失败 {fail_count} 封',
            'errors': errors
        })
    def _build_special_email_html(grants, employee, template, columns, config=None):
        """生成单个员工的特殊事项邮件 HTML"""
        # 确定列
        if not columns:
            # 默认显示金额和备注
            columns = [{'name': '金额(元)', 'display': '金额'}, {'name': '备注', 'display': '备注'}]

        # 构建表格行
        rows_html = ''
        total_amount = 0.0
        for g in grants:
            extra = json.loads(g.extra_data) if g.extra_data else {}
            row = '<tr>'
            for col in columns:
                col_name = col['name']
                if col_name == '金额(元)' or col_name == '金额':
                    val = g.amount
                    total_amount += val
                    row += f'<td style="text-align:right">{val:,.2f}</td>'
                elif col_name == '备注':
                    row += f'<td>{g.remark or ""}</td>'
                else:
                    # 从 extra_data 取自定义字段
                    val = extra.get(col_name, '')
                    row += f'<td>{val}</td>'
            row += '</tr>'
            rows_html += row

        if len(grants) > 1:
            rows_html += f'<tr style="font-weight:bold;background:#f0f0f0"><td colspan="{len(columns)}">合计</td><td style="text-align:right">{total_amount:,.2f}</td></tr>'

        table_html = f'''
        <table style="width:100%; border-collapse:collapse; margin:15px 0;">
            <thead><tr style="background:#f0f0f0;">
                {''.join(f'<th>{col["display"]}</th>' for col in columns)}
            </tr></thead>
            <tbody>{rows_html}</tbody>
        </table>'''

        html = f"""<html><head><style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; }}
                .container {{ max-width: 800px; margin: 0 auto; padding: 20px; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .footer {{ margin-top: 30px; font-size: 0.9em; color: #666; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ padding: 8px; text-align: left; border: 1px solid #ddd; }}
                th {{ background-color: #f2f2f2; }}
        </style></head><body><div class="container">
        <div class="header"><h2>{template.name} 发放明细</h2></div>
        <p>尊敬的{employee.name}：</p>
        <p>您好！以下是您的 {template.name} 明细：</p>
        {table_html}
        <div class="footer"><p>注：此邮件为系统自动发送，请勿直接回复</p><p>如有疑问，请联系办公室</p><p>发送时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p></div></div></body></html>"""
        return html
    @app.route('/debug_details')
    def debug_details():
        from models import SalaryRecord
        record = SalaryRecord.query.filter_by(employee_id=1, month='2025-01').first()
        if record:
            return f"details = {record.details}"
        else:
            return "No record found"

    @app.route('/debug_salary/<int:employee_id>/<month>')
    @login_required
    def debug_salary(employee_id, month):
        record = SalaryRecord.query.filter_by(employee_id=employee_id, month=month).first()
        if record:
            return jsonify({
                'employee_id': record.employee_id,
                'month': record.month,
                'details': record.details,
                'total': record.total,
                'remark': record.remark
            })
        else:
            return jsonify({'error': 'Record not found'}), 404

    @app.route('/api/comprehensive_report', methods=['POST'])
    @login_required
    def comprehensive_report():
        data = request.get_json()
        if not data:
            return jsonify({'error': '无效的请求数据'}), 400

        start_month = data.get('start_month')
        end_month = data.get('end_month')
        unit_id = data.get('unit_id')
        group_by = data.get('group_by', 'employee')
        employee_ids = data.get('employee_ids', [])
        salary_item_names = data.get('salary_items', [])
        special_template_ids = data.get('special_templates', [])

        current_app.logger.info(
            f"综合报表请求: start={start_month}, end={end_month}, unit={unit_id}, "
            f"group={group_by}, emp_ids={employee_ids}, salary={salary_item_names}, special={special_template_ids}"
        )

        if not start_month or not end_month:
            return jsonify({'error': '缺少起止月份'}), 400
        if not salary_item_names and not special_template_ids:
            return jsonify({'error': '请至少选择一个工资项或特殊事项'}), 400

        # 员工过滤：只选择参与工资计算的员工
        employee_query = Employee.query.filter_by(active_for_payroll=True)
        if employee_ids:
            emp_ids = [int(eid) for eid in employee_ids if eid]
            if emp_ids:
                employee_query = employee_query.filter(Employee.id.in_(emp_ids))
        elif unit_id:
            employee_query = employee_query.filter_by(unit_id=unit_id)

        employees = employee_query.all()
        if not employees:
            return jsonify({'columns': [], 'rows': [], 'totals': {}})

        employee_map = {emp.id: emp for emp in employees}
        unit_map = {unit.id: unit.name for unit in Unit.query.all()}

        # 常规工资记录
        salary_records = SalaryRecord.query.filter(
            SalaryRecord.month >= start_month,
            SalaryRecord.month <= end_month,
            SalaryRecord.employee_id.in_(employee_map.keys())
        ).all()

        # 特殊事项
        special_grants = EmployeeSpecialGrant.query.filter(
            EmployeeSpecialGrant.year >= int(start_month[:4]),
            EmployeeSpecialGrant.year <= int(end_month[:4]),
            EmployeeSpecialGrant.employee_id.in_(employee_map.keys())
        )
        if special_template_ids:
            special_grants = special_grants.filter(EmployeeSpecialGrant.template_id.in_(special_template_ids))
        special_grants = special_grants.all()

        template_map = {}
        if special_template_ids:
            # 按 order 字段升序排列，保持与拖拽排序后的顺序一致
            templates = SpecialItemTemplate.query.filter(
                SpecialItemTemplate.id.in_(special_template_ids)
            ).order_by(SpecialItemTemplate.order).all()
            template_map = {t.id: t.name for t in templates}

        if group_by == 'employee':
            show_unit = data.get('show_unit', False)
            return _build_employee_report(employees, salary_records, special_grants,
                                          salary_item_names, template_map, start_month, end_month, show_unit)
        elif group_by == 'unit':
            return _build_unit_report(employees, salary_records, special_grants,
                                      salary_item_names, template_map, unit_map, start_month, end_month)
        elif group_by == 'month':
            return _build_monthly_report(employees, salary_records, special_grants,
                                         salary_item_names, template_map, start_month, end_month)

    def _build_employee_report(employees, salary_records, special_grants, salary_item_names, template_map, start_month,
                               end_month, show_unit=False):
        """按员工汇总（收入项加、扣款项减）"""
        # 分离收入项和扣款项
        all_items = SalaryItem.query.all()
        item_type_map = {item.name: item.item_type for item in all_items}
        calc_dir_map = {item.name: getattr(item, 'calc_direction', 'income') for item in all_items if
                        item.item_type == 'calculation'}

        income_items = []
        deduction_items = []
        for name in salary_item_names:
            itype = item_type_map.get(name, None)
            if itype == 'income':
                income_items.append(name)
            elif itype == 'deduction':
                deduction_items.append(name)
            elif itype == 'calculation':
                cdir = calc_dir_map.get(name, 'income')
                if cdir == 'income':
                    income_items.append(name)
                else:
                    deduction_items.append(name)
            else:
                # 未知项按收入处理
                income_items.append(name)

        columns = ['员工ID', '员工姓名']
        if show_unit:
            columns.append('单位')
        # 1. 收入项明细
        columns.extend(income_items)
        # 2. 特殊事项列
        for tpl_name in template_map.values():
            columns.append(f'{tpl_name}(元)')
        # 3. 收入合计
        columns.append('收入合计(元)')
        # 4. 扣款项明细
        columns.extend(deduction_items)
        # 5. 扣款合计
        columns.append('扣款合计(元)')
        # 6. 合计金额(净额)
        columns.append('合计金额(元)')

        # 初始化数据
        emp_data = {}
        for emp in employees:
            row_data = {
                '员工ID': emp.id,
                '员工姓名': emp.name,
            }
            if show_unit:
                row_data['单位'] = emp.unit.name
            row_data.update({item: 0.0 for item in salary_item_names})
            row_data.update({f'{tpl_name}(元)': 0.0 for tpl_name in template_map.values()})
            row_data['收入合计(元)'] = 0.0
            row_data['扣款合计(元)'] = 0.0
            row_data['合计金额(元)'] = 0.0
            emp_data[emp.id] = row_data

        # 处理常规工资记录
        for rec in salary_records:
            if rec.employee_id not in emp_data:
                continue
            details = rec.details
            if isinstance(details, str):
                try:
                    details = json.loads(details)
                except:
                    details = {}
            for item_name in salary_item_names:
                val = details.get(item_name, 0)
                try:
                    val = float(val)
                except:
                    val = 0.0
                is_deduction = (item_type_map.get(item_name) == 'deduction') or (
                        calc_dir_map.get(item_name) == 'deduction')
                if is_deduction:
                    emp_data[rec.employee_id][item_name] += abs(val)
                    emp_data[rec.employee_id]['扣款合计(元)'] += abs(val)
                else:
                    emp_data[rec.employee_id][item_name] += val
                    emp_data[rec.employee_id]['收入合计(元)'] += val

        # 处理特殊事项（直接加入收入合计）
        for grant in special_grants:
            if grant.employee_id not in emp_data:
                continue
            if grant.month:
                month_str = f"{grant.year}-{str(grant.month).zfill(2)}"
                if month_str < start_month or month_str > end_month:
                    continue
            else:
                if grant.year < int(start_month[:4]) or grant.year > int(end_month[:4]):
                    continue
            tpl_name = template_map.get(grant.template_id)
            if not tpl_name:
                continue
            col_name = f'{tpl_name}(元)'
            emp_data[grant.employee_id][col_name] += grant.amount
            emp_data[grant.employee_id]['收入合计(元)'] += grant.amount

        # 构建结果行和合计行
        rows = []
        totals = {col: 0.0 for col in columns if col not in ['员工ID', '员工姓名', '单位']}
        for emp_id, data in emp_data.items():
            income = data['收入合计(元)']
            deduction = data['扣款合计(元)']
            data['合计金额(元)'] = income - deduction
            rows.append(data)
            for col in totals.keys():
                if col in data:
                    totals[col] += data.get(col, 0)

        rows.sort(key=lambda x: x['合计金额(元)'], reverse=True)
        return jsonify({'columns': columns, 'rows': rows, 'totals': totals})

    def _build_unit_report(employees, salary_records, special_grants, salary_item_names, template_map, unit_map,
                           start_month, end_month):
        valid_unit_ids = {emp.unit_id for emp in employees}
        if not valid_unit_ids:
            return jsonify({'columns': [], 'rows': [], 'totals': {}})

        # 分离收入项和扣款项
        all_items = SalaryItem.query.all()
        item_type_map = {item.name: item.item_type for item in all_items}
        calc_dir_map = {item.name: getattr(item, 'calc_direction', 'income') for item in all_items if
                        item.item_type == 'calculation'}

        income_items = []
        deduction_items = []
        for name in salary_item_names:
            itype = item_type_map.get(name, None)
            if itype == 'income':
                income_items.append(name)
            elif itype == 'deduction':
                deduction_items.append(name)
            elif itype == 'calculation':
                cdir = calc_dir_map.get(name, 'income')
                if cdir == 'income':
                    income_items.append(name)
                else:
                    deduction_items.append(name)
            else:
                income_items.append(name)

        columns = ['单位ID', '单位名称']
        columns.extend(income_items)
        for tpl_name in template_map.values():
            columns.append(f'{tpl_name}(元)')
        columns.append('收入合计(元)')
        columns.extend(deduction_items)
        columns.append('扣款合计(元)')
        columns.append('合计金额(元)')

        unit_data = {}
        for unit_id in valid_unit_ids:
            unit_data[unit_id] = {
                '单位ID': unit_id,
                '单位名称': unit_map.get(unit_id, '未知'),
                **{item: 0.0 for item in salary_item_names},
                **{f'{tpl_name}(元)': 0.0 for tpl_name in template_map.values()},
                '收入合计(元)': 0.0,
                '扣款合计(元)': 0.0,
                '合计金额(元)': 0.0
            }

        emp_to_unit = {emp.id: emp.unit_id for emp in employees}

        # 常规工资
        for rec in salary_records:
            unit_id = emp_to_unit.get(rec.employee_id)
            if unit_id not in unit_data:
                continue
            details = rec.details
            if isinstance(details, str):
                try:
                    details = json.loads(details)
                except:
                    details = {}
            for item_name in salary_item_names:
                val = details.get(item_name, 0)
                try:
                    val = float(val)
                except:
                    val = 0.0
                is_deduction = (item_type_map.get(item_name) == 'deduction') or (
                        calc_dir_map.get(item_name) == 'deduction')
                if is_deduction:
                    unit_data[unit_id][item_name] += abs(val)
                    unit_data[unit_id]['扣款合计(元)'] += abs(val)
                else:
                    unit_data[unit_id][item_name] += val
                    unit_data[unit_id]['收入合计(元)'] += val

        # 特殊事项
        for grant in special_grants:
            unit_id = emp_to_unit.get(grant.employee_id)
            if unit_id not in unit_data:
                continue
            if grant.month:
                month_str = f"{grant.year}-{str(grant.month).zfill(2)}"
                if month_str < start_month or month_str > end_month:
                    continue
            else:
                if grant.year < int(start_month[:4]) or grant.year > int(end_month[:4]):
                    continue
            tpl_name = template_map.get(grant.template_id)
            if not tpl_name:
                continue
            col_name = f'{tpl_name}(元)'
            unit_data[unit_id][col_name] += grant.amount
            unit_data[unit_id]['收入合计(元)'] += grant.amount

        rows = []
        totals = {col: 0.0 for col in columns if col not in ['单位ID', '单位名称']}
        for unit_id, data in unit_data.items():
            data['合计金额(元)'] = data['收入合计(元)'] - data['扣款合计(元)']
            rows.append(data)
            for col in totals.keys():
                totals[col] += data.get(col, 0)

        rows.sort(key=lambda x: x['合计金额(元)'], reverse=True)
        return jsonify({'columns': columns, 'rows': rows, 'totals': totals})

    def _build_monthly_report(employees, salary_records, special_grants, salary_item_names, template_map, start_month,
                              end_month):
        # 生成月份列表
        start = datetime.strptime(start_month, '%Y-%m')
        end = datetime.strptime(end_month, '%Y-%m')
        months = []
        current = start
        while current <= end:
            months.append(current.strftime('%Y-%m'))
            current += relativedelta(months=1)

        # 分离收入项和扣款项
        all_items = SalaryItem.query.all()
        item_type_map = {item.name: item.item_type for item in all_items}
        calc_dir_map = {item.name: getattr(item, 'calc_direction', 'income') for item in all_items if
                        item.item_type == 'calculation'}

        income_items = []
        deduction_items = []
        for name in salary_item_names:
            itype = item_type_map.get(name, None)
            if itype == 'income':
                income_items.append(name)
            elif itype == 'deduction':
                deduction_items.append(name)
            elif itype == 'calculation':
                cdir = calc_dir_map.get(name, 'income')
                if cdir == 'income':
                    income_items.append(name)
                else:
                    deduction_items.append(name)
            else:
                income_items.append(name)

        columns = ['月份']
        columns.extend(income_items)
        for tpl_name in template_map.values():
            columns.append(f'{tpl_name}(元)')
        columns.append('收入合计(元)')
        columns.extend(deduction_items)
        columns.append('扣款合计(元)')
        columns.append('合计金额(元)')

        month_data = {}
        for month in months:
            month_data[month] = {
                '月份': month,
                **{item: 0.0 for item in salary_item_names},
                **{f'{tpl_name}(元)': 0.0 for tpl_name in template_map.values()},
                '收入合计(元)': 0.0,
                '扣款合计(元)': 0.0,
                '合计金额(元)': 0.0
            }

        # 常规工资
        for rec in salary_records:
            month = rec.month
            if month not in month_data:
                continue
            details = rec.details
            if isinstance(details, str):
                try:
                    details = json.loads(details)
                except:
                    details = {}
            for item_name in salary_item_names:
                val = details.get(item_name, 0)
                try:
                    val = float(val)
                except:
                    val = 0.0
                is_deduction = (item_type_map.get(item_name) == 'deduction') or (
                        calc_dir_map.get(item_name) == 'deduction')
                if is_deduction:
                    month_data[month][item_name] += abs(val)
                    month_data[month]['扣款合计(元)'] += abs(val)
                else:
                    month_data[month][item_name] += val
                    month_data[month]['收入合计(元)'] += val

        # 特殊事项（仅包含有月份的）
        for grant in special_grants:
            if grant.month is None:
                continue  # 年度事项不纳入月度报表
            # 安全转换：grant.year 和 grant.month 可能为字符串或整数
            try:
                year_int = int(grant.year)
                month_int = int(grant.month)
                month_str = f"{year_int:04d}-{month_int:02d}"
            except (ValueError, TypeError):
                continue  # 格式异常则跳过
            if month_str not in month_data:
                continue
            tpl_name = template_map.get(grant.template_id)
            if not tpl_name:
                continue
            col_name = f'{tpl_name}(元)'
            month_data[month_str][col_name] += grant.amount
            month_data[month_str]['收入合计(元)'] += grant.amount

        rows = []
        totals = {col: 0.0 for col in columns if col != '月份'}
        for month in months:
            data = month_data[month]
            data['合计金额(元)'] = data['收入合计(元)'] - data['扣款合计(元)']
            rows.append(data)
            for col in totals.keys():
                totals[col] += data.get(col, 0)

        return jsonify({'columns': columns, 'rows': rows, 'totals': totals})

    @app.route('/api/export_backup', methods=['GET'])
    @login_required
    @admin_required
    def export_backup():
        """导出年度备份ZIP包（扩展：包含辅助计算表单和批量计算模板）"""
        import zipfile
        import io
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        year = request.args.get('year', type=int)
        if not year:
            return jsonify({'error': '请指定年份'}), 400

        export_employees = request.args.get('employees', 'true').lower() == 'true'
        export_salary = request.args.get('salary', 'true').lower() == 'true'
        export_special = request.args.get('special', 'true').lower() == 'true'
        export_configs = request.args.get('configs', 'false').lower() == 'true'
        # 新增两个导出选项
        export_auxiliary = request.args.get('auxiliary', 'false').lower() == 'true'
        export_batch_templates = request.args.get('batch_templates', 'false').lower() == 'true'

        # 创建内存中的ZIP文件
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:

            # 1. 导出员工花名册
            if export_employees:
                employees = Employee.query.all()
                data = [{
                    'ID': e.id,
                    '姓名': e.name,
                    '单位': e.unit.name if e.unit else '',
                    '身份证号': e.id_card,
                    '性别': e.gender,
                    '出生日期': e.birth_date.strftime('%Y-%m-%d') if e.birth_date else '',
                    '人员类型': e.employee_type,
                    '人员身份': e.employee_identity,
                    '入职日期': e.join_date.strftime('%Y-%m-%d') if e.join_date else '',
                    '岗位级别': e.position_level,
                    '薪资级别': e.salary_level,
                    '银行卡号': e.bank_account,
                    '是否退役军人': '是' if e.is_veteran else '否',
                    '参与工资计算': '是' if e.active_for_payroll else '否',
                    '邮箱': e.email or ''
                } for e in employees]
                df = pd.DataFrame(data)
                excel_buffer = io.BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='员工花名册', index=False)
                zip_file.writestr(f'员工花名册_{year}年.xlsx', excel_buffer.getvalue())

            # 2. 导出工资记录
            if export_salary:
                records = SalaryRecord.query.filter(SalaryRecord.month.like(f'{year}-%')).all()
                if records:
                    # 按月份分组
                    records_by_month = {}
                    for rec in records:
                        month = rec.month
                        if month not in records_by_month:
                            records_by_month[month] = []
                        records_by_month[month].append(rec)

                    for month, month_records in records_by_month.items():
                        data = []
                        for rec in month_records:
                            emp = rec.employee
                            if not emp:
                                continue
                            row = {
                                '员工ID': emp.id,
                                '员工姓名': emp.name,
                                '单位': emp.unit.name if emp.unit else '',
                                '月份': rec.month,
                                '实发工资': rec.total or 0,
                                '备注': rec.remark or ''
                            }
                            # 添加工资项明细
                            if rec.details:
                                if isinstance(rec.details, str):
                                    try:
                                        details = json.loads(rec.details)
                                    except:
                                        details = {}
                                else:
                                    details = rec.details
                                for key, val in details.items():
                                    if key not in ['应发工资', '扣款合计', '实发工资']:
                                        row[key] = val
                            data.append(row)

                        if data:
                            df = pd.DataFrame(data)
                            excel_buffer = io.BytesIO()
                            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                df.to_excel(writer, sheet_name=month.replace('-', '_'), index=False)
                            zip_file.writestr(f'工资记录/{month}_工资表.xlsx', excel_buffer.getvalue())

            # 3. 导出特殊事项
            if export_special:
                grants = EmployeeSpecialGrant.query.filter_by(year=year).all()
                if grants:
                    data = []
                    for g in grants:
                        emp = db.session.get(Employee, g.employee_id)
                        tpl = db.session.get(SpecialItemTemplate, g.template_id)
                        extra = json.loads(g.extra_data) if g.extra_data else {}
                        row = {
                            '员工ID': g.employee_id,
                            '员工姓名': emp.name if emp else '',
                            '单位': emp.unit.name if emp and emp.unit else '',
                            '模板名称': tpl.name if tpl else '',
                            '年份': g.year,
                            '月份': g.month if g.month else '全年',
                            '金额': g.amount,
                            '备注': g.remark or ''
                        }
                        # 添加额外字段
                        for key, val in extra.items():
                            row[f'额外_{key}'] = val
                        data.append(row)

                    if data:
                        df = pd.DataFrame(data)
                        excel_buffer = io.BytesIO()
                        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                            df.to_excel(writer, sheet_name='特殊事项', index=False)
                        zip_file.writestr(f'特殊事项_{year}年.xlsx', excel_buffer.getvalue())

            # 4. 导出系统配置（原有单位、工资项、特殊事项模板）
            if export_configs:
                wb = Workbook()

                # 单位列表
                units = Unit.query.all()
                ws1 = wb.create_sheet('单位列表')
                ws1.append(['ID', '单位名称'])
                for u in units:
                    ws1.append([u.id, u.name])

                # 工资项
                items = SalaryItem.query.all()
                ws2 = wb.create_sheet('工资项')
                ws2.append(['ID', '名称', '类型', '公式', '默认值', '顺序'])
                for i in items:
                    ws2.append([i.id, i.name, i.item_type, i.formula or '', i.default_value, i.order])

                # 特殊事项模板
                templates = SpecialItemTemplate.query.order_by(SpecialItemTemplate.order).all()
                ws3 = wb.create_sheet('特殊事项模板')
                ws3.append(['ID', '名称', '周期', '计算公式', '基础金额'])
                for t in templates:
                    ws3.append([t.id, t.name, t.frequency, t.formula or '', t.base_amount])

                # 删除默认Sheet
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']

                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                zip_file.writestr('系统配置.xlsx', excel_buffer.getvalue())

            # ========== 5. 新增：导出辅助计算表单 ==========
            if export_auxiliary:
                aux_forms = AuxiliaryForm.query.all()
                if aux_forms:
                    data = []
                    for f in aux_forms:
                        data.append({
                            'ID': f.id,
                            '名称': f.name,
                            '描述': f.description,
                            '配置(JSON)': f.config,
                            '创建人ID': f.created_by,
                            '创建时间': f.created_at.strftime('%Y-%m-%d %H:%M:%S') if f.created_at else '',
                            '更新时间': f.updated_at.strftime('%Y-%m-%d %H:%M:%S') if f.updated_at else ''
                        })
                    df = pd.DataFrame(data)
                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='辅助计算表单', index=False)
                    zip_file.writestr('辅助计算表单.xlsx', excel_buffer.getvalue())

            # ========== 6. 新增：导出批量计算模板 ==========
            if export_batch_templates:
                batch_tpls = BatchCalcTemplate.query.all()
                if batch_tpls:
                    data = []
                    for t in batch_tpls:
                        data.append({
                            'ID': t.id,
                            '模板名称': t.name,
                            '年份': t.year,
                            '配置(JSON)': t.config,
                            '创建人ID': t.created_by,
                            '创建时间': t.created_at.strftime('%Y-%m-%d %H:%M:%S') if t.created_at else ''
                        })
                    df = pd.DataFrame(data)
                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='批量计算模板', index=False)
                    zip_file.writestr('批量计算模板.xlsx', excel_buffer.getvalue())

        zip_buffer.seek(0)

        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'工资系统备份_{year}年_{datetime.now().strftime("%Y%m%d")}.zip'
        )

    @app.route('/export_monthly_special_excel')
    @login_required
    def export_monthly_special_excel():
        """后端生成规范的月度特殊事项发放表 Excel"""
        month = request.args.get('month')  # 如 2026-01
        item_id = request.args.get('itemId')
        unit_id = request.args.get('unitId')

        # 调用已有的数据获取函数拿到员工数据（和展示时一样）
        # 直接复用之前用于预览数据的 /api/monthly_special_table 的逻辑
        # 为了避免重复代码，我们可以在函数内部调用 api_monthly_special_table 的查询逻辑
        # 为清晰起见，我们直接写查询步骤：

        from models import Unit, Employee, EmployeeSpecialGrant, SpecialItemTemplate
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        if not month or not item_id or not unit_id:
            return "参数错误", 400

        unit = db.session.get(Unit, unit_id)
        if not unit:
            return "单位不存在", 404

        template = db.session.get(SpecialItemTemplate, item_id)
        if not template or template.frequency != 'monthly':
            return "事项模板不存在或非月度模板", 404

        # 解析月份
        year = int(month[:4])
        month_num = int(month[5:7])

        employees = Employee.query.filter_by(unit_id=unit_id, active_for_payroll=True).all()

        rows = []
        for emp in employees:
            grant = EmployeeSpecialGrant.query.filter_by(
                employee_id=emp.id,
                template_id=template.id,
                year=year,
                month=month_num
            ).first()

            amount = 0.0
            remark = ''
            extra_cols = []
            if grant:
                amount = grant.amount
                remark = grant.remark or ''
                extra = json.loads(grant.extra_data or '{}')
                # 转换数字字符串
                for k, v in extra.items():
                    if isinstance(v, str) and v.replace('.', '', 1).isdigit():
                        extra[k] = float(v) if '.' in v else int(v)
                # 按模板自定义字段顺序取数据
                if template.extra_fields:
                    try:
                        fields = json.loads(template.extra_fields)
                    except:
                        fields = []
                    for f in fields:
                        extra_cols.append(extra.get(f['name'], ''))
                else:
                    extra_cols = []
            else:
                extra_cols = []

            rows.append({
                'name': emp.name,
                'extra_cols': extra_cols,
                'amount': amount,
                'remark': remark
            })

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = f"{unit.name} {year}年{month_num}月"

        # 样式定义
        title_font = Font(name='微软雅黑', bold=True, size=14)
        header_font = Font(name='微软雅黑', bold=True, size=10)
        normal_font = Font(name='微软雅黑', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # 标题行合并
        col_count = 2 + len(template.extra_fields or []) + 2  # 序号、姓名、... 、金额、备注
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        ws['A1'] = f"{unit.name} {year}年{month_num}月 {template.name}发放表"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align

        # 表头
        headers = ['序号', '姓名']
        if template.extra_fields:
            try:
                fields = json.loads(template.extra_fields)
            except:
                fields = []
            for f in fields:
                headers.append(f.get('label', f.get('name')))
        headers.append('金额(元)')
        headers.append('备注')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = header_fill

        # 数据行
        for idx, emp_row in enumerate(rows, 1):
            row_num = idx + 2
            ws.cell(row=row_num, column=1, value=idx).alignment = center_align
            ws.cell(row=row_num, column=2, value=emp_row['name']).alignment = center_align
            ws.cell(row=row_num, column=2).border = thin_border
            ws.cell(row=row_num, column=1).border = thin_border

            col_offset = 3
            # 自定义字段
            for ecol in emp_row['extra_cols']:
                cell = ws.cell(row=row_num, column=col_offset, value=ecol)
                cell.alignment = center_align
                cell.border = thin_border
                col_offset += 1
            # 金额
            cell_amt = ws.cell(row=row_num, column=col_offset, value=emp_row['amount'])
            cell_amt.number_format = '#,##0.00'
            cell_amt.alignment = right_align
            cell_amt.border = thin_border
            col_offset += 1
            # 备注
            cell_remark = ws.cell(row=row_num, column=col_offset, value=emp_row['remark'])
            cell_remark.alignment = center_align
            cell_remark.border = thin_border

        # 合计行
        total_row = len(rows) + 3
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
        ws.cell(row=total_row, column=1, value='合计').font = header_font
        ws.cell(row=total_row, column=1).alignment = center_align
        ws.cell(row=total_row, column=1).border = thin_border
        ws.cell(row=total_row, column=2).border = thin_border

        col_offset = 3
        if template.extra_fields:
            for _ in fields:
                cell = ws.cell(row=total_row, column=col_offset, value='——')
                cell.alignment = center_align
                cell.border = thin_border
                col_offset += 1
        # 金额合计
        total_amount = sum(emp_row['amount'] for emp_row in rows)
        cell_total = ws.cell(row=total_row, column=col_offset, value=total_amount)
        cell_total.number_format = '#,##0.00'
        cell_total.font = header_font
        cell_total.alignment = right_align
        cell_total.border = thin_border
        col_offset += 1
        ws.cell(row=total_row, column=col_offset, value='——').alignment = center_align
        ws.cell(row=total_row, column=col_offset).border = thin_border

        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        for col in range(3, col_count + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

        # 冻结窗口（标题和表头）
        ws.freeze_panes = 'A3'

        # 输出文件
        export_dir = os.path.join(current_dir, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = f"{unit.name}_{year}年{month_num}月_{template.name}_发放表.xlsx"
        filepath = os.path.join(export_dir, filename)
        wb.save(filepath)
        return send_file(filepath, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/export_quarter_special_excel')

    @login_required
    def export_quarter_special_excel():
        quarter = request.args.get('quarter')
        year = request.args.get('year')
        item_id = request.args.get('itemId')
        unit_id = request.args.get('unitId')
        employee_ids_str = request.args.get('employeeIds', '')

        if not quarter or not year or not item_id:
            return "参数错误", 400

        from models import Unit, Employee, EmployeeSpecialGrant, SpecialItemTemplate
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        template = db.session.get(SpecialItemTemplate, item_id)
        if not template:
            return "事项模板不存在", 404

        # 季度月份
        quarter_months = {'Q1': [1, 2, 3], 'Q2': [4, 5, 6], 'Q3': [7, 8, 9], 'Q4': [10, 11, 12]}
        months = quarter_months[quarter]
        month_names = [f'{m}月' for m in months]

        unit_name = ''
        if unit_id:
            unit = db.session.get(Unit, unit_id)
            if unit:
                unit_name = unit.name

        # 获取员工列表
        if employee_ids_str:
            emp_id_list = [int(eid) for eid in employee_ids_str.split(',') if eid.strip()]
            query = Employee.query.filter(Employee.id.in_(emp_id_list))
            if unit_id:
                query = query.filter_by(unit_id=unit_id)
            employees = query.filter_by(active_for_payroll=True).all()
        elif unit_id:
            employees = Employee.query.filter_by(unit_id=unit_id, active_for_payroll=True).all()
        else:
            employees = Employee.query.filter_by(active_for_payroll=True).all()

        # 收集数据（与 /api/quarter_salary_table 逻辑完全一致）
        emp_data = {}  # {emp_id: {...}}
        for emp in employees:
            emp_data[emp.id] = {
                'name': emp.name,
                'assessments': {f'{m}月': '' for m in months},
                'amounts': {f'{m}月': 0.0 for m in months},
                'total': 0.0,
                'remark': ''
            }

        # 查询发放记录
        grants = EmployeeSpecialGrant.query.filter(
            EmployeeSpecialGrant.template_id == item_id,
            EmployeeSpecialGrant.year == int(year),
            EmployeeSpecialGrant.employee_id.in_(emp_data.keys()),
            EmployeeSpecialGrant.month.in_(months)  # 只取季度月份
        ).all()

        for g in grants:
            if g.employee_id not in emp_data:
                continue
            month_key = f'{g.month}月'
            extra = json.loads(g.extra_data or '{}')
            assess_val = extra.get('assessment_value', '')
            emp_data[g.employee_id]['assessments'][month_key] = assess_val
            emp_data[g.employee_id]['amounts'][month_key] = g.amount
            emp_data[g.employee_id]['total'] += g.amount
            emp_data[g.employee_id]['remark'] = g.remark or ''

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = f"{unit_name}_{year}_{quarter}"

        # 样式
        title_font = Font(name='微软雅黑', bold=True, size=14)
        header_font = Font(name='微软雅黑', bold=True, size=10)
        normal_font = Font(name='微软雅黑', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # 表头
        headers = ['序号', '姓名']
        for m in months:
            headers.append(f'{m}月考核')
            headers.append(f'{m}月金额(元)')
        headers.append('合计(元)')
        headers.append('备注')
        col_count = len(headers)

        # 标题行
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        ws['A1'] = f"{unit_name} {year}年{quarter} {template.name}发放表"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = header_fill

        # 数据行
        row_num = 3
        emp_list = [emp_data[eid] for eid in employees if eid in emp_data]
        for idx, emp in enumerate(emp_list, 1):
            ws.cell(row=row_num, column=1, value=idx).alignment = center_align
            ws.cell(row=row_num, column=1).border = thin_border
            ws.cell(row=row_num, column=2, value=emp['name']).alignment = center_align
            ws.cell(row=row_num, column=2).border = thin_border
            col = 3
            for m in months:
                # 考核
                cell_a = ws.cell(row=row_num, column=col, value=emp['assessments'][f'{m}月'])
                cell_a.alignment = center_align
                cell_a.border = thin_border
                col += 1
                # 金额
                cell_amt = ws.cell(row=row_num, column=col, value=emp['amounts'][f'{m}月'])
                cell_amt.number_format = '#,##0.00'
                cell_amt.alignment = right_align
                cell_amt.border = thin_border
                col += 1
            # 合计
            cell_total = ws.cell(row=row_num, column=col, value=emp['total'])
            cell_total.number_format = '#,##0.00'
            cell_total.font = header_font
            cell_total.alignment = right_align
            cell_total.border = thin_border
            col += 1
            # 备注
            ws.cell(row=row_num, column=col, value=emp['remark']).alignment = center_align
            ws.cell(row=row_num, column=col).border = thin_border
            row_num += 1

        # 合计行
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        ws.cell(row=row_num, column=1, value='合计').font = header_font
        ws.cell(row=row_num, column=1).alignment = center_align
        ws.cell(row=row_num, column=1).border = thin_border
        ws.cell(row=row_num, column=2).border = thin_border
        col = 3
        for m in months:
            ws.cell(row=row_num, column=col, value='——').alignment = center_align
            ws.cell(row=row_num, column=col).border = thin_border
            col += 1
            total_m = sum(emp['amounts'][f'{m}月'] for emp in emp_list)
            cell = ws.cell(row=row_num, column=col, value=total_m)
            cell.number_format = '#,##0.00'
            cell.font = header_font
            cell.alignment = right_align
            cell.border = thin_border
            col += 1
        # 季度合计
        total_all = sum(emp['total'] for emp in emp_list)
        cell = ws.cell(row=row_num, column=col, value=total_all)
        cell.number_format = '#,##0.00'
        cell.font = header_font
        cell.alignment = right_align
        cell.border = thin_border
        col += 1
        ws.cell(row=row_num, column=col, value='——').alignment = center_align
        ws.cell(row=row_num, column=col).border = thin_border

        # 列宽
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 10
        for c in range(3, col_count + 1):
            ws.column_dimensions[get_column_letter(c)].width = 14

        ws.freeze_panes = 'A3'

        # 发送文件
        export_dir = os.path.join(current_dir, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = f"{unit_name}_{year}_{quarter}_{template.name}_发放表.xlsx"
        filepath = os.path.join(export_dir, filename)
        wb.save(filepath)
        return send_file(filepath, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/export_template_excel')
    @login_required
    def export_template_excel():
        from models import PrintTemplate
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        # 1. 获取请求参数
        template_id = request.args.get('template_id', type=int)
        if not template_id:
            return "缺少模板ID", 400

        month = request.args.get('month')
        year = request.args.get('year', type=int)
        quarter = request.args.get('quarter')
        unit_id = request.args.get('unit_id', type=int)
        employee_ids = request.args.getlist('employee_ids')
        empty_mode = request.args.get('empty_mode', 'false').lower() == 'true'

        # 2. 获取模板对象
        template = db.session.get(PrintTemplate, template_id)
        if not template:
            return "模板不存在", 404

        config = json.loads(template.config) if template.config else {}
        columns = config.get('columns', [])
        column_display = config.get('column_display', {})
        show_total = config.get('show_total_row', True)
        title = config.get('title', '')
        template_type = template.template_type

        # 3. 获取预览数据
        preview_data = _get_preview_data(
            template_type=template_type,
            template_id=template_id,
            month=month,
            year=year,
            quarter=quarter,
            unit_id=unit_id,
            employee_ids=employee_ids,
            empty_mode=empty_mode
        )
        if preview_data.get('error'):
            return preview_data['error'], 400

        rows = preview_data.get('rows', [])
        meta = preview_data.get('meta', {})

        # 4. 构建表头（序号、姓名 + 自定义列）
        display_headers = ['序号', '姓名']
        for col in columns:
            display_headers.append(column_display.get(col, col))
        col_count = len(display_headers)

        # 5. 创建工作簿及样式
        wb = Workbook()
        ws = wb.active
        ws.title = meta.get('template_name', '导出表')[:31]

        title_font = Font(name='微软雅黑', bold=True, size=14)
        header_font = Font(name='微软雅黑', bold=True, size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # 6. 标题行
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        unit_name = meta.get('unit_name', '')
        year_str = str(year or (meta.get('year', '')))
        month_str = month.split('-')[1] if month else ''
        quarter_display = {'Q1': '第一季度', 'Q2': '第二季度', 'Q3': '第三季度', 'Q4': '第四季度'}.get(quarter,
                                                                                                       quarter or '')
        title_final = title.replace('{unit_name}', unit_name) \
            .replace('{year}', year_str) \
            .replace('{month}', month_str) \
            .replace('{quarter}', quarter_display) \
            .replace('{template_name}', meta.get('template_name', ''))
        ws['A1'] = title_final
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align

        # 7. 表头行（第2行）
        for col_idx, header in enumerate(display_headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = header_fill

        # 8. 数据行
        for idx, row_data in enumerate(rows, 1):
            row_num = idx + 2
            # 序号
            ws.cell(row=row_num, column=1, value=idx).alignment = center_align
            ws.cell(row=row_num, column=1).border = thin_border
            # 姓名
            emp_name = row_data.get('employee_name') or row_data.get('姓名') or ''
            ws.cell(row=row_num, column=2, value=emp_name).alignment = center_align
            ws.cell(row=row_num, column=2).border = thin_border

            # 自定义列
            for col_idx, col_name in enumerate(columns, 3):
                # 空表模式：数据单元格留空
                if empty_mode:
                    val = ''
                else:
                    val = row_data.get(col_name, '')
                is_number = is_number_column(col_name, template_type)
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.alignment = right_align if is_number else center_align
                cell.border = thin_border
                if is_number and not empty_mode and val != '':
                    try:
                        num = float(val)
                        cell.value = num
                        cell.number_format = '#,##0.00'
                    except:
                        pass

        # 9. 合计行
        if show_total:
            total_row_num = len(rows) + 3
            # 合并前两列
            ws.merge_cells(start_row=total_row_num, start_column=1, end_row=total_row_num, end_column=2)
            ws.cell(row=total_row_num, column=1, value='合计').font = header_font
            ws.cell(row=total_row_num, column=1).alignment = center_align
            ws.cell(row=total_row_num, column=1).border = thin_border
            ws.cell(row=total_row_num, column=2).border = thin_border

            col_index = 3
            for col_name in columns:
                cell = ws.cell(row=total_row_num, column=col_index)
                cell.border = thin_border

                if empty_mode:
                    cell.value = ''
                    cell.alignment = center_align
                else:
                    # 新增：备注、身份证号等非数值列直接显示 ——
                    if col_name in ['备注', '身份证号']:
                        cell.value = '——'
                        cell.alignment = center_align
                    else:
                        # 原有数值求和逻辑
                        total = 0.0
                        all_numeric = True
                        for r in rows:
                            val = r.get(col_name, '')
                            if val == '' or val is None:
                                continue
                            try:
                                total += float(val)
                            except (ValueError, TypeError):
                                all_numeric = False
                                break
                        if all_numeric and rows:
                            cell.value = total
                            cell.number_format = '#,##0.00'
                            cell.font = header_font
                            cell.alignment = right_align
                        else:
                            cell.value = '——'
                            cell.alignment = center_align
                col_index += 1

        # 10. 列宽和冻结窗格
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 12
        for col in range(3, col_count + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.freeze_panes = 'A3'

        # 11. 导出文件
        export_dir = os.path.join(current_dir, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        suffix = "_空表" if empty_mode else ""
        filename = f"{template.name}{suffix}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        filepath = os.path.join(export_dir, filename)
        wb.save(filepath)

        return send_file(filepath, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/special_annual')
    @login_required
    def special_annual():
        return render_template('special_annual.html')

    # ==================== 新增：批量计算增强模块（替代原 advanced_calculator） ====================
    @app.route('/batch_calc_enhanced')
    @login_required
    @admin_required
    def batch_calc_enhanced():
        """增强版批量计算表单页面"""
        return render_template('batch_calc_enhanced.html')

    @app.route('/api/batch_base_total', methods=['POST'])
    @login_required
    @admin_required
    def batch_base_total():
        """批量获取员工工资项基数合计（支持单月/全年/区间，个人/单位/全公司聚合）"""
        data = request.get_json()
        emp_ids = data.get('employee_ids', [])
        items = data.get('salary_items', [])
        time_range = data.get('time_range', {})
        agg_type = data.get('aggregate', 'employee')

        if not emp_ids or not items:
            return jsonify({'error': '缺少员工ID或工资项'}), 400

        # 处理时间范围
        months = []
        tr_type = time_range.get('type')
        if tr_type == 'single':
            month = time_range.get('month')
            if month:
                months = [month]
        elif tr_type == 'year':
            year = time_range.get('year')
            if year:
                for m in range(1, 13):
                    months.append(f"{year}-{str(m).zfill(2)}")
        elif tr_type == 'range':
            start = time_range.get('start')
            end = time_range.get('end')
            if start and end:
                cur = datetime.strptime(start, '%Y-%m')
                end_dt = datetime.strptime(end, '%Y-%m')
                while cur <= end_dt:
                    months.append(cur.strftime('%Y-%m'))
                    cur += relativedelta(months=1)

        if not months:
            return jsonify({'error': '无效的时间范围'}), 400

        # 根据聚合类型计算
        result = {}
        if agg_type == 'employee':
            for eid in emp_ids:
                total = 0.0
                for m in months:
                    rec = SalaryRecord.query.filter_by(employee_id=eid, month=m).first()
                    if rec and rec.details:
                        details = rec.details if isinstance(rec.details, dict) else json.loads(rec.details)
                        for item in items:
                            total += float(details.get(item, 0))
                result[eid] = round(total, 2)
        elif agg_type == 'unit':
            # 获取员工单位映射
            emp_units = {e.id: e.unit_id for e in Employee.query.filter(Employee.id.in_(emp_ids)).all()}
            units = set(emp_units.values())
            unit_totals = {u: 0.0 for u in units}
            # 先计算每个员工的个人合计，再累加到单位
            for eid in emp_ids:
                total = 0.0
                for m in months:
                    rec = SalaryRecord.query.filter_by(employee_id=eid, month=m).first()
                    if rec and rec.details:
                        details = rec.details if isinstance(rec.details, dict) else json.loads(rec.details)
                        for item in items:
                            total += float(details.get(item, 0))
                unit_totals[emp_units[eid]] += total
            for eid in emp_ids:
                result[eid] = round(unit_totals[emp_units[eid]], 2)
        elif agg_type == 'company':
            # 全公司合计（所有在职员工）
            all_emp_ids = [e.id for e in Employee.query.filter_by(active_for_payroll=True).all()]
            company_total = 0.0
            for eid in all_emp_ids:
                for m in months:
                    rec = SalaryRecord.query.filter_by(employee_id=eid, month=m).first()
                    if rec and rec.details:
                        details = rec.details if isinstance(rec.details, dict) else json.loads(rec.details)
                        for item in items:
                            company_total += float(details.get(item, 0))
            for eid in emp_ids:
                result[eid] = round(company_total, 2)
        else:
            return jsonify({'error': '不支持的聚合方式'}), 400

        return jsonify(result)

    @app.route('/api/employee_coefficient', methods=['POST'])
    @login_required
    @admin_required
    def employee_coefficient():
        """获取员工系数（支持固定值、员工字段、其他工资项、特殊事项）"""
        data = request.get_json()
        emp_ids = data.get('employee_ids', [])
        coeff_config = data.get('config', {})
        if not emp_ids or not coeff_config:
            return jsonify({'error': '参数不足'}), 400

        result = {}
        coeff_type = coeff_config.get('type')
        for eid in emp_ids:
            emp = db.session.get(Employee, eid)
            if not emp:
                result[eid] = 1.0
                continue
            if coeff_type == 'fixed':
                val = float(coeff_config.get('value', 1.0))
            elif coeff_type == 'field':
                field = coeff_config.get('field')
                if field == 'is_veteran':
                    val = 1.2 if emp.is_veteran else 1.0
                elif field == 'position_level':
                    level_map = {'初级': 0.8, '中级': 1.0, '高级': 1.2}
                    val = level_map.get(emp.position_level, 1.0)
                elif field == 'salary_level':
                    val = 1.0  # 可扩展
                elif field.startswith('extra_fields.'):
                    extra = json.loads(emp.extra_fields or '{}')
                    key = field.split('.')[1]
                    val = float(extra.get(key, 1.0))
                else:
                    val = 1.0
            elif coeff_type == 'salary':
                item_name = coeff_config.get('item_name')
                month = coeff_config.get('month')
                if not month:
                    val = 1.0
                else:
                    rec = SalaryRecord.query.filter_by(employee_id=eid, month=month).first()
                    val = float(rec.details.get(item_name, 0)) if rec and rec.details else 0.0
            elif coeff_type == 'special':
                tpl_id = coeff_config.get('template_id')
                period = coeff_config.get('period')  # 'year' or 'month'
                year = coeff_config.get('year')
                month = coeff_config.get('month') if period == 'month' else None
                grant = EmployeeSpecialGrant.query.filter_by(
                    employee_id=eid, template_id=tpl_id, year=year, month=month
                ).first()
                val = grant.amount if grant else 1.0
            else:
                val = 1.0
            result[eid] = round(val, 4)
        return jsonify(result)

    @app.route('/api/batch_update_salary', methods=['POST'])
    @login_required
    @admin_required
    def batch_update_salary():
        """批量更新工资项（覆盖或累加）"""
        data = request.get_json()
        target_item = data.get('target_item')
        target_month = data.get('target_month')
        write_mode = data.get('write_mode')  # 'overwrite' or 'append'
        updates = data.get('updates', [])  # [{employee_id, amount, remark}]

        if not target_item or not target_month or not updates:
            return jsonify({'success': False, 'error': '缺少必要参数'}), 400

        count = 0
        for upd in updates:
            emp_id = upd['employee_id']
            amount = float(upd['amount'])
            remark = upd.get('remark', '')
            record = SalaryRecord.query.filter_by(employee_id=emp_id, month=target_month).first()
            if not record:
                record = SalaryRecord(employee_id=emp_id, month=target_month, details={}, total=0)
                db.session.add(record)
            details = record.details
            if isinstance(details, str):
                details = json.loads(details)
            old_val = details.get(target_item, 0)
            if write_mode == 'overwrite':
                details[target_item] = amount
            else:
                details[target_item] = old_val + amount
            record.details = details
            if remark:
                record.remark = remark
            count += 1

        try:
            db.session.commit()
            return jsonify({'success': True, 'updated': count})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/api/auxiliary_forms', methods=['GET'])
    @login_required
    @admin_required
    def list_auxiliary_forms():
        forms = AuxiliaryForm.query.filter_by(created_by=current_user.id).order_by(
            AuxiliaryForm.updated_at.desc()).all()
        return jsonify([{
            'id': f.id,
            'name': f.name,
            'description': f.description,
            'updated_at': f.updated_at.strftime('%Y-%m-%d %H:%M') if f.updated_at else f.created_at.strftime(
                '%Y-%m-%d %H:%M')
        } for f in forms])

    # 保存或更新辅助表单（POST）
    @app.route('/api/auxiliary_forms', methods=['POST'])
    @login_required
    @admin_required
    def save_auxiliary_form():
        data = request.get_json()
        form_id = data.get('id')
        name = data.get('name')
        description = data.get('description', '')
        config = data.get('config')
        if not name or not config:
            return jsonify({'success': False, 'error': '名称和配置不能为空'}), 400
        if form_id:
            form = db.session.get(AuxiliaryForm, form_id)
            if not form or form.created_by != current_user.id:
                return jsonify({'success': False, 'error': '表单不存在或无权限'}), 404
            form.name = name
            form.description = description
            form.config = json.dumps(config, ensure_ascii=False)
            form.updated_at = datetime.now()
        else:
            form = AuxiliaryForm(
                name=name,
                description=description,
                config=json.dumps(config, ensure_ascii=False),
                created_by=current_user.id
            )
            db.session.add(form)
        db.session.commit()
        return jsonify({'success': True, 'id': form.id})


    @app.route('/api/auxiliary_forms/<int:form_id>', methods=['GET'])
    @login_required
    @admin_required
    def get_auxiliary_form(form_id):
        """获取单个表单的完整配置"""
        form = AuxiliaryForm.query.get_or_404(form_id)
        if form.created_by != current_user.id:
            return jsonify({'error': '无权限'}), 403
        return jsonify({
            'id': form.id,
            'name': form.name,
            'description': form.description,
            'config': json.loads(form.config)
        })

    # ==================== 批量计算快照 API ====================
    @app.route('/api/batch_snapshots', methods=['GET'])
    @login_required
    def list_batch_snapshots():
        """获取快照列表（可筛选年份、模板、单位）"""
        year = request.args.get('year', type=int)
        template_id = request.args.get('template_id', type=int)
        unit_id = request.args.get('unit_id', type=int)  # 新增：单位ID过滤

        query = BatchCalcSnapshot.query
        if year:
            query = query.filter_by(year=year)
        if template_id:
            query = query.filter_by(template_id=template_id)

        # 非管理员权限过滤（原有）
        if not current_user.is_admin:
            query = query.filter(
                (BatchCalcSnapshot.created_by == current_user.id) | (BatchCalcSnapshot.is_public == True)
            )

        snapshots = query.all()

        # 如果传入了 unit_id，则进一步过滤（解析 result_data 中的 filters.unit_id）
        if unit_id:
            filtered = []
            for s in snapshots:
                try:
                    data = json.loads(s.result_data)
                    filters = data.get('filters', {})
                    if filters.get('unit_id') == str(unit_id):  # 注意单位ID可能是字符串
                        filtered.append(s)
                except:
                    # 解析失败则跳过（旧快照可能没有 filters）
                    pass
            snapshots = filtered

        return jsonify([{
            'id': s.id,
            'name': s.name,
            'year': s.year,
            'created_at': s.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'template_id': s.template_id
        } for s in snapshots])

    @app.route('/api/batch_snapshots', methods=['POST'])
    @login_required
    def create_batch_snapshot():
        """保存当前计算结果为快照"""
        data = request.get_json()
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'error': '快照名称不能为空'}), 400
        description = data.get('description', '')
        is_public = data.get('is_public', False)
        template_id = data.get('template_id')
        year = data.get('year')
        result_data = data.get('result_data')

        if not template_id or not year or not result_data:
            return jsonify({'error': '缺少必要参数 (template_id, year, result_data)'}), 400

        snapshot = BatchCalcSnapshot(
            name=name,
            description=description,
            year=year,
            template_id=template_id,
            result_data=result_data,
            created_by=current_user.id,
            is_public=is_public
        )
        db.session.add(snapshot)
        db.session.commit()
        return jsonify({'id': snapshot.id, 'message': '快照保存成功'})

    @app.route('/api/batch_snapshots/<int:sid>', methods=['GET'])
    @login_required
    def get_batch_snapshot(sid):
        """获取单个快照详情"""
        snap = BatchCalcSnapshot.query.get_or_404(sid)
        if not current_user.is_admin and snap.created_by != current_user.id and not snap.is_public:
            abort(403)
        return jsonify({
            'id': snap.id,
            'name': snap.name,
            'description': snap.description,
            'year': snap.year,
            'template_id': snap.template_id,
            'result_data': snap.result_data,
            'is_public': snap.is_public,
            'created_at': snap.created_at.strftime('%Y-%m-%d %H:%M:%S')
        })

    @app.route('/api/batch_snapshots/<int:sid>', methods=['DELETE'])
    @login_required
    def delete_batch_snapshot(sid):
        """删除快照"""
        snap = BatchCalcSnapshot.query.get_or_404(sid)
        if not current_user.is_admin and snap.created_by != current_user.id:
            abort(403)
        db.session.delete(snap)
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/batch_import_manual_values', methods=['POST'])
    @login_required
    def import_manual_values():
        """从 Excel 导入指定手工字段的值"""
        if 'file' not in request.files:
            return jsonify({'error': '未上传文件'}), 400
        file = request.files['file']
        field_name = request.form.get('field_name')
        if not field_name:
            return jsonify({'error': '缺少字段名'}), 400

        try:
            df = pd.read_excel(file)
        except Exception as e:
            return jsonify({'error': f'读取 Excel 失败: {str(e)}'}), 400

        # 检查必要列
        if '员工姓名' not in df.columns and '员工ID' not in df.columns:
            return jsonify({'error': 'Excel 必须包含“员工姓名”或“员工ID”列'}), 400

        # 确定值列：取除了员工标识列之外的第一列（可以优化，这里简单取第一个数值列）
        value_col = None
        for col in df.columns:
            if col not in ['员工姓名', '员工ID']:
                value_col = col
                break
        if not value_col:
            return jsonify({'error': '未找到值列'}), 400

        result = []
        for idx, row in df.iterrows():
            emp = None
            if '员工ID' in df.columns and pd.notna(row['员工ID']):
                emp = db.session.get(Employee, int(row['员工ID']))
            else:
                name = str(row['员工姓名']).strip()
                emp = Employee.query.filter_by(name=name).first()
            if not emp:
                continue
            val = row[value_col]
            try:
                val = float(val)
            except:
                continue
            result.append({'employee_id': emp.id, 'value': val})

        return jsonify({'success': True, 'data': result})

    @app.route('/api/batch_evaluate_expressions', methods=['POST'])
    @login_required
    def batch_evaluate_expressions():
        data = request.get_json()
        emp_ids = data.get('employees', [])
        expressions = data.get('expressions', [])  # [{name, expr}]
        base_totals = data.get('base_totals', {})
        year = data.get('year')
        if not emp_ids or not expressions:
            return jsonify([])
        employees = Employee.query.filter(Employee.id.in_(emp_ids)).all()
        emp_map = {e.id: e for e in employees}
        result = []
        for emp_id in emp_ids:
            emp = emp_map.get(emp_id)
            if not emp:
                continue
            context = {
                'base_total': base_totals.get(str(emp_id), 0),
                'year': year,
                '姓名': emp.name,
                '人员身份': emp.employee_identity,
                '岗位级别': emp.position_level,
                '人员类型': emp.employee_type,
                '性别': emp.gender,
                '是否退役军人': 1 if emp.is_veteran else 0,
                'employee_id': emp_id
            }
            row_res = {'employee_id': emp_id, 'results': {}}
            for exp in expressions:
                if exp['expr']:
                    try:
                        # 调用 evaluate_formula_with_context 支持本年合计
                        val = evaluate_formula_with_context(exp['expr'], emp_id, year, context, None)
                    except:
                        val = 0
                    row_res['results'][exp['name']] = val
            result.append(row_res)
        return jsonify(result)
    @app.route('/api/batch_templates', methods=['GET'])
    @login_required
    @admin_required
    def list_batch_templates():
        templates = BatchCalcTemplate.query.order_by(BatchCalcTemplate.year.desc()).all()
        return jsonify([{
            'id': t.id,
            'name': t.name,
            'year': t.year
        } for t in templates])

    @app.route('/api/batch_templates', methods=['POST'])
    @login_required
    @admin_required
    def save_batch_template():
        data = request.get_json()
        tpl_id = data.get('id')
        name = data['name']
        year = data['year']
        config = json.dumps(data['config'], ensure_ascii=False)
        if tpl_id:
            tpl = db.session.get(BatchCalcTemplate, tpl_id)
            if not tpl:
                return jsonify({'error': '模板不存在'}), 404
            tpl.name = name
            tpl.year = year
            tpl.config = config
        else:
            tpl = BatchCalcTemplate(name=name, year=year, config=config, created_by=current_user.id)
            db.session.add(tpl)
        db.session.commit()
        return jsonify({'success': True, 'id': tpl.id, 'message': '保存成功'})

    @app.route('/api/batch_templates/<int:tpl_id>', methods=['GET'])
    @login_required
    @admin_required
    def get_batch_template(tpl_id):
        tpl = BatchCalcTemplate.query.get_or_404(tpl_id)
        return jsonify({
            'id': tpl.id,
            'name': tpl.name,
            'year': tpl.year,
            'config': json.loads(tpl.config)
        })

    @app.route('/api/batch_templates/<int:tpl_id>/copy', methods=['POST'])
    @login_required
    @admin_required
    def copy_batch_template(tpl_id):
        src = BatchCalcTemplate.query.get_or_404(tpl_id)
        data = request.get_json()
        target_year = data.get('targetYear', src.year + 1)
        new_name = f"{src.name} (复制至{target_year})"
        new_config = json.loads(src.config)  # 深拷贝
        new_tpl = BatchCalcTemplate(
            name=new_name,
            year=target_year,
            config=json.dumps(new_config, ensure_ascii=False),
            created_by=current_user.id
        )
        db.session.add(new_tpl)
        db.session.commit()
        return jsonify({'success': True, 'new_id': new_tpl.id, 'message': f'已复制到 {target_year} 年'})

    @app.route('/api/batch_templates/<int:tpl_id>', methods=['DELETE'])
    @login_required
    @admin_required
    def delete_batch_template(tpl_id):
        tpl = BatchCalcTemplate.query.get_or_404(tpl_id)
        db.session.delete(tpl)
        db.session.commit()
        return jsonify({'success': True, 'message': '模板已删除'})
    # 计算员工基数总和（工资项+特殊事项）
    @app.route('/api/batch_base_total_with_special', methods=['POST'])
    @login_required
    def batch_base_total_with_special():
        data = request.get_json()
        emp_id = data['employee_id']
        base_items = data.get('base_items', [])  # 工资项名称列表
        special_tpl_ids = data.get('special_templates', [])  # 特殊事项模板ID列表
        time_range = data.get('time_range', {})

        total = 0.0

        # ---------- 1. 解析时间范围，生成月份列表和年份列表 ----------
        tr_type = time_range.get('type')
        months = []  # 存放具体月份 "YYYY-MM"
        years = set()  # 存放涉及的年份（用于特殊事项年度查询）

        if tr_type == 'single':
            month = time_range.get('singleMonth')
            if month and re.match(r'\d{4}-\d{2}', month):
                months = [month]
                years.add(int(month[:4]))
        elif tr_type == 'year':
            year = time_range.get('yearTotal')
            if year:
                year = int(year)
                years.add(year)
                # 生成该年所有月份
                for m in range(1, 13):
                    months.append(f"{year}-{str(m).zfill(2)}")
        elif tr_type == 'range':
            start = time_range.get('rangeStart')
            end = time_range.get('rangeEnd')
            if start and end:
                cur = datetime.strptime(start, '%Y-%m')
                end_dt = datetime.strptime(end, '%Y-%m')
                while cur <= end_dt:
                    month_str = cur.strftime('%Y-%m')
                    months.append(month_str)
                    years.add(cur.year)
                    cur += relativedelta(months=1)
        else:
            # 无时间范围，直接返回0
            return jsonify({'total': 0.0})

        # ---------- 2. 累加工资项 ----------
        if base_items:
            for month in months:
                rec = SalaryRecord.query.filter_by(employee_id=emp_id, month=month).first()
                if rec and rec.details:
                    details = rec.details if isinstance(rec.details, dict) else json.loads(rec.details)
                    for item in base_items:
                        total += float(details.get(item, 0))

        # ---------- 3. 累加特殊事项 ----------
        if special_tpl_ids:
            # 获取模板信息（频率）
            from models import SpecialItemTemplate
            templates = {t.id: t for t in
                         SpecialItemTemplate.query.filter(SpecialItemTemplate.id.in_(special_tpl_ids)).all()}

            for tpl_id in special_tpl_ids:
                tpl = templates.get(tpl_id)
                if not tpl:
                    continue

                # 根据时间范围类型分别处理
                if tr_type == 'single':
                    # 单月：精确匹配年月
                    year = int(months[0][:4])
                    month = int(months[0][5:7])
                    grant = EmployeeSpecialGrant.query.filter_by(
                        employee_id=emp_id, template_id=tpl_id, year=year, month=month
                    ).first()
                    if grant:
                        total += grant.amount

                elif tr_type == 'year':
                    year = int(time_range.get('yearTotal'))
                    if tpl.frequency == 'monthly':
                        # 月度事项：累加该年所有月份
                        grants = EmployeeSpecialGrant.query.filter(
                            EmployeeSpecialGrant.employee_id == emp_id,
                            EmployeeSpecialGrant.template_id == tpl_id,
                            EmployeeSpecialGrant.year == year,
                            EmployeeSpecialGrant.month.isnot(None)
                        ).all()
                        total += sum(g.amount for g in grants)
                    else:
                        # 年度事项：取 month IS NULL 的记录
                        grant = EmployeeSpecialGrant.query.filter_by(
                            employee_id=emp_id, template_id=tpl_id, year=year, month=None
                        ).first()
                        if grant:
                            total += grant.amount

                elif tr_type == 'range':
                    # 区间：可能跨年，需要根据模板频率累加
                    if tpl.frequency == 'monthly':
                        # 月度事项：逐月查询（利用 months 列表）
                        for month_str in months:
                            year = int(month_str[:4])
                            month = int(month_str[5:7])
                            grant = EmployeeSpecialGrant.query.filter_by(
                                employee_id=emp_id, template_id=tpl_id, year=year, month=month
                            ).first()
                            if grant:
                                total += grant.amount
                    else:
                        # 年度事项：只要年份在 years 集合内，就累加该年记录
                        for year in years:
                            grant = EmployeeSpecialGrant.query.filter_by(
                                employee_id=emp_id, template_id=tpl_id, year=year, month=None
                            ).first()
                            if grant:
                                total += grant.amount

        return jsonify({'total': round(total, 2)})

    # 表达式求值（支持公式计算，变量传入）
    @app.route('/api/evaluate_expression', methods=['POST'])
    @login_required
    def evaluate_expression():
        data = request.get_json()
        expr = data['expression']
        variables = data.get('variables', {})
        employee_id = data.get('employee_id')
        year = data.get('year')
        month = data.get('month')  # 前端传入的基准月份（用于单月取值）
        # ========== 新增：单独处理身份证号 ==========
        if expr.strip() == '身份证号':
            id_card = variables.get('身份证号')
            if not id_card and employee_id:
                emp = db.session.get(Employee, employee_id)
                if emp:
                    id_card = emp.id_card or ''
            return jsonify({'result': id_card or ''})
        # ========== 支持 row_字段名 语法（将其转换为普通字段名） ==========
        import re
        expr = re.sub(r'\brow_([a-zA-Z\u4e00-\u9fa5_][a-zA-Z0-9\u4e00-\u9fa5_]*)\b', r'\1', expr)

        # ========== 新增：员工属性注入 ==========
        if employee_id:
            employee = db.session.get(Employee, employee_id)
            if employee:
                # 中文变量名（推荐使用）
                variables['姓名'] = employee.name or ''  # ← 添加这一行
                variables['人员身份'] = employee.employee_identity or ''
                variables['岗位级别'] = employee.position_level or ''
                variables['人员类型'] = employee.employee_type or ''
                variables['性别'] = employee.gender or ''
                variables['是否退役军人'] = 1 if employee.is_veteran else 0
                variables['身份证号'] = employee.id_card or ''
                # 保留英文变量名以兼容旧公式
                variables['name'] = employee.name or ''  # ← 可选，供英文公式使用
                variables['employee_identity'] = employee.employee_identity or ''
                variables['position_level'] = employee.position_level or ''
                variables['employee_type'] = employee.employee_type or ''
                variables['gender'] = employee.gender or ''
                variables['is_veteran'] = 1 if employee.is_veteran else 0
        if employee_id and year:
            # 正则：匹配 special_模板名(参数) 或 special_模板名 (无括号)
            pattern_with_args = r'special_([a-zA-Z\u4e00-\u9fa5_][a-zA-Z0-9\u4e00-\u9fa5_]*)\(([^)]*)\)'
            pattern_no_args = r'(?<!_)special_([a-zA-Z\u4e00-\u9fa5_][a-zA-Z0-9\u4e00-\u9fa5_]*)(?!\()'

            # 处理带参数的特殊事项
            def replace_with_args(match):
                tpl_name = match.group(1)
                args_str = match.group(2).strip()
                parts = [p.strip() for p in args_str.split(',') if p.strip()]
                tpl = SpecialItemTemplate.query.filter_by(name=tpl_name).first()
                if not tpl:
                    return '0'
                if len(parts) == 1:
                    p = parts[0]
                    if p.isdigit():
                        if len(p) == 4:  # 年份
                            target_year = int(p)
                            query = EmployeeSpecialGrant.query.filter_by(
                                employee_id=employee_id,
                                template_id=tpl.id,
                                year=target_year
                            )
                            if tpl.frequency == 'monthly':
                                grants = query.all()
                                total = sum(g.amount for g in grants)
                            else:
                                grant = query.filter_by(month=None).first()
                                total = grant.amount if grant else 0
                            return str(total)
                        else:  # 月份数字（1-12）
                            target_year = year
                            target_month = int(p)
                            grant = EmployeeSpecialGrant.query.filter_by(
                                employee_id=employee_id,
                                template_id=tpl.id,
                                year=target_year,
                                month=target_month
                            ).first()
                            return str(grant.amount if grant else 0)
                elif len(parts) == 2:
                    y = int(parts[0])
                    m = int(parts[1])
                    grant = EmployeeSpecialGrant.query.filter_by(
                        employee_id=employee_id,
                        template_id=tpl.id,
                        year=y,
                        month=m
                    ).first()
                    return str(grant.amount if grant else 0)
                elif len(parts) == 3:
                    y = int(parts[0])
                    start_m = int(parts[1])
                    end_m = int(parts[2])
                    grants = EmployeeSpecialGrant.query.filter(
                        EmployeeSpecialGrant.employee_id == employee_id,
                        EmployeeSpecialGrant.template_id == tpl.id,
                        EmployeeSpecialGrant.year == y,
                        EmployeeSpecialGrant.month >= start_m,
                        EmployeeSpecialGrant.month <= end_m
                    ).all()
                    total = sum(g.amount for g in grants)
                    return str(total)
                else:
                    return '0'

            # 处理无参数的特殊事项
            def replace_no_args(match):
                tpl_name = match.group(1)
                tpl = SpecialItemTemplate.query.filter_by(name=tpl_name).first()
                if not tpl:
                    return '0'
                query = EmployeeSpecialGrant.query.filter_by(
                    employee_id=employee_id,
                    template_id=tpl.id,
                    year=year
                )
                if tpl.frequency == 'monthly' and month:
                    query = query.filter_by(month=month)
                else:
                    query = query.filter_by(month=None)
                grant = query.first()
                return str(grant.amount if grant else 0)

            expr = re.sub(pattern_with_args, replace_with_args, expr)
            expr = re.sub(pattern_no_args, replace_no_args, expr)

            # 调用增强公式计算（支持上年合计、指定年工资项合计等）
            result = evaluate_formula_with_context(
                expr, employee_id, year, variables,
                salary_item_names=None
            )
            return jsonify({'result': result})
        else:
            result = evaluate_formula(expr, variables)
            return jsonify({'result': result})
    @app.route('/api/auxiliary_forms/<int:form_id>', methods=['DELETE'])
    @login_required
    @admin_required
    def delete_auxiliary_form(form_id):
        """删除表单"""
        form = AuxiliaryForm.query.get_or_404(form_id)
        if form.created_by != current_user.id:
            return jsonify({'error': '无权限'}), 403
        db.session.delete(form)
        db.session.commit()
        return jsonify({'success': True})
    return app  # 这一行绝对不能少


if __name__ == '__main__':
    app = create_app()
    app.run(debug=True, host='0.0.0.0', port=5000)